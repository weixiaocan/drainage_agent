from __future__ import annotations

import json
from copy import copy
from pathlib import Path

import pytest
from openpyxl import load_workbook

from analysis.baselines import (
    BaselinePreconditionError,
    FilterBaselineService,
    FilterRequest,
)
from analysis.runs import AnalysisRequest, AnalysisRunner
from agent.tools.filter_baselines import (
    confirm_filter_baseline,
    run_filter_analysis,
)
from analysis.io.standard import STANDARD_FLOW_COLUMNS, STANDARD_FLOW_UNITS
from web.projects import ProjectRepository


def write_standard_flow(batch_workspace: Path, *, point_id: str = "W1") -> None:
    standard = batch_workspace / "standard"
    standard.mkdir(parents=True, exist_ok=True)
    rows = [
        ",".join(STANDARD_FLOW_COLUMNS),
        *[
            f"2026-03-{day:02d}T00:00:00,D1,{point_id},{day}.0,1.0,0.5"
            for day in range(1, 6)
        ],
    ]
    (standard / "flow.csv").write_text("\n".join(rows) + "\n", encoding="utf-8")
    (standard / "manifest.json").write_text(
        json.dumps(
            {
                "contract_version": 1,
                "kind": "standard_flow",
                "columns": STANDARD_FLOW_COLUMNS,
                "units": STANDARD_FLOW_UNITS,
                "source_import_id": "import-1",
                "source_sha256": "source-sha",
                "source_encoding": "utf-8",
                "mapping": {},
                "source_units": {},
                "file": "flow.csv",
            }
        ),
        encoding="utf-8",
    )


def test_filter_uses_confirmed_standard_v1_and_waits_for_confirmation(
    tmp_path: Path,
) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("筛选项目")
    batch = projects.create_batch(project.id, "筛选批次")
    workspace = projects.batch_workspace(project.id, batch.id)
    write_standard_flow(workspace)
    raw = workspace / "inputs" / "raw-1"
    raw.mkdir(parents=True)
    (raw / "flow.csv").write_text(
        "timestamp,point_id,flow_lps\n2026-03-02,W9,999\n",
        encoding="utf-8",
    )

    result = FilterBaselineService(database, files_root).run_filter(
        FilterRequest(
            project_id=project.id,
            batch_id=batch.id,
            expected_rows_per_day=1,
        )
    )

    assert result.status == "awaiting_confirmation"
    assert result.summary == {
        "point_count": 1,
        "selected_point_days": 3,
        "selected_days": {"W1": ["2026-03-02", "2026-03-03", "2026-03-04"]},
        "exclusion_reasons": {"W1": ["无明显异常"]},
    }
    assert result.identity["project_id"] == project.id
    assert result.identity["batch_id"] == batch.id
    assert result.identity["standard_input"]["contract_version"] == 1
    assert result.identity["parameters"]["expected_rows_per_day"] == 1
    assert result.artifact.endswith("/filter_result.xlsx")
    assert (workspace / result.artifact).is_file()
    assert FilterBaselineService(database, files_root).current_baseline(
        project.id, batch.id
    ) is None


def test_filter_never_falls_back_to_raw_input(tmp_path: Path) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("无标准数据项目")
    batch = projects.create_batch(project.id, "无标准数据批次")
    raw = projects.batch_workspace(project.id, batch.id) / "inputs" / "raw-1"
    raw.mkdir(parents=True)
    (raw / "flow.csv").write_text(
        "timestamp,point_id,flow_lps\n2026-03-02,W9,999\n",
        encoding="utf-8",
    )

    with pytest.raises(BaselinePreconditionError, match="标准数据"):
        FilterBaselineService(database, files_root).run_filter(
            FilterRequest(project.id, batch.id)
        )


def test_uploaded_revision_is_validated_and_confirmed_as_immutable_baseline(
    tmp_path: Path,
) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("人工修改项目")
    batch = projects.create_batch(project.id, "人工修改批次")
    workspace = projects.batch_workspace(project.id, batch.id)
    write_standard_flow(workspace)
    service = FilterBaselineService(database, files_root)
    generated = service.run_filter(
        FilterRequest(project.id, batch.id, expected_rows_per_day=1)
    )
    edited = tmp_path / "edited.xlsx"
    edited.write_bytes((workspace / generated.artifact).read_bytes())
    workbook = load_workbook(edited)
    sheet = workbook["筛选结果"]
    sheet.cell(row=3, column=3).fill = copy(sheet.cell(row=3, column=2).fill)
    workbook.save(edited)

    revision = service.upload_revision(
        project.id,
        batch.id,
        generated.filter_id,
        "edited.xlsx",
        edited.read_bytes(),
    )
    baseline = service.confirm(project.id, batch.id, revision.filter_id)

    assert revision.version == 2
    assert revision.status == "awaiting_confirmation"
    assert revision.identity["file_sha256"] != generated.identity["file_sha256"]
    assert baseline.identity == {
        "kind": "confirmed_filter",
        "identity": baseline.identity["identity"],
        "project_id": project.id,
        "batch_id": batch.id,
        "standard_input": generated.identity["standard_input"],
        "parameters": generated.identity["parameters"],
        "file_sha256": revision.identity["file_sha256"],
    }
    assert service.current_baseline(project.id, batch.id) == baseline
    baseline_path = workspace / baseline.artifact
    assert baseline_path.is_file()
    assert baseline_path.read_bytes() == edited.read_bytes()
    assert baseline.artifact != revision.artifact
    selected_flow = service.load_flow(project.id, batch.id)
    assert set(selected_flow["timestamp"].dt.strftime("%Y-%m-%d")) == {
        "2026-03-03",
        "2026-03-04",
    }


def test_invalid_revision_and_stale_confirmation_are_rejected(
    tmp_path: Path,
) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("校验项目")
    batch = projects.create_batch(project.id, "校验批次")
    workspace = projects.batch_workspace(project.id, batch.id)
    write_standard_flow(workspace)
    service = FilterBaselineService(database, files_root)
    generated = service.run_filter(
        FilterRequest(project.id, batch.id, expected_rows_per_day=1)
    )

    with pytest.raises(ValueError, match="xlsx"):
        service.upload_revision(
            project.id, batch.id, generated.filter_id, "bad.csv", b"bad"
        )
    with pytest.raises(ValueError, match="结构"):
        service.upload_revision(
            project.id, batch.id, generated.filter_id, "bad.xlsx", b"bad"
        )

    with (workspace / "standard" / "flow.csv").open("a", encoding="utf-8") as flow:
        flow.write("2026-03-06T00:00:00,D1,W1,6.0,1.0,0.5\n")
    with pytest.raises(BaselinePreconditionError, match="已过期"):
        service.confirm(project.id, batch.id, generated.filter_id)


def test_changed_parameters_or_baseline_file_invalidate_current_without_overwrite(
    tmp_path: Path,
) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("版本项目")
    batch = projects.create_batch(project.id, "版本批次")
    workspace = projects.batch_workspace(project.id, batch.id)
    write_standard_flow(workspace)
    service = FilterBaselineService(database, files_root)
    first_filter = service.run_filter(
        FilterRequest(project.id, batch.id, expected_rows_per_day=1)
    )
    first = service.confirm(project.id, batch.id, first_filter.filter_id)
    first_path = workspace / first.artifact
    first_bytes = first_path.read_bytes()

    second_filter = service.run_filter(
        FilterRequest(
            project.id,
            batch.id,
            expected_rows_per_day=1,
            mean_lower_ratio=0.25,
        )
    )

    assert service.current_baseline(project.id, batch.id) is None
    with pytest.raises(BaselinePreconditionError, match="已过期"):
        service.confirm(project.id, batch.id, first_filter.filter_id)
    assert first_path.read_bytes() == first_bytes
    second = service.confirm(project.id, batch.id, second_filter.filter_id)
    assert second.version == 2
    assert first_path.is_file()
    (workspace / second.artifact).write_bytes(b"changed outside the service")
    assert service.current_baseline(project.id, batch.id) is None
    with pytest.raises(BaselinePreconditionError, match="当前已确认分析基线"):
        service.load_flow(project.id, batch.id)


def test_cross_project_filter_access_is_rejected(tmp_path: Path) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    north = projects.create("北区")
    south = projects.create("南区")
    north_batch = projects.create_batch(north.id, "北区批次")
    south_batch = projects.create_batch(south.id, "南区批次")
    write_standard_flow(projects.batch_workspace(north.id, north_batch.id))
    write_standard_flow(projects.batch_workspace(south.id, south_batch.id))
    service = FilterBaselineService(database, files_root)
    north_filter = service.run_filter(
        FilterRequest(north.id, north_batch.id, expected_rows_per_day=1)
    )

    with pytest.raises(LookupError, match="筛选结果不存在"):
        service.confirm(south.id, south_batch.id, north_filter.filter_id)


def test_web_filter_download_upload_and_explicit_confirmation_workflow(
    tmp_path: Path,
) -> None:
    from fastapi.testclient import TestClient

    from quality.tests.test_web_app import FakeAgent, make_deps
    from web.app import create_app

    app = create_app(
        tmp_path,
        deps_factory=make_deps,
        agent_factory=lambda _deps: FakeAgent(),
    )
    client = TestClient(app)
    project = client.post("/api/projects", json={"name": "Web 筛选"}).json()
    batch = client.post(
        f"/api/projects/{project['id']}/batches",
        json={"name": "Web 筛选批次"},
    ).json()
    write_standard_flow(
        app.state.projects.batch_workspace(project["id"], batch["id"])
    )
    base_url = f"/api/projects/{project['id']}/batches/{batch['id']}"

    filtered = client.post(
        f"{base_url}/filters", json={"expected_rows_per_day": 1}
    )
    assert filtered.status_code == 201
    candidate = filtered.json()
    assert candidate["status"] == "awaiting_confirmation"
    assert candidate["summary"]["selected_point_days"] == 3
    downloaded = client.get(
        f"{base_url}/filters/{candidate['filter_id']}/download"
    )
    assert downloaded.status_code == 200
    assert downloaded.headers["content-disposition"].endswith(
        'filename="filter_result.xlsx"'
    )

    revised = client.post(
        f"{base_url}/filters/{candidate['filter_id']}/revisions",
        files={
            "file": (
                "modified.xlsx",
                downloaded.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert revised.status_code == 201
    revision = revised.json()
    assert revision["version"] == 2
    assert client.post(
        f"{base_url}/filters/{revision['filter_id']}/confirmation",
        json={"confirm": False},
    ).status_code == 400

    confirmation = client.post(
        f"{base_url}/filters/{revision['filter_id']}/confirmation",
        json={"confirm": True},
    )
    assert confirmation.status_code == 200
    baseline = confirmation.json()
    assert baseline["identity"]["project_id"] == project["id"]
    assert client.get(f"{base_url}/baseline").json()["baseline_id"] == baseline[
        "baseline_id"
    ]


def test_web_workbench_exposes_filter_confirmation_controls(tmp_path: Path) -> None:
    from fastapi.testclient import TestClient

    from quality.tests.test_web_app import FakeAgent, make_deps
    from web.app import create_app

    response = TestClient(
        create_app(
            tmp_path,
            deps_factory=make_deps,
            agent_factory=lambda _deps: FakeAgent(),
        )
    ).get("/")

    assert response.status_code == 200
    assert 'id="runFilterButton"' in response.text
    assert 'id="filterSummary"' in response.text
    assert 'id="filterRevisionFile"' in response.text
    assert 'id="confirmFilterButton"' in response.text
    assert "/filters" in response.text


def test_agent_adapter_and_web_share_filter_baseline_service(tmp_path: Path) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("Agent 筛选")
    batch = projects.create_batch(project.id, "Agent 筛选批次")
    write_standard_flow(projects.batch_workspace(project.id, batch.id))
    service = FilterBaselineService(database, files_root)

    result = run_filter_analysis(
        service,
        project_id=project.id,
        batch_id=batch.id,
        expected_rows_per_day=1,
    )
    confirmed = confirm_filter_baseline(
        service,
        project_id=project.id,
        batch_id=batch.id,
        filter_id=result["data"]["filter_id"],
        confirmed=True,
    )

    assert result["status"] == "needs_confirmation"
    assert result["data"]["selected_point_days"] == 3
    assert confirmed["status"] == "ok"
    assert confirmed["data"]["baseline_id"] == service.current_baseline(
        project.id, batch.id
    ).baseline_id


def test_agent_data_filter_tool_uses_project_baseline_service(
    tmp_path: Path,
) -> None:
    from agent.tools.module_tools import (
        confirm_pending_filter_result,
        data_filter_impl,
    )
    from quality.tests.test_web_app import make_deps

    database = tmp_path / "var" / "drainage.sqlite3"
    files_root = tmp_path / "var" / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("Agent 工具")
    batch = projects.create_batch(project.id, "Agent 工具批次")
    write_standard_flow(projects.batch_workspace(project.id, batch.id))
    service = FilterBaselineService(database, files_root)
    deps = make_deps(tmp_path)
    deps.filter_baselines = service
    deps.current_project_id = project.id
    deps.current_batch_id = batch.id

    result = data_filter_impl(deps, expected_rows_per_day=1)
    confirmed_path = confirm_pending_filter_result(deps)

    assert result["status"] == "needs_confirmation"
    assert deps.session.pending_filter_id is None
    assert confirmed_path == service.baseline_artifact_path(project.id, batch.id)


def test_data_quality_identity_remains_baseline_free_after_confirmation(
    tmp_path: Path,
) -> None:
    database = tmp_path / "state" / "drainage.sqlite3"
    files_root = tmp_path / "projects"
    projects = ProjectRepository(database, files_root)
    project = projects.create("质量身份")
    batch = projects.create_batch(project.id, "质量身份批次")
    write_standard_flow(projects.batch_workspace(project.id, batch.id))
    service = FilterBaselineService(database, files_root)
    filtered = service.run_filter(
        FilterRequest(project.id, batch.id, expected_rows_per_day=1)
    )
    service.confirm(project.id, batch.id, filtered.filter_id)

    result = AnalysisRunner(
        database, files_root, baseline_service=service
    ).run(AnalysisRequest(project.id, batch.id, "data_quality"))

    assert result.identity["baseline"] == {"kind": "none", "identity": None}
