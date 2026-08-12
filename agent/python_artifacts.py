from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import stat
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image


RESOURCE_FILES = {
    "confirmed_flow": "standard/flow.csv",
    "rainfall": "standard/rainfall.csv",
    "site_info": "standard/sites.csv",
}
ALLOWED_EXTENSIONS = {".csv", ".json", ".png", ".xlsx"}
SAFE_NAME = re.compile(r"^[\w.-]{1,128}$", re.UNICODE)


@dataclass(frozen=True)
class SnapshotFile:
    resource: str
    source: str
    file: str
    sha256: str
    size_bytes: int


@dataclass(frozen=True)
class InputSnapshot:
    snapshot_id: str
    job_root: Path
    files: tuple[SnapshotFile, ...]


@dataclass(frozen=True)
class ValidatedArtifact:
    name: str
    size_bytes: int
    sha256: str


class UnsafeArtifact(ValueError):
    pass


def create_input_snapshot(batch_root: Path, jobs_root: Path, *, project_id: str,
                          batch_id: str, resources: list[str] | tuple[str, ...],
                          snapshot_id: str | None = None) -> InputSnapshot:
    batch_root = batch_root.resolve()
    jobs_root = jobs_root.resolve()
    snapshot_id = snapshot_id or uuid.uuid4().hex
    if not snapshot_id.isascii() or not snapshot_id.replace("-", "").isalnum():
        raise ValueError("snapshot_id must be an opaque ASCII identifier")
    job_root = (jobs_root / snapshot_id).resolve()
    if not job_root.is_relative_to(jobs_root):
        raise ValueError("snapshot path escapes jobs root")
    input_root = job_root / "input"
    input_root.mkdir(parents=True, exist_ok=False)
    (job_root / "code").mkdir()
    (job_root / "output").mkdir()
    files: list[SnapshotFile] = []
    for resource in resources:
        relative = RESOURCE_FILES.get(resource)
        if relative is None:
            raise ValueError(f"unsupported snapshot resource: {resource}")
        source = (batch_root / relative).resolve()
        if not source.is_relative_to(batch_root) or not source.is_file() or source.is_symlink():
            raise FileNotFoundError(f"authorized resource unavailable: {resource}")
        target = input_root / Path(relative).name
        shutil.copyfile(source, target)
        target.chmod(stat.S_IREAD)
        files.append(SnapshotFile(resource, relative, target.name, _sha256(target), target.stat().st_size))
    manifest = {
        "contract_version": 1, "snapshot_id": snapshot_id, "project_id": project_id,
        "batch_id": batch_id, "created_at": datetime.now(timezone.utc).isoformat(),
        "files": [item.__dict__ for item in files],
    }
    (input_root / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2),
                                               encoding="utf-8")
    return InputSnapshot(snapshot_id, job_root, tuple(files))


def validate_and_receive_artifacts(output_root: Path, exports_root: Path, *, overwrite: bool,
                                   max_files: int = 20, max_file_bytes: int = 16 * 1024 * 1024,
                                   max_total_bytes: int = 64 * 1024 * 1024) -> tuple[ValidatedArtifact, ...]:
    output_root = output_root.resolve()
    exports_root = exports_root.resolve()
    candidates = list(output_root.iterdir()) if output_root.is_dir() else []
    if len(candidates) > max_files:
        raise UnsafeArtifact("artifact count exceeds limit")
    validated: list[tuple[Path, ValidatedArtifact]] = []
    total = 0
    for path in candidates:
        info = os.lstat(path)
        if not stat.S_ISREG(info.st_mode) or path.is_symlink():
            raise UnsafeArtifact("artifact must be a regular non-link file")
        if not SAFE_NAME.fullmatch(path.name) or path.suffix.lower() not in ALLOWED_EXTENSIONS:
            raise UnsafeArtifact("artifact name or extension is not allowed")
        if info.st_size > max_file_bytes:
            raise UnsafeArtifact("artifact exceeds per-file limit")
        total += info.st_size
        if total > max_total_bytes:
            raise UnsafeArtifact("artifacts exceed total size limit")
        _parse_artifact(path)
        validated.append((path, ValidatedArtifact(path.name, info.st_size, _sha256(path))))
    exports_root.mkdir(parents=True, exist_ok=True)
    for source, artifact in validated:
        target = (exports_root / artifact.name).resolve()
        if not target.is_relative_to(exports_root):
            raise UnsafeArtifact("artifact target escapes exports")
        if target.exists() and not overwrite:
            raise UnsafeArtifact("artifact would overwrite existing export")
    for source, artifact in validated:
        shutil.copyfile(source, exports_root / artifact.name)
    return tuple(item for _, item in validated)


def _parse_artifact(path: Path) -> None:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            frame = pd.read_csv(path)
            for value in frame.to_numpy().flat:
                if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
                    raise UnsafeArtifact("CSV formula injection is not allowed")
        elif suffix == ".json":
            json.loads(path.read_text(encoding="utf-8"))
        elif suffix == ".png":
            with Image.open(path) as image:
                if image.format != "PNG":
                    raise UnsafeArtifact("file is not a PNG")
                image.verify()
        elif suffix == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        if any(cell.data_type == "f" for cell in row):
                            raise UnsafeArtifact("Excel formulas are not allowed")
            finally:
                workbook.close()
    except UnsafeArtifact:
        raise
    except Exception as exc:
        raise UnsafeArtifact(f"artifact cannot be safely parsed: {path.name}") from exc


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
