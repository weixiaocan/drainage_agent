from __future__ import annotations

from copy import deepcopy
import json
import pickle
import re
import time
import traceback
from pathlib import Path
from typing import Any, Callable

import pandas as pd

from agent.deps import AgentDeps
from agent.tools.manifest import data_fingerprint, load_manifest, record_result
from agent.types import ToolResult, error, needs_input, ok
from analysis import io
from analysis.dry_curves import build_dry_curves, dry_statistics
from analysis.event_response import analyze_event_response
from analysis.filtering import FilterConfig, run_data_filter
from analysis.patterns import analyze_patterns
from analysis.rainfall import analyze_rainfall
from analysis.rdii import analyze_rdii
from analysis.reporting import build_report
from analysis.risk import assess_risk
from analysis.schema import to_display_columns
from analysis.stats import check_data


SHEET_TABLE_TYPES = {
    "数据体检": "data_check",
    "数据收集率统计": "data_check",
    "日降雨量统计": "rainfall_daily",
    "降雨概况": "rainfall_daily",
    "场次降雨统计": "rainfall_events",
    "降雨场次分析": "rainfall_events",
    "排污规律分析": "patterns",
    "旱天分析": "dry_stats",
    "旱天风险": "dry_risk",
    "雨天溢流风险": "rainy_risk",
    "雨天事件统计": "event_response",
    "RDII总量统计": "rdii",
}


def _rel(deps: AgentDeps, path: Path) -> str:
    try:
        return path.resolve().relative_to(deps.paths.root).as_posix()
    except ValueError:
        return str(path)


def _result_artifacts(deps: AgentDeps, data: dict[str, Any]) -> list[str]:
    """Return only files owned by this tool result, never historical output files."""
    candidates: list[Any] = []
    for destination in data.get("result_destinations", []):
        if isinstance(destination, dict) and destination.get("path"):
            candidates.append(destination["path"])
    for key in ("chart_paths", "curve_images", "output_file"):
        if key in data:
            candidates.append(data[key])

    paths: list[str] = []

    def collect(value: Any) -> None:
        if isinstance(value, dict):
            for child in value.values():
                collect(child)
        elif isinstance(value, (list, tuple, set)):
            for child in value:
                collect(child)
        elif isinstance(value, (str, Path)):
            path = Path(value)
            if not path.is_absolute():
                path = deps.paths.root / path
            if path.is_file():
                relative = _rel(deps, path)
                if relative not in paths:
                    paths.append(relative)

    collect(candidates)
    return paths


def _destination_note(destinations: list[dict[str, Any]]) -> str:
    notes: list[str] = []
    for destination in destinations:
        kind = destination.get("kind")
        path = destination.get("path")
        if kind == "combined_xlsx" and path:
            sheet = destination.get("sheet")
            notes.append(f"已写入 {path}" + (f"（{sheet}）" if sheet else ""))
        elif kind == "csv" and path:
            notes.append(f"已导出 CSV：{path}")
        elif kind == "not_persisted":
            notes.append("本次结果未落盘")
    return "；".join(notes)


class ToolLLMClient:
    def __init__(self, deps: AgentDeps):
        from openai import OpenAI

        kwargs: dict[str, Any] = {"api_key": deps.settings.api_key}
        if deps.settings.base_url:
            kwargs["base_url"] = deps.settings.base_url
        self._client = OpenAI(**kwargs)
        self._model = deps.settings.model
        self._prompt_dirs = [deps.paths.root / "agent" / "prompts", deps.paths.root / "prompts"]

    def load_prompt(self, name: str) -> str:
        for prompt_dir in self._prompt_dirs:
            path = prompt_dir / f"{name}.txt"
            if path.exists():
                return path.read_text(encoding="utf-8")
        raise FileNotFoundError(name)

    def chat_json(self, prompt: str, temperature: float = 0.1) -> str:
        last_exc: Exception | None = None
        for attempt in range(3):
            try:
                response = self._client.chat.completions.create(
                    model=self._model,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=temperature,
                    response_format={"type": "json_object"},
                )
                return response.choices[0].message.content or "{}"
            except Exception as exc:
                last_exc = exc
                if attempt < 2:
                    time.sleep(2**attempt)
        raise last_exc or RuntimeError("LLM JSON call failed")


def _build_tool_llm_client(deps: AgentDeps) -> ToolLLMClient | None:
    if not deps.settings.api_key:
        return None
    return ToolLLMClient(deps)


def _write_sheet(path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    display_df = to_display_columns(df, SHEET_TABLE_TYPES.get(sheet_name, ""))
    mode = "a" if path.exists() else "w"
    if mode == "a":
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            display_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            display_df.to_excel(writer, sheet_name=sheet_name, index=False)


def _site_point_ids(deps: AgentDeps) -> set[str]:
    sites = io.load_sites(root=deps.paths.root)
    for column in ("point_id", "点位编号", "监测点编号", "点位"):
        if column in sites.columns:
            return set(sites[column].dropna().astype(str))
    return set()


def is_full_network(points: list[str] | None, deps: AgentDeps) -> bool:
    if not points:
        return True
    all_points = _site_point_ids(deps)
    if not all_points:
        return False
    values = {str(point).strip() for point in points if str(point).strip()}
    full_scope_aliases = {"全网", "全部点", "全部点位", "所有点", "所有点位"}
    if values.intersection(full_scope_aliases):
        return True
    for value in values:
        match = re.fullmatch(r"(\d+)\s*个?\s*点(?:位)?", value)
        if match and int(match.group(1)) == len(all_points):
            return True
    return all_points.issubset(values)


def _normalize_point_scope(points: list[str] | None, deps: AgentDeps) -> list[str] | None:
    """Use one canonical representation for full-network scope."""
    if is_full_network(points, deps):
        return None
    return list(dict.fromkeys(str(point).strip() for point in points or [] if str(point).strip()))


def is_full_time_range(start: str | None = None, end: str | None = None) -> bool:
    return start is None and end is None


def is_complete_scope(
    points: list[str] | None,
    deps: AgentDeps,
    start: str | None = None,
    end: str | None = None,
) -> bool:
    return is_full_network(points, deps) and is_full_time_range(start, end)


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", value.strip())
    return cleaned.strip(" ._") or "结果"


def _point_result_prefix(points: list[str] | None) -> str:
    values = sorted({str(point) for point in points or []})
    return "_".join(_safe_filename_part(value) for value in values) or "部分点位"


def _time_result_prefix(start: str | None, end: str | None) -> str:
    if is_full_time_range(start, end):
        return "全时段"

    def format_bound(value: str | None, fallback: str) -> str:
        if value is None:
            return fallback
        parsed = pd.to_datetime(value, errors="coerce")
        if not pd.isna(parsed):
            text = str(value).strip()
            if len(text) > 10:
                return parsed.strftime("%Y-%m-%d_%H-%M-%S")
            return parsed.strftime("%Y-%m-%d")
        return _safe_filename_part(value)

    if start is not None and end is None:
        return f"{format_bound(start, '起始')}_之后"
    if start is None and end is not None:
        return f"{format_bound(end, '结束')}_之前"
    return f"{format_bound(start, '起始')}_{format_bound(end, '结束')}"


def _range_result_prefix(
    points: list[str] | None,
    deps: AgentDeps,
    start: str | None,
    end: str | None,
) -> str:
    point_prefix = "全网" if is_full_network(points, deps) else _point_result_prefix(points)
    return f"{point_prefix}_{_time_result_prefix(start, end)}"


def _route_table_result(
    deps: AgentDeps,
    df: pd.DataFrame,
    sheet_name: str,
    points: list[str] | None,
    export: bool,
    start: str | None = None,
    end: str | None = None,
) -> dict[str, Any]:
    if export:
        filename = f"{_range_result_prefix(points, deps, start, end)}_{_safe_filename_part(sheet_name)}.csv"
        output_path = deps.paths.outputs / filename
        output_path.parent.mkdir(parents=True, exist_ok=True)
        display_df = to_display_columns(df, SHEET_TABLE_TYPES.get(sheet_name, ""))
        display_df.to_csv(output_path, index=False, encoding="utf-8-sig")
        return {"kind": "csv", "path": _rel(deps, output_path), "sheet": None}
    return {"kind": "not_persisted", "path": None, "sheet": None}


REPORT_COMBINED_SHEETS: tuple[tuple[str, str], ...] = (
    ("data_collection", "数据收集率统计"),
    ("rainfall_daily", "降雨概况"),
    ("rainfall_events", "降雨场次分析"),
    ("pattern_analysis", "排污规律分析"),
    ("dry_analysis", "旱天分析"),
    ("dry_risk", "旱天风险"),
    ("rainy_overflow_risk", "雨天溢流风险"),
)


def _write_report_combined_workbook(
    deps: AgentDeps,
    tables: dict[str, pd.DataFrame],
    output_path: Path,
) -> list[str]:
    """Write only the tables included in the current report."""
    written: list[str] = []
    if output_path.exists():
        output_path.unlink()
    if deps.paths.combined_xlsx != output_path and deps.paths.combined_xlsx.exists():
        deps.paths.combined_xlsx.unlink()
    for key, sheet_name in REPORT_COMBINED_SHEETS:
        table = tables.get(key)
        if table is None or table.empty:
            continue
        _write_sheet(output_path, sheet_name, table)
        written.append(sheet_name)
    return written


def _report_combined_workbook_path(output_file: Path) -> Path:
    report_stem = output_file.stem.removesuffix("_分析报告")
    return output_file.with_name(f"{report_stem}_综合分析结果.xlsx")


def _remove_sheet(path: Path, sheet_name: str) -> None:
    if not path.exists():
        return
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        return
    if len(workbook.sheetnames) <= 1:
        path.unlink()
        return
    workbook.remove(workbook[sheet_name])
    workbook.save(path)


def _add_rainfall_excel_charts(path: Path, daily: pd.DataFrame) -> None:
    if daily.empty or not path.exists():
        return
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    sheet_name = "降雨概况"
    workbook = load_workbook(path)
    if sheet_name not in workbook.sheetnames:
        return
    sheet = workbook[sheet_name]

    daily_rows = len(daily)
    if daily_rows == 0:
        workbook.save(path)
        return

    pie_col = 5
    rainy_days = int(pd.to_numeric(daily["rain_mm"], errors="coerce").fillna(0).gt(0).sum())
    non_rainy_days = int(daily_rows - rainy_days)
    sheet.cell(row=1, column=pie_col, value="类型")
    sheet.cell(row=1, column=pie_col + 1, value="天数")
    sheet.cell(row=2, column=pie_col, value="降雨日")
    sheet.cell(row=2, column=pie_col + 1, value=rainy_days)
    sheet.cell(row=3, column=pie_col, value="非降雨日")
    sheet.cell(row=3, column=pie_col + 1, value=non_rainy_days)

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "日降雨量时间序列"
    bar_chart.y_axis.title = "降雨量(mm)"
    bar_chart.x_axis.title = "日期"
    bar_chart.width = 20
    bar_chart.height = 10
    bar_chart.y_axis.majorGridlines = None
    bar_chart.x_axis.majorGridlines = None
    data_ref = Reference(sheet, min_col=2, min_row=1, max_row=daily_rows + 1)
    cats_ref = Reference(sheet, min_col=1, min_row=2, max_row=daily_rows + 1)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    sheet.add_chart(bar_chart, "A8")

    pie_chart = PieChart()
    pie_chart.width = 10
    pie_chart.height = 10
    pie_chart.legend = None
    pie_data = Reference(sheet, min_col=pie_col + 1, min_row=1, max_row=3)
    pie_cats = Reference(sheet, min_col=pie_col, min_row=2, max_row=3)
    pie_chart.add_data(pie_data, titles_from_data=True)
    pie_chart.set_categories(pie_cats)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.dataLabels.showVal = True
    pie_chart.dataLabels.showCatName = True
    sheet.add_chart(pie_chart, "A28")

    workbook.save(path)


def _save_rainfall_png_charts(
    daily: pd.DataFrame,
    output_dir: Path,
    scope_prefix: str,
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "daily_bar": output_dir / f"{scope_prefix}_日降雨量时间序列图.png",
        "rainy_ratio": output_dir / f"{scope_prefix}_降雨日占比饼图.png",
    }
    if daily.empty:
        return {key: str(value) for key, value in paths.items()}
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return {key: str(value) for key, value in paths.items()}

    plot_df = daily.copy()
    plot_df["date"] = pd.to_datetime(plot_df["date"], errors="coerce")
    plot_df["rain_mm"] = pd.to_numeric(plot_df["rain_mm"], errors="coerce").fillna(0)
    plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    fig, ax = plt.subplots(figsize=(9.2, 4.8), dpi=180)
    labels = [d.strftime("%Y-%m-%d") if not pd.isna(d) else "" for d in plot_df["date"]]
    ax.bar(range(len(plot_df)), plot_df["rain_mm"], color="#5B9BD5", edgecolor="#2F5597", linewidth=0.6)
    ax.set_xticks(range(len(plot_df)))
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=7.5)
    ax.set_ylabel("降雨量(mm)")
    ax.set_xlabel("日期")
    ax.set_title("日降雨量时间序列")
    ax.grid(False)
    fig.tight_layout()
    fig.savefig(paths["daily_bar"], bbox_inches="tight")
    plt.close(fig)

    rainy_days = int(plot_df["rain_mm"].gt(0).sum())
    non_rainy_days = int(len(plot_df) - rainy_days)
    total_days = max(1, rainy_days + non_rainy_days)
    labels = ["降雨日", "非降雨日"]

    def autopct(pct: float) -> str:
        count = int(round(pct * total_days / 100.0))
        label = labels.pop(0)
        return f"{label}\n{count}天\n{pct:.0f}%"

    fig, ax = plt.subplots(figsize=(4.8, 4.8), dpi=180)
    ax.pie(
        [rainy_days, non_rainy_days],
        labels=["", ""],
        autopct=autopct,
        pctdistance=0.58,
        startangle=90,
        colors=["#5B9BD5", "#ED7D31"],
        wedgeprops={"edgecolor": "white", "linewidth": 1.0},
        textprops={"fontsize": 10, "color": "black", "ha": "center"},
    )
    ax.axis("equal")
    fig.tight_layout()
    fig.savefig(paths["rainy_ratio"], bbox_inches="tight")
    plt.close(fig)
    return {key: str(value) for key, value in paths.items()}


def _save_rdii_curve_pngs(
    rdii_curve_data: dict[int, dict[str, pd.DataFrame]],
    rain: pd.DataFrame,
    events: pd.DataFrame,
    output_dir: Path,
    delay_hours: float = 12.0,
    selected_events: list[int] | None = None,
) -> dict[int, dict[str, str]]:
    saved: dict[int, dict[str, str]] = {}
    if not rdii_curve_data or rain.empty or events.empty:
        return saved
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib as mpl
        from matplotlib import gridspec
        import matplotlib.pyplot as plt
    except Exception:
        return saved

    mpl.rcParams["font.sans-serif"] = ["SimHei"]
    mpl.rcParams["font.serif"] = ["SimHei"]
    mpl.rcParams["axes.unicode_minus"] = False

    rain_df = rain.copy()
    rain_df["timestamp"] = pd.to_datetime(rain_df["timestamp"], errors="coerce")
    rain_df = rain_df.dropna(subset=["timestamp"]).sort_values("timestamp").set_index("timestamp")
    if "rain_mm" not in rain_df.columns:
        return saved

    event_ids = sorted(int(event_id) for event_id in rdii_curve_data.keys())
    if selected_events:
        wanted = {int(event_id) for event_id in selected_events}
        event_ids = [event_id for event_id in event_ids if event_id in wanted]

    for event_id in event_ids:
        event_rows = events[events["event_id"].astype(int) == int(event_id)]
        if event_rows.empty:
            continue
        event = event_rows.iloc[0]
        start = pd.to_datetime(event["start_time"], errors="coerce")
        end = pd.to_datetime(event["end_time"], errors="coerce")
        if pd.isna(start) or pd.isna(end):
            continue
        plot_end = end + pd.Timedelta(hours=delay_hours)
        event_rain = rain_df.loc[start:plot_end, "rain_mm"].copy()

        event_saved: dict[str, str] = {}
        for point_id, rdii_df in rdii_curve_data.get(event_id, {}).items():
            data_to_plot = rdii_df.copy()
            if data_to_plot.empty:
                continue
            data_to_plot.index = pd.to_datetime(data_to_plot.index, errors="coerce")
            data_to_plot = data_to_plot[~data_to_plot.index.isna()].sort_index()
            rename_map = {
                "rain_flow_lps": "雨天流量",
                "dry_flow_lps": "旱天流量",
                "rdii_lps": "RDII",
            }
            data_to_plot = data_to_plot.rename(columns=rename_map)
            keep_cols = [col for col in ("雨天流量", "旱天流量", "RDII") if col in data_to_plot.columns]
            if not keep_cols:
                continue

            event_dir = output_dir / "rdii_curve" / f"event{event_id}_{start.month}_{start.day}"
            event_dir.mkdir(parents=True, exist_ok=True)
            fig = plt.figure(figsize=(10, 5))
            grid = gridspec.GridSpec(2, 1, height_ratios=[1, 3])
            ax_flow = plt.subplot(grid[1])
            ax_rain = plt.subplot(grid[0])
            ax_rain.get_xaxis().set_visible(False)
            fig.subplots_adjust(hspace=0)

            data_to_plot[keep_cols].plot(ax=ax_flow, legend=True)
            ax_flow.set_xlabel('时间', fontsize='large')
            ax_flow.set_ylabel('流量/(L/s)', fontsize='large')

            rain_to_plot = _regularize_rain_series_for_plot(event_rain)
            if len(rain_to_plot) > 500:
                rain_to_plot = rain_to_plot.resample("10min").sum()
                rain_to_plot = _regularize_rain_series_for_plot(rain_to_plot)
            ax_rain.bar(range(len(rain_to_plot)), rain_to_plot.to_numpy(dtype=float), width=0.8)
            ax_rain.set_ylabel('降雨/mm', fontsize='large')
            if len(rain_to_plot) > 100:
                n_ticks = min(10, len(rain_to_plot))
                step = len(rain_to_plot) // n_ticks
                tick_positions = list(range(0, len(rain_to_plot), step))
                ax_rain.set_xticks(tick_positions)
                ax_rain.set_xticklabels(
                    [rain_to_plot.index[i].strftime('%m-%d %H:%M') for i in tick_positions],
                    rotation=45,
                    ha='right',
                )

            image_path = event_dir / f"{point_id}_event{event_id}.png"
            fig.savefig(image_path, dpi=300, bbox_inches="tight")
            plt.cla()
            plt.clf()
            plt.close(fig)
            event_saved[str(point_id)] = str(image_path)

        if event_saved:
            saved[event_id] = event_saved

    return saved


def _regularize_rain_series_for_plot(series: pd.Series) -> pd.Series:
    if series.empty:
        return series
    result = pd.to_numeric(series, errors="coerce").fillna(0.0).copy()
    index = pd.DatetimeIndex(pd.to_datetime(result.index, errors="coerce"))
    valid = ~index.isna()
    result = result[valid]
    index = pd.DatetimeIndex(index[valid])
    if result.empty:
        return result

    freq = index.freq or pd.infer_freq(index)
    if freq is None and len(index) > 1:
        deltas = index.to_series().diff().dropna()
        positive_deltas = deltas[deltas > pd.Timedelta(0)]
        if not positive_deltas.empty:
            freq = positive_deltas.min()
    if freq is not None:
        full_index = pd.date_range(index.min(), index.max(), freq=freq)
        result.index = index
        result = result.groupby(level=0).sum().reindex(full_index, fill_value=0.0)
        return result

    result.index = index
    return result


def _daily_curve_frame(dry_flow: pd.DataFrame, point_id: str, value_col: str) -> pd.DataFrame:
    if dry_flow.empty or value_col not in dry_flow.columns:
        return pd.DataFrame()

    point_flow = dry_flow[dry_flow["point_id"].astype(str) == str(point_id)].copy()
    if point_flow.empty:
        return pd.DataFrame()

    point_flow["timestamp"] = pd.to_datetime(point_flow["timestamp"], errors="coerce")
    point_flow = point_flow.dropna(subset=["timestamp"]).sort_values("timestamp")
    if point_flow.empty:
        return pd.DataFrame()

    day_index = pd.date_range("00:00:00", "23:59:00", freq="min")
    daily_df = pd.DataFrame(index=day_index)
    for date, day_data in point_flow.groupby(point_flow["timestamp"].dt.strftime("%Y-%m-%d"), sort=True):
        values = pd.to_numeric(day_data[value_col], errors="coerce").dropna().to_numpy()
        if len(values) == 0:
            continue
        padded = [0.0] * 1440
        limit = min(len(values), 1440)
        padded[:limit] = values[:limit]
        daily_df[str(date)] = padded
    return daily_df


def _plot_pipeline_pattern_curve(
    daily_df: pd.DataFrame,
    curve: pd.DataFrame,
    value_col: str,
    output_path: Path,
    daily_label: str,
    curve_label: str,
    y_label: str,
) -> bool:
    if daily_df.empty or daily_df.shape[1] == 0 or value_col not in curve.columns:
        return False

    import matplotlib.dates as mdates
    import matplotlib.pyplot as plt

    plot_curve = curve.copy()
    if "minute_of_day" in plot_curve.columns:
        plot_curve = plot_curve.sort_values("minute_of_day")
    curve_values = pd.to_numeric(plot_curve[value_col], errors="coerce").to_numpy()
    if len(curve_values) == 0:
        return False

    day_index = pd.date_range("00:00:00", "23:59:00", freq="min")
    padded_curve = [0.0] * 1440
    limit = min(len(curve_values), 1440)
    padded_curve[:limit] = curve_values[:limit]
    curve_series = pd.Series(padded_curve, index=day_index)

    fig = plt.figure(figsize=(10, 5), dpi=120)
    ax = fig.add_subplot(1, 1, 1)
    valid_days = list(daily_df.columns)
    for idx, day in enumerate(valid_days):
        label = daily_label if idx == len(valid_days) - 1 else ""
        daily_df[day].plot(ax=ax, color="#D3D3D3", label=label, legend=bool(label), alpha=0.5)
    curve_series.plot(ax=ax, color="#1E90FF", label=curve_label, legend=True)

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(byhour=range(3, 24, 3)))
    ax.set_xlabel("时间")
    ax.set_ylabel(y_label)
    ax.grid(False)
    ax.legend(loc="upper right")
    plt.tight_layout()
    fig.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return True


def _save_pattern_curve_pngs(
    curves: dict[str, pd.DataFrame],
    dry_flow: pd.DataFrame,
    output_dir: Path,
    scope_prefix: str,
) -> dict[str, list[str]]:
    output_dir = output_dir / scope_prefix
    output_dir.mkdir(parents=True, exist_ok=True)
    saved: dict[str, list[str]] = {}
    if not curves:
        return saved
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return saved

    plt.rcParams["font.sans-serif"] = ["SimSun", "Microsoft YaHei", "SimHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    hour_ticks = list(range(0, 1441, 120))
    hour_labels = [f"{hour // 60:02d}:00" for hour in hour_ticks]

    for point_id, curve in curves.items():
        point_paths: list[str] = []
        plot_df = curve.copy()
        minutes = pd.to_numeric(plot_df["minute_of_day"], errors="coerce").fillna(0)

        if "flow_lps" in plot_df.columns:
            flow_path = output_dir / f"{point_id}_流量特征曲线.png"
            flow_daily = _daily_curve_frame(dry_flow, point_id, "flow_lps")
            if _plot_pipeline_pattern_curve(
                flow_daily,
                plot_df,
                "flow_lps",
                flow_path,
                "每日流量",
                "流量特征曲线_总体",
                "流量/(L/s)",
            ):
                point_paths.append(str(flow_path))

        if "level_m" in plot_df.columns:
            level_path = output_dir / f"{point_id}_液位特征曲线.png"
            level_daily = _daily_curve_frame(dry_flow, point_id, "level_m")
            if _plot_pipeline_pattern_curve(
                level_daily,
                plot_df,
                "level_m",
                level_path,
                "每日液位",
                "液位特征曲线",
                "液位/(m)",
            ):
                point_paths.append(str(level_path))

        saved[point_id] = point_paths
    return saved


def _save_partial_pattern_curve_png(
    curves: dict[str, pd.DataFrame],
    dry_flow: pd.DataFrame,
    points: list[str] | None,
    output_dir: Path,
    scope_prefix: str,
) -> dict[str, list[str]]:
    selected_points = {str(value) for value in points or []}
    selected_curves = {
        point_id: curve
        for point_id, curve in curves.items()
        if not selected_points or point_id in selected_points
    }
    return _save_pattern_curve_pngs(
        selected_curves,
        dry_flow,
        output_dir / "特征曲线图",
        scope_prefix,
    )


def _save_partial_rdii_curve_png(
    curve_data: dict[int, dict[str, pd.DataFrame]],
    points: list[str] | None,
    output_dir: Path,
    event_ids: list[int] | None,
) -> dict[int, dict[str, str]]:
    selected_points = {str(point) for point in points or []}
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return {}

    fig = plt.figure(figsize=(10, 5), dpi=120)
    ax = fig.add_subplot(1, 1, 1)
    plotted = False
    for event_id in sorted(curve_data):
        for point_id, frame in sorted(curve_data[event_id].items()):
            if selected_points and str(point_id) not in selected_points:
                continue
            if frame.empty or "rdii_lps" not in frame.columns:
                continue
            values = pd.to_numeric(frame["rdii_lps"], errors="coerce")
            ax.plot(pd.to_datetime(frame.index, errors="coerce"), values, label=f"{point_id}-事件{event_id}")
            plotted = True
    if not plotted:
        plt.close(fig)
        return {}
    output_dir.mkdir(parents=True, exist_ok=True)
    ax.set_xlabel("时间")
    ax.set_ylabel("RDII/(L/s)")
    ax.legend(loc="upper right")
    ax.grid(False)
    fig.tight_layout()
    event_prefix = "_".join(f"event{event_id}" for event_id in sorted(event_ids or [])) or "event未指定"
    path = output_dir / f"{_point_result_prefix(points)}_{event_prefix}_RDII曲线.png"
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return {0: {"selected": str(path)}}


def _run(
    deps: AgentDeps,
    tool_name: str,
    fn: Callable[[], tuple[str, dict[str, Any]]],
    params: dict[str, Any] | None = None,
    use_cache: bool = True,
) -> ToolResult:
    cache_key = _analysis_cache_key(deps, tool_name, params or {})
    if use_cache and cache_key in deps.session.analysis_cache:
        cached = deepcopy(deps.session.analysis_cache[cache_key])
        return ok(cached["summary"], artifacts=_result_artifacts(deps, cached["data"]), **cached["data"])
    try:
        deps.paths.outputs.mkdir(parents=True, exist_ok=True)
        summary, data = fn()
        destination_note = _destination_note(data.get("result_destinations", []))
        if destination_note:
            summary = f"{summary} 落盘去向：{destination_note}。"
        artifacts = _result_artifacts(deps, data)
        record_result(deps, tool_name, artifacts, params=params)
        if use_cache:
            deps.session.analysis_cache[cache_key] = {"summary": summary, "data": deepcopy(data)}
        _store_report_data_components(deps, tool_name, params or {}, data)
        return ok(summary, artifacts=artifacts, **data)
    except Exception as exc:
        deps.logger.exception("%s failed", tool_name)
        return error(f"{tool_name} 执行失败: {exc}", traceback=traceback.format_exc(limit=8))


def _analysis_cache_key(deps: AgentDeps, tool_name: str, params: dict[str, Any]) -> str:
    normalized_params = dict(params)
    for key in ("points", "event_ids", "sections"):
        value = normalized_params.get(key)
        if isinstance(value, list):
            normalized_params[key] = sorted(value, key=str)
    payload = {
        "tool": tool_name,
        "params": normalized_params,
        "data_fingerprint": data_fingerprint(deps)["digest"],
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)


def _report_data_key(
    deps: AgentDeps,
    component: str,
    points: list[str] | None = None,
    start: str | None = None,
    end: str | None = None,
    event_ids: list[int] | None = None,
) -> str:
    payload = {
        "component": component,
        "points": sorted(points or [], key=str),
        "start": start,
        "end": end,
        "event_ids": sorted(event_ids or []),
        "data_fingerprint": data_fingerprint(deps)["digest"],
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)


def _store_report_data_components(
    deps: AgentDeps,
    tool_name: str,
    params: dict[str, Any],
    data: dict[str, Any],
) -> None:
    points = params.get("points") or []
    start = params.get("start")
    end = params.get("end")
    event_ids = params.get("event_ids") or []
    components: list[tuple[str, str, list[int] | None]] = []
    if tool_name == "check_data":
        components = [("data_collection", "table", None)]
    elif tool_name == "analyze_rainfall":
        time_range = params.get("time_range") or []
        start = time_range[0] if len(time_range) == 2 else None
        end = time_range[1] if len(time_range) == 2 else None
        components = [
            ("rainfall_daily", "daily", None),
            ("rainfall_events", "events", None),
            ("rainfall_chart_paths", "chart_paths", None),
        ]
    elif tool_name == "analyze_patterns":
        components = [
            ("pattern_analysis", "table", None),
            ("pattern_chart_paths", "curve_images", None),
        ]
    elif tool_name == "assess_risk":
        scope = params.get("scope", "all")
        if scope in {"dry", "all"} and data.get("dry_analysis") is not None:
            components.append(("dry_analysis", "dry_analysis", None))
        if scope in {"dry", "all"} and data.get("dry_risk") is not None:
            components.append(("dry_risk", "dry_risk", None))
        if scope in {"rainy", "all"} and data.get("rainy_risk") is not None:
            components.append(("rainy_overflow_risk", "rainy_risk", event_ids))
    for component, data_key, component_events in components:
        records = data.get(data_key)
        if records is None:
            continue
        key = _report_data_key(deps, component, points, start, end, component_events)
        deps.session.report_data_cache[key] = deepcopy(records)


def _cached_report_frame(
    deps: AgentDeps,
    component: str,
    points: list[str] | None = None,
    start: str | None = None,
    end: str | None = None,
    event_ids: list[int] | None = None,
) -> pd.DataFrame | None:
    key = _report_data_key(deps, component, points, start, end, event_ids)
    records = deps.session.report_data_cache.get(key)
    return pd.DataFrame(deepcopy(records)) if records is not None else None


def _cached_report_value(
    deps: AgentDeps,
    component: str,
    points: list[str] | None = None,
    start: str | None = None,
    end: str | None = None,
    event_ids: list[int] | None = None,
) -> Any | None:
    key = _report_data_key(deps, component, points, start, end, event_ids)
    value = deps.session.report_data_cache.get(key)
    return deepcopy(value) if value is not None else None


def _manifest_stale(deps: AgentDeps, tool_name: str) -> bool:
    manifest = load_manifest(deps)
    item = manifest.get("results", {}).get(tool_name)
    if not item:
        return True
    from agent.tools.manifest import data_fingerprint

    return item.get("data_fingerprint") != data_fingerprint(deps)["digest"]


def _intermediate_dir(deps: AgentDeps) -> Path:
    path = deps.paths.outputs / "intermediate"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _curves_path(deps: AgentDeps) -> Path:
    return _intermediate_dir(deps) / "dry_curves.pkl"


def _rdii_curves_path(deps: AgentDeps) -> Path:
    return _intermediate_dir(deps) / "rdii_curves.pkl"


def _save_curves(deps: AgentDeps, curves: dict[str, pd.DataFrame]) -> None:
    with _curves_path(deps).open("wb") as fh:
        pickle.dump(curves, fh)


def _save_rdii_curves(deps: AgentDeps, curves: dict[int, dict[str, pd.DataFrame]]) -> None:
    with _rdii_curves_path(deps).open("wb") as fh:
        pickle.dump(curves, fh)


def _load_curves(deps: AgentDeps) -> dict[str, pd.DataFrame]:
    path = _curves_path(deps)
    if not path.exists():
        return {}
    with path.open("rb") as fh:
        return pickle.load(fh)


def data_filter_impl(
    deps: AgentDeps,
    missing_rate_threshold: float = 0.1,
    expected_rows_per_day: int = 1440,
    rain_day_filter_threshold: float = 2.0,
    zero_like_threshold: float = 0.02,
    high_zero_ratio_threshold: float = 0.5,
    high_zero_ratio_normal_days_threshold: int = 5,
    zero_day_drop_min_nonzero_keep_days: int = 3,
    mean_lower_ratio: float = 0.5,
    mean_upper_ratio: float = 2.0,
    output_file: str | None = None,
) -> ToolResult:
    params = {
        "missing_rate_threshold": missing_rate_threshold,
        "expected_rows_per_day": expected_rows_per_day,
        "rain_day_filter_threshold": rain_day_filter_threshold,
        "zero_like_threshold": zero_like_threshold,
        "high_zero_ratio_threshold": high_zero_ratio_threshold,
        "high_zero_ratio_normal_days_threshold": high_zero_ratio_normal_days_threshold,
        "zero_day_drop_min_nonzero_keep_days": zero_day_drop_min_nonzero_keep_days,
        "mean_lower_ratio": mean_lower_ratio,
        "mean_upper_ratio": mean_upper_ratio,
        "output_file": output_file or "",
    }

    def work() -> tuple[str, dict[str, Any]]:
        flow = io.load_flow(root=deps.paths.root)
        rain = io.load_rain(root=deps.paths.root)
        out_path = Path(output_file) if output_file else deps.paths.filter_result
        if not out_path.is_absolute():
            out_path = deps.paths.root / out_path
        selected = run_data_filter(
            flow=flow,
            rain=rain,
            output_xlsx=out_path,
            config=FilterConfig(
                missing_rate_threshold=missing_rate_threshold,
                expected_rows_per_day=expected_rows_per_day,
                rain_day_filter_threshold=rain_day_filter_threshold,
                zero_like_threshold=zero_like_threshold,
                high_zero_ratio_threshold=high_zero_ratio_threshold,
                high_zero_ratio_normal_days_threshold=high_zero_ratio_normal_days_threshold,
                zero_day_drop_min_nonzero_keep_days=zero_day_drop_min_nonzero_keep_days,
                mean_lower_ratio=mean_lower_ratio,
                mean_upper_ratio=mean_upper_ratio,
            ),
        )
        point_count = len(selected)
        total_days = sum(len(days) for days in selected.values())
        summary = f"数据筛选完成：处理 {point_count} 个点位，筛出有效旱天 {total_days} 个点位日，输出 {_rel(deps, out_path)}。"
        return summary, {"selected": selected, "output_file": str(out_path)}

    return _run(deps, "data_filter", work, params=params)


def _load_event_table(deps: AgentDeps) -> pd.DataFrame:
    rain = io.load_rain(root=deps.paths.root)
    return analyze_rainfall(rain)["events"]


def _event_options(events: pd.DataFrame) -> list[dict[str, Any]]:
    options: list[dict[str, Any]] = []
    for _, row in events.iterrows():
        options.append(
            {
                "event_id": int(row["event_id"]),
                "label": f"场次{int(row['event_id'])}: {row['start_time']} 至 {row['end_time']}，总雨量 {float(row['total_rain_mm']):.1f} mm",
            }
        )
    return options


def _require_event_ids(deps: AgentDeps, event_ids: list[int] | None) -> ToolResult | None:
    if event_ids:
        deps.session.selected_event_ids = event_ids
        return None
    if deps.session.selected_event_ids:
        return None
    events = _load_event_table(deps)
    return needs_input(
        "event_ids",
        "请从 options 中选择降雨场次编号；也可以回复“只出旱天报告”。",
        summary="需要先选择降雨场次编号，才能分析雨天响应、RDII 或雨天风险。",
        options=_event_options(events),
    )


def _source_event_ids(deps: AgentDeps, event_ids: list[int]) -> list[int]:
    """Translate window-local public IDs to stable source IDs for calculations."""
    mapping = deps.session.window_event_id_map
    return [mapping.get(int(event_id), int(event_id)) for event_id in event_ids]


def _public_event_frame(deps: AgentDeps, frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    if result.empty or "event_id" not in result.columns or not deps.session.window_event_id_map:
        return result
    source_to_local = {source: local for local, source in deps.session.window_event_id_map.items()}
    result["event_id"] = result["event_id"].map(
        lambda value: source_to_local.get(int(value), int(value)) if pd.notna(value) else value
    )
    return result


def _event_data_coverage(
    deps: AgentDeps,
    event_ids: list[int],
    points: list[str] | None = None,
    delay_hours: float = 12.0,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str], list[dict[str, str]]]:
    flow = io.load_flow(points=points, root=deps.paths.root)
    events = _load_event_table(deps)

    wanted = {int(event_id) for event_id in event_ids}
    selected_events = events[events["event_id"].astype(int).isin(wanted)].copy() if not events.empty else events
    requested_points = [str(point) for point in points] if points else sorted(flow["point_id"].astype(str).unique())
    covered: list[str] = []
    excluded: list[dict[str, str]] = []
    for point_id in requested_points:
        point_flow = flow[flow["point_id"].astype(str) == point_id]
        has_coverage = False
        for _, event in selected_events.iterrows():
            start = pd.to_datetime(event.get("start_time"), errors="coerce")
            end = pd.to_datetime(event.get("end_time"), errors="coerce")
            if pd.isna(start) or pd.isna(end):
                continue
            end = end + pd.Timedelta(hours=delay_hours)
            if not point_flow[(point_flow["timestamp"] >= start) & (point_flow["timestamp"] <= end)].empty:
                has_coverage = True
                break
        if has_coverage:
            covered.append(point_id)
        else:
            excluded.append({"point_id": point_id, "reason": "该时段/该点位无数据，无法分析"})

    covered_flow = flow[flow["point_id"].astype(str).isin(covered)].copy()
    return covered_flow, events, covered, excluded


def _coverage_guard_result(
    deps: AgentDeps,
    event_ids: list[int],
    covered: list[str],
    excluded: list[dict[str, str]],
) -> ToolResult | None:
    if covered:
        return None
    deps.session.unavailable_event_ids = sorted(
        set(deps.session.unavailable_event_ids).union(event_ids)
    )
    point_labels = [item["point_id"] for item in excluded] or ["全部点位"]
    return needs_input(
        "data_coverage",
        "请选择覆盖该时段的点位或其他降雨事件。",
        summary=f"该时段/该点位无数据，无法分析：场次 {event_ids}，点位 {point_labels}。",
        options=excluded,
    )


def _window_bounds(start: str | None, end: str | None) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    start_ts = pd.to_datetime(start) if start else None
    end_ts = pd.to_datetime(end) if end else None
    if end_ts is not None and isinstance(end, str) and len(end.strip()) <= 10:
        end_ts = end_ts + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
    if start_ts is not None and end_ts is not None and start_ts > end_ts:
        raise ValueError("start must be earlier than or equal to end")
    return start_ts, end_ts


def _window_data_coverage(
    deps: AgentDeps,
    points: list[str] | None,
    start: str | None,
    end: str | None,
) -> tuple[pd.DataFrame, list[str], list[dict[str, str]], dict[str, str | None]]:
    flow = _load_filtered_dry_flow(deps, points=points)
    start_ts, end_ts = _window_bounds(start, end)
    requested_points = [str(point) for point in points] if points else sorted(flow["point_id"].astype(str).unique())
    covered: list[str] = []
    excluded: list[dict[str, str]] = []
    frames: list[pd.DataFrame] = []
    for point_id in requested_points:
        point_flow = flow[flow["point_id"].astype(str) == point_id]
        if start_ts is not None:
            point_flow = point_flow[point_flow["timestamp"] >= start_ts]
        if end_ts is not None:
            point_flow = point_flow[point_flow["timestamp"] <= end_ts]
        if point_flow.empty:
            excluded.append({"point_id": point_id, "reason": "该时段/该点位无数据，无法分析"})
            continue
        covered.append(point_id)
        frames.append(point_flow)

    window_flow = pd.concat(frames, ignore_index=True) if frames else flow.iloc[0:0].copy()
    actual_start = window_flow["timestamp"].min() if not window_flow.empty else None
    actual_end = window_flow["timestamp"].max() if not window_flow.empty else None
    coverage = {
        "requested_start": str(start) if start is not None else None,
        "requested_end": str(end) if end is not None else None,
        "actual_start": actual_start.isoformat(sep=" ") if actual_start is not None else None,
        "actual_end": actual_end.isoformat(sep=" ") if actual_end is not None else None,
    }
    return window_flow.reset_index(drop=True), covered, excluded, coverage


def _window_coverage_guard_result(
    covered: list[str],
    excluded: list[dict[str, str]],
    start: str | None,
    end: str | None,
) -> ToolResult | None:
    if covered:
        return None
    point_labels = [item["point_id"] for item in excluded] or ["全部点位"]
    return needs_input(
        "data_coverage",
        "请选择有数据覆盖的时间窗或点位。",
        summary=f"时间窗 [{start or '不限'}, {end or '不限'}] 内点位 {point_labels} 无数据覆盖，无法分析。",
        options=excluded,
    )


def _window_coverage_note(coverage: dict[str, str | None], excluded: list[dict[str, str]]) -> str:
    note = f"；实际分析范围 [{coverage['actual_start']}, {coverage['actual_end']}]"
    if excluded:
        note += f"；剔除无覆盖点位 {[item['point_id'] for item in excluded]}"
    return note


def _ensure_filter_result(deps: AgentDeps) -> Path:
    if deps.paths.filter_result.exists():
        return deps.paths.filter_result
    flow = io.load_flow(root=deps.paths.root)
    rain = io.load_rain(root=deps.paths.root)
    run_data_filter(flow=flow, rain=rain, output_xlsx=deps.paths.filter_result, config=FilterConfig())
    return deps.paths.filter_result


def _load_filtered_dry_flow(
    deps: AgentDeps,
    points: list[str] | None = None,
    time_range: list[str] | None = None,
) -> pd.DataFrame:
    filter_result = _ensure_filter_result(deps)
    return io.load_filtered_flow(points=points, time_range=time_range, root=deps.paths.root)


def check_data_impl(
    deps: AgentDeps,
    points: list[str] | None = None,
    export: bool = False,
    start: str | None = None,
    end: str | None = None,
) -> ToolResult:
    windowed = start is not None or end is not None

    def work() -> tuple[str, dict[str, Any]]:
        flow = io.load_flow(points=points, root=deps.paths.root)
        coverage = None
        if windowed:
            start_ts, end_ts = _window_bounds(start, end)
            if start_ts is not None:
                flow = flow[flow["timestamp"] >= start_ts]
            if end_ts is not None:
                flow = flow[flow["timestamp"] <= end_ts]
            coverage = {
                "requested_start": start,
                "requested_end": end,
                "actual_start": flow["timestamp"].min().isoformat(sep=" ") if not flow.empty else None,
                "actual_end": flow["timestamp"].max().isoformat(sep=" ") if not flow.empty else None,
            }
        stats_df = check_data(flow)
        if windowed and not stats_df.empty:
            effective_start = start_ts or flow["timestamp"].min()
            effective_end = end_ts or flow["timestamp"].max()
            expected = max(int((effective_end - effective_start) / pd.Timedelta(minutes=1)) + 1, 1)
            stats_df["monitoring_days"] = max(int((expected + 1439) // 1440), 1)
            stats_df["theoretical_count"] = expected
            stats_df["collection_rate"] = (stats_df["record_count"] / expected).clip(upper=1.0)
        destination = _route_table_result(
            deps, stats_df, "数据收集率统计", points, export, start=start, end=end
        )
        if destination["kind"] == "combined_xlsx":
            _remove_sheet(deps.paths.combined_xlsx, "数据体检")
        avg = float(stats_df["collection_rate"].mean()) if not stats_df.empty else 0.0
        summary = f"数据收集率统计完成：处理 {len(stats_df)} 个点位，平均收集率 {avg:.1%}。"
        data = {"table": stats_df.to_dict(orient="records"), "result_destinations": [destination]}
        if coverage is not None:
            data["window_coverage"] = coverage
        return summary, data

    return _run(
        deps,
        "check_data",
        work,
        params={"points": points or [], "export": export, "start": start, "end": end},
    )


def _rainfall_window_bounds(time_range: list[str]) -> tuple[pd.Timestamp, pd.Timestamp]:
    start_text, end_text = time_range
    start = pd.to_datetime(start_text)
    end = pd.to_datetime(end_text)
    if isinstance(end_text, str) and len(end_text.strip()) <= 10:
        end = end + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
    return start, end


def _filter_rainfall_result_to_window(
    rain: pd.DataFrame,
    result: dict[str, pd.DataFrame],
    time_range: list[str],
) -> dict[str, pd.DataFrame]:
    start, end = _rainfall_window_bounds(time_range)
    window_rain = rain[(rain["timestamp"] >= start) & (rain["timestamp"] <= end)].copy()
    window_daily = analyze_rainfall(window_rain)["daily"]

    events = result["events"].copy()
    if not events.empty:
        event_starts = pd.to_datetime(events["start_time"], errors="coerce")
        event_ends = pd.to_datetime(events["end_time"], errors="coerce")
        events = events[(event_ends >= start) & (event_starts <= end)].copy()
        events.insert(0, "source_event_id", events["event_id"].astype(int))
        events["event_id"] = range(1, len(events) + 1)
    return {"daily": window_daily, "events": events.reset_index(drop=True)}


def analyze_rainfall_impl(
    deps: AgentDeps,
    time_range: list[str] | None = None,
    output: str = "all",
    rainfall_gap_hours: int = 12,
    export: bool = False,
) -> ToolResult:
    params = {
        "time_range": time_range or [],
        "output": output,
        "rainfall_gap_hours": rainfall_gap_hours,
        "export": export,
    }

    def work() -> tuple[str, dict[str, Any]]:
        rain = io.load_rain(root=deps.paths.root)
        result = analyze_rainfall(rain, gap_hours=rainfall_gap_hours)
        if time_range:
            result = _filter_rainfall_result_to_window(rain, result, time_range)
            deps.session.window_event_id_map = {
                int(local): int(source)
                for local, source in zip(result["events"]["event_id"], result["events"]["source_event_id"])
            }
        else:
            deps.session.window_event_id_map = {}
        range_start = time_range[0] if time_range else None
        range_end = time_range[1] if time_range else None
        chart_paths: dict[str, str] = {}
        destinations: list[dict[str, Any]] = []
        if output in {"all", "daily"}:
            destination = _route_table_result(
                deps,
                result["daily"],
                "降雨概况",
                None,
                export,
                start=range_start,
                end=range_end,
            )
            destinations.append(destination)
            if destination["kind"] == "combined_xlsx":
                _remove_sheet(deps.paths.combined_xlsx, "日降雨量统计")
                _add_rainfall_excel_charts(deps.paths.combined_xlsx, result["daily"])
            chart_paths = _save_rainfall_png_charts(
                result["daily"],
                deps.paths.outputs / "降雨分析图",
                _range_result_prefix(None, deps, range_start, range_end),
            )
        if output in {"all", "events"}:
            destination = _route_table_result(
                deps,
                result["events"],
                "降雨场次分析",
                None,
                export,
                start=range_start,
                end=range_end,
            )
            destinations.append(destination)
            if destination["kind"] == "combined_xlsx":
                _remove_sheet(deps.paths.combined_xlsx, "场次降雨统计")
        rainy_days = int(result["daily"]["is_rainy"].sum()) if not result["daily"].empty else 0
        total = float(result["daily"]["rain_mm"].sum()) if not result["daily"].empty else 0.0
        summary = f"降雨分析完成：雨日 {rainy_days} 天，总雨量 {total:.1f} mm，场次 {len(result['events'])} 场。"
        data = {key: df.to_dict(orient="records") for key, df in result.items()}
        data["has_rainfall_coverage"] = bool(rainy_days or not result["events"].empty)
        data["chart_paths"] = chart_paths
        data["result_destinations"] = destinations
        return summary, data

    return _run(deps, "analyze_rainfall", work, params=params)


def _dry_inputs(deps: AgentDeps, points: list[str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, pd.DataFrame]]:
    dry_flow = _load_filtered_dry_flow(deps, points=points)
    stats_df = dry_statistics(dry_flow, io.load_sites(root=deps.paths.root))
    curves = build_dry_curves(dry_flow)
    _save_curves(deps, curves)
    return dry_flow, stats_df, curves


def analyze_patterns_impl(
    deps: AgentDeps,
    points: list[str] | None = None,
    output: str = "all",
    export: bool = False,
    start: str | None = None,
    end: str | None = None,
    report_charts: bool = False,
) -> ToolResult:
    params = {
        "points": points or [],
        "start": start,
        "end": end,
        "output": output,
        "export": export,
        "report_charts": report_charts,
    }
    windowed = start is not None or end is not None
    coverage: dict[str, str | None] | None = None
    covered: list[str] = []
    excluded: list[dict[str, str]] = []
    window_flow = pd.DataFrame()
    if windowed:
        try:
            window_flow, covered, excluded, coverage = _window_data_coverage(deps, points, start, end)
        except ValueError as exc:
            return error(str(exc))
        coverage_failure = _window_coverage_guard_result(covered, excluded, start, end)
        if coverage_failure:
            return coverage_failure

    def work() -> tuple[str, dict[str, Any]]:
        dry_flow = window_flow if windowed else _load_filtered_dry_flow(deps, points=points)
        llm_client = _build_tool_llm_client(deps)
        result = analyze_patterns(dry_flow, llm_client=llm_client)
        patterns = result["patterns"]
        curves = result["curves"]
        _save_curves(deps, curves)
        scope_prefix = _range_result_prefix(points, deps, start, end)
        if is_full_network(points, deps) or report_charts:
            curve_images = _save_pattern_curve_pngs(
                curves,
                dry_flow,
                deps.paths.outputs / "特征曲线图",
                scope_prefix,
            )
        elif export:
            curve_images = _save_partial_pattern_curve_png(
                curves, dry_flow, points, deps.paths.outputs, scope_prefix
            )
        else:
            curve_images = {}
        destination = _route_table_result(
            deps, patterns, "排污规律分析", points, export, start=start, end=end
        )
        llm_note = "描述由 LLM JSON 生成并经规则后处理" if llm_client is not None else "未配置 LLM，描述使用规则兜底生成"
        summary = (
            f"排污规律分析完成：分析 {len(patterns)} 个点位，生成 {len(curves)} 条旱天曲线。"
            f"{llm_note}。基于筛选结果 {_rel(deps, deps.paths.filter_result)}。"
        )
        if coverage is not None:
            summary += _window_coverage_note(coverage, excluded)
        data = {
            "table": patterns.to_dict(orient="records"),
            "curve_images": curve_images,
            "result_destinations": [destination],
        }
        if windowed:
            data["window_coverage"] = coverage
            data["covered_points"] = covered
            data["excluded_points"] = excluded
        return summary, data

    return _run(deps, "analyze_patterns", work, params=params)


def analyze_event_response_impl(
    deps: AgentDeps,
    event_ids: list[int] | None = None,
    points: list[str] | None = None,
    export: bool = False,
) -> ToolResult:
    precheck = _require_event_ids(deps, event_ids)
    if precheck:
        return precheck
    event_ids = event_ids or deps.session.selected_event_ids
    source_event_ids = _source_event_ids(deps, event_ids or [])
    params = {"event_ids": source_event_ids, "points": points or [], "export": export}

    flow, events, covered, excluded = _event_data_coverage(deps, source_event_ids, points)
    coverage_failure = _coverage_guard_result(deps, source_event_ids, covered, excluded)
    if coverage_failure:
        return coverage_failure

    def work() -> tuple[str, dict[str, Any]]:
        response = analyze_event_response(flow, events, source_event_ids)
        public_response = _public_event_frame(deps, response)
        destination = _route_table_result(deps, public_response, "雨天事件统计", points, export)
        if response.empty:
            deps.session.unavailable_event_ids = sorted(
                set(deps.session.unavailable_event_ids).union(event_ids or [])
            )
            selected = points or ["全部点位"]
            summary = (
                f"雨天事件统计无可用数据：场次 {event_ids} 与点位 {selected} 的监测数据无时间重叠，"
                "无法计算事件响应指标。"
            )
            return summary, {
                "table": [],
                "no_data": True,
                "event_ids": event_ids,
                "points": points or [],
                "result_destinations": [destination],
            }
        excluded_note = f"；剔除无覆盖点位 {[item['point_id'] for item in excluded]}" if excluded else ""
        summary = f"雨天事件统计完成：场次 {event_ids}，输出 {len(response)} 个点位统计{excluded_note}。"
        return summary, {
            "table": public_response.to_dict(orient="records"),
            "no_data": False,
            "covered_points": covered,
            "excluded_points": excluded,
            "result_destinations": [destination],
        }

    return _run(deps, "analyze_event_response", work, params=params)


def analyze_rdii_impl(
    deps: AgentDeps,
    event_ids: list[int] | None = None,
    points: list[str] | None = None,
    output: str = "all",
    export: bool = False,
) -> ToolResult:
    precheck = _require_event_ids(deps, event_ids)
    if precheck:
        return precheck
    event_ids = event_ids or deps.session.selected_event_ids
    source_event_ids = _source_event_ids(deps, event_ids or [])
    params = {"event_ids": source_event_ids, "points": points or [], "output": output, "export": export}

    flow, events, covered, excluded = _event_data_coverage(deps, source_event_ids, points)
    coverage_failure = _coverage_guard_result(deps, source_event_ids, covered, excluded)
    if coverage_failure:
        return coverage_failure

    def work() -> tuple[str, dict[str, Any]]:
        dry_flow = _load_filtered_dry_flow(deps, points=covered)
        dry_curves = build_dry_curves(dry_flow)
        _save_curves(deps, dry_curves)
        result = analyze_rdii(flow, dry_curves, events, source_event_ids)
        table = result["rdii_total"]
        public_table = _public_event_frame(deps, table)
        _save_rdii_curves(deps, result["rdii_curve_data"])
        if table.empty:
            deps.session.unavailable_event_ids = sorted(
                set(deps.session.unavailable_event_ids).union(event_ids or [])
            )
            summary = (
                f"RDII 分析无可用数据：场次 {event_ids} 与点位 {points or ['全部点位']} 的监测数据"
                "无时间重叠，无法计算 RDII。"
            )
            destination = _route_table_result(deps, public_table, "RDII总量统计", points, export)
            return summary, {
                "table": [],
                "chart_paths": {},
                "no_data": True,
                "event_ids": event_ids,
                "result_destinations": [destination],
            }
        rain = io.load_rain(root=deps.paths.root)
        if is_full_network(points, deps):
            chart_paths = _save_rdii_curve_pngs(
                result["rdii_curve_data"],
                rain,
                events,
                deps.paths.outputs,
                selected_events=source_event_ids,
            )
        elif export:
            chart_paths = _save_partial_rdii_curve_png(
                result["rdii_curve_data"], points, deps.paths.outputs, source_event_ids
            )
        else:
            chart_paths = {}
        destination = _route_table_result(deps, public_table, "RDII总量统计", points, export)
        chart_count = sum(len(point_paths) for point_paths in chart_paths.values())
        excluded_note = f"；剔除无覆盖点位 {[item['point_id'] for item in excluded]}" if excluded else ""
        summary = f"RDII 分析完成：场次 {event_ids}，输出 {len(table)} 行统计，生成 {chart_count} 张 RDII 曲线图{excluded_note}。"
        return summary, {
            "table": public_table.to_dict(orient="records"),
            "chart_paths": chart_paths,
            "no_data": False,
            "covered_points": covered,
            "excluded_points": excluded,
            "result_destinations": [destination],
        }

    return _run(deps, "analyze_rdii", work, params=params)


def assess_risk_impl(
    deps: AgentDeps,
    scope: str = "all",
    event_ids: list[int] | None = None,
    points: list[str] | None = None,
    export: bool = False,
    start: str | None = None,
    end: str | None = None,
) -> ToolResult:
    scope = {"旱天": "dry", "雨天": "rainy", "全部": "all"}.get(scope, scope)
    if scope in {"rainy", "all"}:
        precheck = _require_event_ids(deps, event_ids)
        if precheck:
            return precheck
    event_ids = event_ids or deps.session.selected_event_ids
    source_event_ids = _source_event_ids(deps, event_ids or [])
    params = {
        "scope": scope,
        "event_ids": source_event_ids,
        "points": points or [],
        "start": start,
        "end": end,
        "export": export,
    }

    windowed = scope in {"dry", "all"} and (start is not None or end is not None)
    dry_window_flow = pd.DataFrame()
    dry_covered: list[str] = []
    dry_excluded: list[dict[str, str]] = []
    window_coverage: dict[str, str | None] | None = None
    if windowed:
        try:
            dry_window_flow, dry_covered, dry_excluded, window_coverage = _window_data_coverage(
                deps, points, start, end
            )
        except ValueError as exc:
            return error(str(exc))
        coverage_failure = _window_coverage_guard_result(dry_covered, dry_excluded, start, end)
        if coverage_failure:
            return coverage_failure

    flow = pd.DataFrame()
    events = pd.DataFrame()
    covered: list[str] = []
    excluded: list[dict[str, str]] = []
    if scope in {"rainy", "all"} and source_event_ids:
        flow, events, covered, excluded = _event_data_coverage(deps, source_event_ids, points)
        coverage_failure = _coverage_guard_result(deps, source_event_ids, covered, excluded)
        if coverage_failure:
            return coverage_failure

    def work() -> tuple[str, dict[str, Any]]:
        if windowed:
            dry_flow = dry_window_flow
            dry_stats = dry_statistics(dry_flow, io.load_sites(root=deps.paths.root))
        else:
            dry_flow, dry_stats, _ = _dry_inputs(deps, points=points)
        sites = io.load_sites(root=deps.paths.root)
        event_table = pd.DataFrame()
        if scope in {"rainy", "all"} and source_event_ids:
            event_table = analyze_event_response(flow, events, source_event_ids)
        result = assess_risk(
            dry_stats,
            event_table,
            scope=scope,
            sites=sites,
            flow=flow,
            events=events,
            event_ids=source_event_ids,
        )
        result["rainy_risk"] = _public_event_frame(deps, result["rainy_risk"])
        destinations = [
            _route_table_result(
                deps, dry_stats, "旱天分析", points, export, start=start, end=end
            )
        ]
        if not result["dry_risk"].empty:
            destinations.append(
                _route_table_result(
                    deps, result["dry_risk"], "旱天风险", points, export, start=start, end=end
                )
            )
        if not result["rainy_risk"].empty:
            destinations.append(
                _route_table_result(
                    deps,
                    result["rainy_risk"],
                    "雨天溢流风险",
                    points,
                    export,
                    start=start,
                    end=end,
                )
            )
        excluded_note = f"；雨天分析剔除无覆盖点位 {[item['point_id'] for item in excluded]}" if excluded else ""
        summary = f"风险评估完成：旱天风险 {len(result['dry_risk'])} 行，雨天风险 {len(result['rainy_risk'])} 行{excluded_note}。"
        if window_coverage is not None:
            summary += _window_coverage_note(window_coverage, dry_excluded)
        data = {key: df.to_dict(orient="records") for key, df in result.items()}
        data["dry_analysis"] = dry_stats.to_dict(orient="records")
        data["covered_points"] = covered
        data["excluded_points"] = excluded
        if windowed:
            data["window_coverage"] = window_coverage
            data["window_covered_points"] = dry_covered
            data["window_excluded_points"] = dry_excluded
        data["result_destinations"] = destinations
        return summary, data

    return _run(deps, "assess_risk", work, params=params)


DEFAULT_REPORT_SECTIONS = ["监测概况", "降雨分析", "旱天排污规律统计分析", "污水系统运行风险分析"]
REPORT_MONITORING_SECTIONS = {"监测概况", "数据概况", "概述与数据质量", "数据体检", "数据质量"}
REPORT_RAINFALL_SECTIONS = {"降雨分析", "降雨统计", "雨天事件统计", "事件响应", "RDII"}
REPORT_PATTERN_SECTIONS = {
    "旱天排污规律统计分析",
    "旱天排污规律",
    "旱天排污规律分析",
    "点位特征对比分析",
    "排污规律",
    "排污规律分析",
    "旱天分析",
}
REPORT_FULL_RISK_SECTIONS = {"污水系统运行风险分析", "污水系统运行风险", "运行风险分析", "风险评估"}
REPORT_DRY_RISK_SECTIONS = {"旱天风险", "旱天运行风险评估", "结论与建议"}
REPORT_RAINY_RISK_SECTIONS = {"雨天风险", "雨天溢流风险", "溢流风险"}
REPORT_RISK_SECTIONS = REPORT_FULL_RISK_SECTIONS | REPORT_DRY_RISK_SECTIONS | REPORT_RAINY_RISK_SECTIONS


def _section_requested(sections: list[str], aliases: set[str]) -> bool:
    return any(
        section == alias or section.startswith(f"{alias}（") or section.startswith(f"{alias}(")
        for section in sections
        for alias in aliases
    )


def _is_dry_only_report_sections(sections: list[str]) -> bool:
    wants_any_dry = _section_requested(sections, REPORT_PATTERN_SECTIONS | REPORT_DRY_RISK_SECTIONS)
    wants_rain = _section_requested(sections, REPORT_RAINFALL_SECTIONS | REPORT_RAINY_RISK_SECTIONS)
    wants_full_risk = _section_requested(sections, REPORT_FULL_RISK_SECTIONS)
    return wants_any_dry and not wants_rain and not wants_full_risk


def _report_actual_time_range(
    deps: AgentDeps,
    points: list[str] | None,
    start: str | None,
    end: str | None,
    *,
    prefer_dry_flow: bool,
) -> tuple[str | None, str | None]:
    try:
        if prefer_dry_flow:
            flow = _load_filtered_dry_flow(deps, points=points)
        else:
            flow = io.load_flow(points=points, root=deps.paths.root)
        start_ts, end_ts = _window_bounds(start, end)
    except Exception:
        return start, end
    if start_ts is not None:
        flow = flow[flow["timestamp"] >= start_ts]
    if end_ts is not None:
        flow = flow[flow["timestamp"] <= end_ts]
    if flow.empty:
        return start, end
    actual_start = flow["timestamp"].min()
    actual_end = flow["timestamp"].max()
    return actual_start.isoformat(sep=" "), actual_end.isoformat(sep=" ")


def generate_report_impl(
    deps: AgentDeps,
    points: list[str] | None = None,
    start: str | None = None,
    end: str | None = None,
    sections: list[str] | None = None,
    event_ids: list[int] | None = None,
) -> ToolResult:
    points = _normalize_point_scope(points, deps)
    sections = sections or list(DEFAULT_REPORT_SECTIONS)
    selected_event_ids = list(event_ids or deps.session.selected_event_ids)
    unavailable = sorted(set(selected_event_ids).intersection(deps.session.unavailable_event_ids))
    if unavailable:
        return error(
            f"无法生成可靠报告：场次 {unavailable} 与监测数据无时间重叠，缺少事件响应、RDII 和雨天风险依据。"
        )

    wants_monitoring = _section_requested(sections, REPORT_MONITORING_SECTIONS)
    wants_rainfall = _section_requested(sections, REPORT_RAINFALL_SECTIONS)
    wants_patterns = _section_requested(sections, REPORT_PATTERN_SECTIONS)
    wants_full_risk = _section_requested(sections, REPORT_FULL_RISK_SECTIONS)
    wants_dry_risk = wants_full_risk or _section_requested(sections, REPORT_DRY_RISK_SECTIONS)
    wants_rainy_risk = wants_full_risk or _section_requested(sections, REPORT_RAINY_RISK_SECTIONS)
    wants_risk = wants_dry_risk or wants_rainy_risk
    dry_only_report = _is_dry_only_report_sections(sections)
    if dry_only_report:
        wants_monitoring = True
    if not any((wants_monitoring, wants_rainfall, wants_patterns, wants_risk)):
        return error(f"无法识别报告章节: {sections}")
    tables: dict[str, pd.DataFrame] = {}
    rainfall_chart_paths: dict[str, str] = {}
    pattern_chart_paths: dict[str, list[str]] = {}
    summaries: list[str] = []
    time_range = _resolved_report_time_range(deps, start, end) if start is not None or end is not None else None
    report_start, report_end = _report_actual_time_range(
        deps,
        points,
        start,
        end,
        prefer_dry_flow=wants_patterns or wants_dry_risk,
    )
    report_sections = list(sections)
    if dry_only_report and not _section_requested(report_sections, REPORT_MONITORING_SECTIONS):
        report_sections.insert(0, "监测概况")

    if wants_monitoring:
        cached = _cached_report_frame(deps, "data_collection", points, start, end)
        if cached is None:
            data_check = check_data_impl(deps, points=points, start=start, end=end)
            if data_check["status"] != "ok":
                return data_check
            cached = _result_frame(data_check, "table")
            summaries.append(data_check["summary"])
        tables["data_collection"] = cached

    rain: ToolResult | None = None
    if wants_rainfall or wants_rainy_risk:
        rain_start = time_range[0] if time_range else None
        rain_end = time_range[1] if time_range else None
        cached_daily = _cached_report_frame(deps, "rainfall_daily", start=rain_start, end=rain_end)
        cached_events = _cached_report_frame(deps, "rainfall_events", start=rain_start, end=rain_end)
        rainfall_chart_paths = _cached_report_value(
            deps, "rainfall_chart_paths", start=rain_start, end=rain_end
        ) or {}
        if cached_daily is None or cached_events is None:
            rain = analyze_rainfall_impl(deps, time_range=time_range)
            if rain["status"] != "ok":
                return rain
            cached_daily = _result_frame(rain, "daily")
            cached_events = _result_frame(rain, "events")
            rainfall_chart_paths = deepcopy(rain.get("data", {}).get("chart_paths", {}))
            summaries.append(rain["summary"])
        tables["rainfall_daily"] = cached_daily
        tables["rainfall_events"] = cached_events

    if wants_patterns:
        cached = _cached_report_frame(deps, "pattern_analysis", points, start, end)
        pattern_chart_paths = _cached_report_value(
            deps, "pattern_chart_paths", points, start, end
        ) or {}
        if cached is None or not pattern_chart_paths:
            patterns = analyze_patterns_impl(
                deps, points=points, start=start, end=end, report_charts=True
            )
            if patterns["status"] != "ok":
                return patterns
            cached = _result_frame(patterns, "table")
            pattern_chart_paths = deepcopy(patterns.get("data", {}).get("curve_images", {}))
            summaries.append(patterns["summary"])
        tables["pattern_analysis"] = cached

    if wants_risk:
        window_events = tables.get("rainfall_events", pd.DataFrame())
        event_id_column = "source_event_id" if time_range and "source_event_id" in window_events.columns else "event_id"
        available_ids = (
            set(pd.to_numeric(window_events.get(event_id_column), errors="coerce").dropna().astype(int).tolist())
            if not window_events.empty and event_id_column in window_events.columns
            else set()
        )
        if wants_rainy_risk and not selected_event_ids:
            return needs_input(
                "event_ids",
                "请选择报告雨天风险所使用的降雨场次编号。",
                summary="默认全套报告包含雨天风险，需要先选择降雨场次。",
                options=_event_options(window_events),
            )
        risk_event_ids = list(selected_event_ids)
        public_event_ids = list(selected_event_ids)
        source_to_local: dict[int, int] = {}
        if time_range and "source_event_id" in window_events.columns:
            local_to_source = {
                int(local): int(source)
                for local, source in zip(window_events["event_id"], window_events["source_event_id"])
            }
            source_to_local = {source: local for local, source in local_to_source.items()}
            selected_set = set(selected_event_ids)
            if selected_set and not selected_set.issubset(available_ids) and selected_set.issubset(local_to_source):
                risk_event_ids = [local_to_source[event_id] for event_id in selected_event_ids]
            public_event_ids = [source_to_local.get(event_id, event_id) for event_id in risk_event_ids]
        outside = sorted(set(risk_event_ids) - available_ids)
        if wants_rainy_risk and time_range and outside:
            return error(f"降雨场次 {outside} 不在报告时间窗 [{start or '不限'}, {end or '不限'}] 内。")
        dry_analysis = _cached_report_frame(deps, "dry_analysis", points, start, end) if wants_dry_risk else pd.DataFrame()
        dry_risk = _cached_report_frame(deps, "dry_risk", points, start, end) if wants_dry_risk else pd.DataFrame()
        rainy_risk = (
            _cached_report_frame(deps, "rainy_overflow_risk", points, start, end, risk_event_ids)
            if wants_rainy_risk
            else pd.DataFrame()
        )
        missing_dry = wants_dry_risk and (dry_analysis is None or dry_risk is None)
        missing_rainy = wants_rainy_risk and rainy_risk is None
        if missing_dry or missing_rainy:
            scope = "all" if missing_dry and missing_rainy else "dry" if missing_dry else "rainy"
            risk = assess_risk_impl(
                deps,
                scope=scope,
                event_ids=risk_event_ids if scope in {"all", "rainy"} else None,
                points=points,
                start=start,
                end=end,
            )
            if risk["status"] != "ok":
                return risk
            if missing_dry:
                dry_analysis = _result_frame(risk, "dry_analysis")
                dry_risk = _result_frame(risk, "dry_risk")
            if missing_rainy:
                rainy_risk = _result_frame(risk, "rainy_risk")
            summaries.append(risk["summary"])
        if wants_dry_risk:
            tables["dry_analysis"] = dry_analysis if dry_analysis is not None else pd.DataFrame()
            tables["dry_risk"] = dry_risk if dry_risk is not None else pd.DataFrame()
        if wants_rainy_risk:
            public_rainy_risk = rainy_risk.copy() if rainy_risk is not None else pd.DataFrame()
            if source_to_local and "event_id" in public_rainy_risk.columns:
                public_rainy_risk["event_id"] = public_rainy_risk["event_id"].map(
                    lambda value: source_to_local.get(int(value), int(value)) if pd.notna(value) else value
                )
            tables["rainy_overflow_risk"] = public_rainy_risk
        if wants_rainy_risk and tables["rainy_overflow_risk"].empty:
            return error("雨天风险计算结果为空，拒绝生成带空雨天风险章节的报告。")

    params = {
        "points": points or [],
        "start": start,
        "end": end,
        "sections": report_sections,
        "event_ids": public_event_ids if wants_risk else selected_event_ids,
    }

    def work() -> tuple[str, dict[str, Any]]:
        output_file = deps.paths.outputs / _report_filename(points, deps, start, end)
        combined_file = _report_combined_workbook_path(output_file)
        result = build_report(
            output_file,
            "排水监测数据分析报告",
            summaries,
            template_file=deps.paths.report_template_file,
            analysis_tables=tables,
            site_info_file=deps.paths.site_info_file,
            outputs_dir=deps.paths.outputs,
            sections=report_sections,
            has_rainfall_data=not tables.get("rainfall_daily", pd.DataFrame()).empty,
            point_ids=points,
            start=report_start,
            end=report_end,
            rainfall_chart_paths=rainfall_chart_paths,
            pattern_chart_paths=pattern_chart_paths,
            artifact_scope=_range_result_prefix(points, deps, start, end),
        )
        combined_sheets = _write_report_combined_workbook(deps, tables, combined_file)
        result["report_combined_sheets"] = combined_sheets
        if combined_sheets:
            result["result_destinations"] = [
                {
                    "kind": "combined_xlsx",
                    "path": _rel(deps, combined_file),
                    "sheet": None,
                }
            ]
        summary = (
            f"报告生成完成：{_rel(deps, output_file)}，范围点位 {points or ['全网']}，"
            f"时间窗 [{start or '全时段'}, {end or '全时段'}]。"
        )
        if report_start or report_end:
            summary += f" 报告正文按实际有效数据范围 [{report_start or '不限'}, {report_end or '不限'}] 填充。"
        if wants_rainy_risk:
            summary += f" 窗口内降雨场次编号 {public_event_ids}。"
        return summary, result

    return _run(deps, "generate_report", work, params=params, use_cache=False)


def _result_frame(result: ToolResult, key: str) -> pd.DataFrame:
    value = result.get("data", {}).get(key, [])
    return pd.DataFrame(value)


def _resolved_report_time_range(deps: AgentDeps, start: str | None, end: str | None) -> list[str]:
    rain = io.load_rain(root=deps.paths.root)
    resolved_start = start or rain["timestamp"].min().strftime("%Y-%m-%d %H:%M:%S")
    resolved_end = end or rain["timestamp"].max().strftime("%Y-%m-%d %H:%M:%S")
    return [resolved_start, resolved_end]


def _report_filename(
    points: list[str] | None,
    deps: AgentDeps,
    start: str | None,
    end: str | None,
) -> str:
    points = _normalize_point_scope(points, deps)
    if points is None:
        point_part = "全网"
    elif len(points) == 1:
        point_part = _safe_filename_part(points[0])[:24]
    else:
        first = _safe_filename_part(sorted(points, key=str)[0])[:12]
        point_part = f"{len(points)}点_{first}等"
    filename = f"{point_part}_{_time_result_prefix(start, end)}_分析报告.docx"
    if len(filename) > 80:
        filename = f"{point_part[:20]}_{_time_result_prefix(start, end)[:36]}_分析报告.docx"
    return filename
