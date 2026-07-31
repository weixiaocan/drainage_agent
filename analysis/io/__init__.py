from __future__ import annotations

import os
from pathlib import Path
from typing import Iterable

import pandas as pd

from analysis.schema import normalize_flow_df, normalize_rain_df
from analysis.io.standard import StandardDataStore, StandardDataUnavailable


def project_root() -> Path:
    env_root = os.getenv("DRAINAGE_AGENT_ROOT")
    if env_root:
        return Path(env_root).resolve()
    current = Path.cwd().resolve()
    for candidate in [current, *current.parents]:
        if ((candidate / "resources" / "data").exists() or (candidate / "data").exists()) and (candidate / "agent").exists():
            return candidate
    return current


def _data_dir(base: Path) -> Path:
    current = base / "resources" / "data"
    return current if current.exists() else base / "data"


def _outputs_dir(base: Path) -> Path:
    current = base / "var" / "outputs"
    return current if current.exists() else base / "outputs"


def _read_csv(path: Path) -> pd.DataFrame:
    last_exc: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError as exc:
            last_exc = exc
    if last_exc:
        raise last_exc
    return pd.read_csv(path)


def _normalize_points(points: Iterable[str] | str | None) -> set[str] | None:
    if points is None or points == "全部":
        return None
    if isinstance(points, str):
        return {points}
    return {str(p) for p in points}


def _apply_time_range(df: pd.DataFrame, time_range: tuple[str, str] | list[str] | None) -> pd.DataFrame:
    if not time_range:
        return df
    start, end = time_range
    result = df
    if start:
        result = result[result["timestamp"] >= pd.to_datetime(start)]
    if end:
        result = result[result["timestamp"] <= pd.to_datetime(end)]
    return result.copy()


def load_rain(time_range: tuple[str, str] | list[str] | None = None, root: Path | None = None) -> pd.DataFrame:
    base = root or project_root()
    standard_path = base / "standard" / "rainfall.csv"
    path = standard_path if standard_path.is_file() else _data_dir(base) / "降雨数据.csv"
    if not path.exists():
        return pd.DataFrame(columns=["timestamp", "rain_mm"])
    rain = normalize_rain_df(_read_csv(path))
    return _apply_time_range(rain, time_range)


def load_flow(
    points: Iterable[str] | str | None = None,
    time_range: tuple[str, str] | list[str] | None = None,
    root: Path | None = None,
) -> pd.DataFrame:
    base = root or project_root()
    standard_path = base / "standard" / "flow.csv"
    if standard_path.is_file():
        flow = pd.read_csv(
            standard_path,
            dtype={"device_id": "string", "point_id": "string"},
        )
        flow["timestamp"] = pd.to_datetime(flow["timestamp"], errors="coerce")
        flow = flow.dropna(subset=["timestamp"])
        selected = _normalize_points(points)
        if selected is not None:
            flow = flow[flow["point_id"].astype(str).isin(selected)]
        return _apply_time_range(flow, time_range).reset_index(drop=True)
    flow_dir = _data_dir(base) / "flow"
    selected_points = _normalize_points(points)
    frames: list[pd.DataFrame] = []
    for csv_path in sorted(flow_dir.glob("*.csv")):
        normalized = normalize_flow_df(_read_csv(csv_path), csv_path)
        if selected_points is not None and str(normalized["point_id"].iloc[0]) not in selected_points:
            continue
        frames.append(normalized)
    if not frames:
        return pd.DataFrame(columns=["timestamp", "device_id", "point_id", "flow_lps", "level_m", "velocity_mps"])
    flow = pd.concat(frames, ignore_index=True)
    return _apply_time_range(flow, time_range).reset_index(drop=True)


def read_selected_days(filter_result: Path) -> dict[str, set[object]]:
    if not filter_result.exists():
        return {}
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(filter_result, data_only=True)
        sheet = workbook["筛选结果"]
    except Exception:
        return {}

    date_cols: list[tuple[int, object]] = []
    for col in range(2, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value and str(value) != "筛选说明":
            date_cols.append((col, pd.to_datetime(value).date()))

    selected: dict[str, set[object]] = {}
    for row in range(3, sheet.max_row + 1):
        point_id = str(sheet.cell(row=row, column=1).value or "").strip()
        if not point_id or point_id.lower() == "nan":
            continue
        days: set[object] = set()
        if "_" in point_id:
            point_id = point_id.split("_", 1)[1]
        for col, day in date_cols:
            fill = sheet.cell(row=row, column=col).fill
            color = str(fill.start_color.index).upper() if fill and fill.start_color else ""
            if color.endswith("92D050"):
                days.add(day)
        selected[point_id] = days
    return selected


def load_flow_by_filter_result(
    filter_result: Path,
    points: Iterable[str] | str | None = None,
    time_range: tuple[str, str] | list[str] | None = None,
    root: Path | None = None,
) -> pd.DataFrame:
    selected_days = read_selected_days(filter_result)
    if not selected_days:
        return pd.DataFrame(columns=["timestamp", "device_id", "point_id", "flow_lps", "level_m", "velocity_mps"])

    flow = load_flow(points=points, time_range=time_range, root=root)
    if flow.empty:
        return flow
    frames: list[pd.DataFrame] = []
    for point_id, point_df in flow.groupby("point_id", sort=True):
        point_df = point_df.sort_values("timestamp").copy()
        full_index = pd.date_range(point_df["timestamp"].min(), point_df["timestamp"].max(), freq="min")
        full_df = pd.DataFrame({"timestamp": full_index})
        merged = full_df.merge(point_df, on="timestamp", how="left")
        merged["point_id"] = str(point_id)
        merged["device_id"] = merged["device_id"].ffill().bfill()
        for col in ("flow_lps", "level_m", "velocity_mps"):
            merged[col] = pd.to_numeric(merged[col], errors="coerce").interpolate(method="linear").fillna(0.0)
        frames.append(merged)
    df = pd.concat(frames, ignore_index=True) if frames else flow.copy()
    df["date"] = df["timestamp"].dt.date
    result = df[
        df.apply(
            lambda row: row["date"] in selected_days.get(str(row["point_id"]), set()),
            axis=1,
        )
    ].drop(columns=["date"])
    return result.reset_index(drop=True)


def load_filtered_flow(
    points: Iterable[str] | str | None = None,
    time_range: tuple[str, str] | list[str] | None = None,
    root: Path | None = None,
) -> pd.DataFrame:
    """读取 data_filter 标准产物中选定的有效旱天数据。"""
    base = root or project_root()
    return load_flow_by_filter_result(
        _outputs_dir(base) / "筛选结果.xlsx",
        points=points,
        time_range=time_range,
        root=base,
    )


def load_sites(root: Path | None = None) -> pd.DataFrame:
    base = root or project_root()
    standard_path = base / "standard" / "sites.csv"
    if standard_path.is_file():
        return pd.read_csv(standard_path, dtype={"point_id": "string"})
    path = _data_dir(base) / "点位信息.xlsx"
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(path)
    except Exception:
        return pd.DataFrame()

