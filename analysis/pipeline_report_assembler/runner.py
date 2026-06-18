from __future__ import annotations

"""报告组装模块入口

统一接口: run(config: Config, logger, dry_curve_data=None) -> dict

输出:
    - config.report_output_path（分析报告.docx）
    - 返回值: {output_file, stats}
"""

import logging
from pathlib import Path
from typing import Any, TYPE_CHECKING

import pandas as pd
from openpyxl import load_workbook

from .assembler import run_report_assembler

if TYPE_CHECKING:
    from typing import Protocol

    class Config(Protocol):
        report_template_path: Path
        combined_xlsx_path: Path
        site_info_path: Path
        filter_result_path: Path
        report_output_path: Path
        baseinfo_path: Path


def _load_dry_curve_data_from_excel(combined_xlsx: Path, logger: logging.Logger) -> dict[str, pd.DataFrame]:
    """从 Excel 读取旱天特征曲线数据"""
    dry_curve_data: dict[str, pd.DataFrame] = {}

    try:
        wb = load_workbook(combined_xlsx, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("特征曲线_"):
                ws = wb[sheet_name]
                point_name = sheet_name.replace("特征曲线_", "")

                data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:
                        data.append(row)

                if data:
                    df = pd.DataFrame(data, columns=["时间", "流量(L/s)", "液位(m)", "流速(m/s)"])
                    df = df.dropna(subset=["时间"])
                    df["时间"] = pd.date_range("00:00:00", "23:59:00", freq="min")[:len(df)]
                    df = df.set_index("时间")
                    df = df.rename(columns={"流量(L/s)": "f", "液位(m)": "l", "流速(m/s)": "velo"})
                    dry_curve_data[point_name] = df

        wb.close()

    except Exception as e:
        logger.warning(f"读取旱天特征曲线数据失败: {e}")

    return dry_curve_data


def run(
    config: Config,
    logger: logging.Logger,
    dry_curve_data: dict[str, pd.DataFrame] | None = None,
    has_rainfall_data: bool = True,
    llm_client=None,
) -> dict[str, Any]:
    """
    报告组装入口。

    输入:
        - 报告模板（从 config.report_template_path）
        - 综合分析结果（从 config.combined_xlsx_path）
        - 点位信息（从 config.site_info_path）
        - 筛选结果（从 config.filter_result_path）
        - 旱天特征曲线数据（从内存传入，或从 Excel 读取）
        - has_rainfall_data: 是否有降雨数据（无则跳过雨天章节）
        - llm_client: LLM客户端（用于生成风险分析段落）

    输出:
        - config.report_output_path（分析报告.docx）

    返回:
        {
            "output_file": Path,
            "stats": dict,
        }
    """
    template_file = config.report_template_path
    combined_xlsx = config.combined_xlsx_path
    site_info_file = config.site_info_path
    filter_result_path = config.filter_result_path
    output_file = config.report_output_path

    logger.info(f"开始报告组装")
    logger.info(f"  报告模板: {template_file}")
    logger.info(f"  综合分析结果: {combined_xlsx}")
    logger.info(f"  点位信息: {site_info_file}")
    logger.info(f"  筛选结果: {filter_result_path}")
    logger.info(f"  输出文件: {output_file}")
    logger.info(f"  降雨数据: {'有' if has_rainfall_data else '无'}")
    logger.info(f"  LLM客户端: {'有' if llm_client else '无'}")

    # 如果没有传入 dry_curve_data，从 Excel 读取
    if dry_curve_data is None:
        logger.info("  从 Excel 读取旱天特征曲线数据...")
        dry_curve_data = _load_dry_curve_data_from_excel(combined_xlsx, logger)

    # 执行组装
    result = run_report_assembler(
        template_file=template_file,
        combined_xlsx=combined_xlsx,
        site_info_file=site_info_file,
        output_file=output_file,
        dry_curve_data=dry_curve_data,
        filter_result_path=filter_result_path,
        config={"baseinfo_path": str(config.baseinfo_path)},
        has_rainfall_data=has_rainfall_data,
        llm_client=llm_client,
    )

    stats = result["stats"]
    logger.info(f"报告组装完成")
    logger.info(f"  填充表格: {stats['tables_filled']} 个")
    logger.info(f"  插入图片: {stats.get('images_inserted', 0)} 张")
    logger.info(f"  处理点位: {stats['points_processed']} 个")
    logger.info(f"  文字替换: {stats.get('text_replaced', 0)} 处")
    logger.info(f"  LLM生成: {stats.get('llm_generated', 0)} 段")

    return result
