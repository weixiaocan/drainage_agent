from __future__ import annotations

import hashlib
import io
import json
import shutil
import sqlite3
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from analysis.io import StandardDataStore, StandardDataUnavailable, read_selected_days
from analysis.modules.filtering import FilterConfig, run_data_filter


class BaselinePreconditionError(ValueError):
    """Raised when filtering or baseline use requires another user action."""


@dataclass(frozen=True)
class FilterRequest:
    project_id: str
    batch_id: str
    missing_rate_threshold: float = 0.1
    expected_rows_per_day: int = 1440
    rain_day_filter_threshold: float = 2.0
    zero_like_threshold: float = 0.02
    high_zero_ratio_threshold: float = 0.5
    high_zero_ratio_normal_days_threshold: int = 5
    zero_day_drop_min_nonzero_keep_days: int = 3
    mean_lower_ratio: float = 0.5
    mean_upper_ratio: float = 2.0


@dataclass(frozen=True)
class FilterResult:
    filter_id: str
    project_id: str
    batch_id: str
    version: int
    status: str
    identity: dict[str, Any]
    summary: dict[str, Any]
    artifact: str
    created_at: str


@dataclass(frozen=True)
class AnalysisBaseline:
    baseline_id: str
    filter_id: str
    project_id: str
    batch_id: str
    version: int
    identity: dict[str, Any]
    artifact: str
    confirmed_at: str


class FilterBaselineService:
    """Shared Web and Agent boundary for filter confirmation and baselines."""

    def __init__(self, database: Path, files_root: Path) -> None:
        self.database = database
        self.files_root = files_root.resolve()
        self.standard_data = StandardDataStore(self.files_root)
        self.database.parent.mkdir(parents=True, exist_ok=True)
        with self._connect() as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS filter_results (
                    filter_id TEXT PRIMARY KEY,
                    project_id TEXT NOT NULL,
                    batch_id TEXT NOT NULL,
                    version INTEGER NOT NULL,
                    status TEXT NOT NULL,
                    standard_sha256 TEXT NOT NULL,
                    parameters_json TEXT NOT NULL,
                    file_sha256 TEXT NOT NULL,
                    result_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    UNIQUE (project_id, batch_id, version)
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS analysis_baselines (
                    baseline_id TEXT PRIMARY KEY,
                    filter_id TEXT NOT NULL,
                    project_id TEXT NOT NULL,
                    batch_id TEXT NOT NULL,
                    version INTEGER NOT NULL,
                    identity_json TEXT NOT NULL,
                    artifact TEXT NOT NULL,
                    confirmed_at TEXT NOT NULL,
                    UNIQUE (project_id, batch_id, version),
                    FOREIGN KEY (filter_id) REFERENCES filter_results (filter_id)
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS current_analysis_baselines (
                    project_id TEXT NOT NULL,
                    batch_id TEXT NOT NULL,
                    baseline_id TEXT NOT NULL,
                    PRIMARY KEY (project_id, batch_id),
                    FOREIGN KEY (baseline_id)
                        REFERENCES analysis_baselines (baseline_id)
                )
                """
            )

    def run_filter(self, request: FilterRequest) -> FilterResult:
        self._require_batch(request.project_id, request.batch_id)
        try:
            flow = self.standard_data.load_flow(
                request.project_id, request.batch_id
            )
        except StandardDataUnavailable as exc:
            raise BaselinePreconditionError(
                f"{exc}；请先导入并确认当前分析批次的标准数据"
            ) from exc
        flow["timestamp"] = pd.to_datetime(flow["timestamp"])

        standard_sha256 = self._standard_sha256(
            request.project_id, request.batch_id
        )
        parameters = self._parameters(request)
        version = self._next_filter_version(request.project_id, request.batch_id)
        filter_id = uuid.uuid4().hex
        artifact = f"baseline/filters/{version}-{filter_id}/filter_result.xlsx"
        artifact_path = (
            self._batch_root(request.project_id, request.batch_id) / artifact
        )
        selected = run_data_filter(
            flow,
            pd.DataFrame(columns=["timestamp", "rain_mm"]),
            artifact_path,
            FilterConfig(**parameters),
        )
        file_sha256 = self._sha256(artifact_path)
        identity = {
            "project_id": request.project_id,
            "batch_id": request.batch_id,
            "standard_input": {
                "contract_version": 1,
                "content_sha256": standard_sha256,
            },
            "parameters": parameters,
            "file_sha256": file_sha256,
        }
        summary = {
            "point_count": len(selected),
            "selected_point_days": sum(len(days) for days in selected.values()),
            "selected_days": selected,
            "exclusion_reasons": self._read_filter_notes(artifact_path),
        }
        created_at = datetime.now(timezone.utc).isoformat()
        result = FilterResult(
            filter_id=filter_id,
            project_id=request.project_id,
            batch_id=request.batch_id,
            version=version,
            status="awaiting_confirmation",
            identity=identity,
            summary=summary,
            artifact=artifact,
            created_at=created_at,
        )
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO filter_results (
                    filter_id, project_id, batch_id, version, status,
                    standard_sha256, parameters_json, file_sha256,
                    result_json, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    filter_id,
                    request.project_id,
                    request.batch_id,
                    version,
                    result.status,
                    standard_sha256,
                    json.dumps(parameters, sort_keys=True),
                    file_sha256,
                    json.dumps(asdict(result), ensure_ascii=False),
                    created_at,
                ),
            )
            connection.execute(
                """
                DELETE FROM current_analysis_baselines
                WHERE project_id = ? AND batch_id = ?
                """,
                (request.project_id, request.batch_id),
            )
        return result

    def current_baseline(
        self, project_id: str, batch_id: str
    ) -> AnalysisBaseline | None:
        self._require_batch(project_id, batch_id)
        with self._connect() as connection:
            row = connection.execute(
                """
                SELECT baseline.identity_json, baseline.baseline_id,
                       baseline.filter_id, baseline.project_id,
                       baseline.batch_id, baseline.version,
                       baseline.artifact, baseline.confirmed_at
                FROM current_analysis_baselines AS current
                JOIN analysis_baselines AS baseline
                  ON baseline.baseline_id = current.baseline_id
                WHERE current.project_id = ? AND current.batch_id = ?
                """,
                (project_id, batch_id),
            ).fetchone()
        if row is None:
            return None
        identity = json.loads(row[0])
        artifact_path = self._batch_root(project_id, batch_id) / row[6]
        try:
            standard_sha256 = self._standard_sha256(project_id, batch_id)
        except BaselinePreconditionError:
            return None
        if (
            identity["standard_input"]["content_sha256"]
            != standard_sha256
            or not artifact_path.is_file()
            or identity["file_sha256"] != self._sha256(artifact_path)
        ):
            return None
        return AnalysisBaseline(
            baseline_id=row[1],
            filter_id=row[2],
            project_id=row[3],
            batch_id=row[4],
            version=row[5],
            identity=identity,
            artifact=row[6],
            confirmed_at=row[7],
        )

    def get_filter(
        self, project_id: str, batch_id: str, filter_id: str
    ) -> FilterResult | None:
        self._require_batch(project_id, batch_id)
        with self._connect() as connection:
            row = connection.execute(
                """
                SELECT status, result_json FROM filter_results
                WHERE project_id = ? AND batch_id = ? AND filter_id = ?
                """,
                (project_id, batch_id, filter_id),
            ).fetchone()
        if row is None:
            return None
        values = json.loads(row[1])
        values["status"] = row[0]
        return FilterResult(**values)

    def list_filters(
        self, project_id: str, batch_id: str
    ) -> list[FilterResult]:
        self._require_batch(project_id, batch_id)
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT status, result_json FROM filter_results
                WHERE project_id = ? AND batch_id = ?
                ORDER BY version DESC
                """,
                (project_id, batch_id),
            ).fetchall()
        results: list[FilterResult] = []
        for status, serialized in rows:
            values = json.loads(serialized)
            values["status"] = status
            results.append(FilterResult(**values))
        return results

    def upload_revision(
        self,
        project_id: str,
        batch_id: str,
        source_filter_id: str,
        filename: str,
        content: bytes,
    ) -> FilterResult:
        if Path(filename).suffix.lower() != ".xlsx":
            raise ValueError("筛选文件只允许 xlsx")
        source = self.get_filter(project_id, batch_id, source_filter_id)
        if source is None:
            raise LookupError("筛选结果不存在")
        if source.version != self._latest_filter_version(project_id, batch_id):
            raise BaselinePreconditionError("筛选结果已过期，请使用当前最新版本")
        if (
            source.identity["standard_input"]["content_sha256"]
            != self._standard_sha256(project_id, batch_id)
        ):
            raise BaselinePreconditionError("筛选结果已过期，请重新运行自动筛选")
        self._validate_workbook(project_id, batch_id, content)

        version = self._next_filter_version(project_id, batch_id)
        filter_id = uuid.uuid4().hex
        artifact = f"baseline/filters/{version}-{filter_id}/filter_result.xlsx"
        artifact_path = self._batch_root(project_id, batch_id) / artifact
        artifact_path.parent.mkdir(parents=True, exist_ok=False)
        artifact_path.write_bytes(content)
        file_sha256 = self._sha256(artifact_path)
        identity = {
            "project_id": project_id,
            "batch_id": batch_id,
            "standard_input": source.identity["standard_input"],
            "parameters": source.identity["parameters"],
            "file_sha256": file_sha256,
        }
        selected = {
            point_id: sorted(day.isoformat() for day in days)
            for point_id, days in read_selected_days(artifact_path).items()
        }
        result = FilterResult(
            filter_id=filter_id,
            project_id=project_id,
            batch_id=batch_id,
            version=version,
            status="awaiting_confirmation",
            identity=identity,
            summary={
                "point_count": len(selected),
                "selected_point_days": sum(len(days) for days in selected.values()),
                "selected_days": selected,
                "exclusion_reasons": self._read_filter_notes(artifact_path),
            },
            artifact=artifact,
            created_at=datetime.now(timezone.utc).isoformat(),
        )
        self._insert_filter(result)
        with self._connect() as connection:
            connection.execute(
                """
                DELETE FROM current_analysis_baselines
                WHERE project_id = ? AND batch_id = ?
                """,
                (project_id, batch_id),
            )
        return result

    def confirm(
        self, project_id: str, batch_id: str, filter_id: str
    ) -> AnalysisBaseline:
        candidate = self.get_filter(project_id, batch_id, filter_id)
        if candidate is None:
            raise LookupError("筛选结果不存在")
        if candidate.version != self._latest_filter_version(
            project_id, batch_id
        ):
            raise BaselinePreconditionError("筛选结果已过期，请确认当前最新版本")
        candidate_path = (
            self._batch_root(project_id, batch_id) / candidate.artifact
        )
        if (
            candidate.identity["standard_input"]["content_sha256"]
            != self._standard_sha256(project_id, batch_id)
            or not candidate_path.is_file()
            or candidate.identity["file_sha256"] != self._sha256(candidate_path)
        ):
            raise BaselinePreconditionError("筛选结果已过期，请重新运行自动筛选")

        version = self._next_baseline_version(project_id, batch_id)
        baseline_id = uuid.uuid4().hex
        artifact = (
            f"baseline/versions/{version}-{baseline_id}/analysis_baseline.xlsx"
        )
        baseline_path = self._batch_root(project_id, batch_id) / artifact
        baseline_path.parent.mkdir(parents=True, exist_ok=False)
        shutil.copyfile(candidate_path, baseline_path)
        bound_identity = {
            "project_id": project_id,
            "batch_id": batch_id,
            "standard_input": candidate.identity["standard_input"],
            "parameters": candidate.identity["parameters"],
            "file_sha256": self._sha256(baseline_path),
        }
        identity_digest = hashlib.sha256(
            json.dumps(
                bound_identity, sort_keys=True, separators=(",", ":")
            ).encode()
        ).hexdigest()
        identity = {
            "kind": "confirmed_filter",
            "identity": identity_digest,
            **bound_identity,
        }
        confirmed_at = datetime.now(timezone.utc).isoformat()
        baseline = AnalysisBaseline(
            baseline_id=baseline_id,
            filter_id=filter_id,
            project_id=project_id,
            batch_id=batch_id,
            version=version,
            identity=identity,
            artifact=artifact,
            confirmed_at=confirmed_at,
        )
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO analysis_baselines (
                    baseline_id, filter_id, project_id, batch_id, version,
                    identity_json, artifact, confirmed_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    baseline_id,
                    filter_id,
                    project_id,
                    batch_id,
                    version,
                    json.dumps(identity, ensure_ascii=False, sort_keys=True),
                    artifact,
                    confirmed_at,
                ),
            )
            connection.execute(
                """
                INSERT INTO current_analysis_baselines (
                    project_id, batch_id, baseline_id
                ) VALUES (?, ?, ?)
                ON CONFLICT (project_id, batch_id)
                DO UPDATE SET baseline_id = excluded.baseline_id
                """,
                (project_id, batch_id, baseline_id),
            )
            connection.execute(
                """
                UPDATE filter_results SET status = 'confirmed'
                WHERE filter_id = ?
                """,
                (filter_id,),
            )
        return baseline

    def load_flow(self, project_id: str, batch_id: str) -> pd.DataFrame:
        baseline = self.current_baseline(project_id, batch_id)
        if baseline is None:
            raise BaselinePreconditionError("当前已确认分析基线不可用")
        flow = self.standard_data.load_flow(project_id, batch_id)
        flow["timestamp"] = pd.to_datetime(flow["timestamp"])
        selected = read_selected_days(
            self._batch_root(project_id, batch_id) / baseline.artifact
        )
        dates = flow["timestamp"].dt.date
        keep = [
            date in selected.get(str(point_id), set())
            for point_id, date in zip(flow["point_id"], dates, strict=True)
        ]
        return flow.loc[keep].reset_index(drop=True)

    def artifact_path(
        self, project_id: str, batch_id: str, filter_id: str
    ) -> Path:
        candidate = self.get_filter(project_id, batch_id, filter_id)
        if candidate is None:
            raise LookupError("筛选结果不存在")
        path = self._batch_root(project_id, batch_id) / candidate.artifact
        if not path.is_file():
            raise FileNotFoundError(path)
        return path

    def baseline_artifact_path(
        self, project_id: str, batch_id: str
    ) -> Path:
        baseline = self.current_baseline(project_id, batch_id)
        if baseline is None:
            raise BaselinePreconditionError("当前已确认分析基线不可用")
        return self._batch_root(project_id, batch_id) / baseline.artifact

    def _require_batch(self, project_id: str, batch_id: str) -> None:
        with self._connect() as connection:
            row = connection.execute(
                """
                SELECT 1 FROM analysis_batches
                WHERE project_id = ? AND id = ?
                """,
                (project_id, batch_id),
            ).fetchone()
        if row is None:
            raise LookupError("分析批次不存在或不属于当前监测项目")

    def _next_filter_version(self, project_id: str, batch_id: str) -> int:
        return self._latest_filter_version(project_id, batch_id) + 1

    def _latest_filter_version(self, project_id: str, batch_id: str) -> int:
        with self._connect() as connection:
            row = connection.execute(
                """
                SELECT COALESCE(MAX(version), 0)
                FROM filter_results
                WHERE project_id = ? AND batch_id = ?
                """,
                (project_id, batch_id),
            ).fetchone()
        return int(row[0])

    def _next_baseline_version(self, project_id: str, batch_id: str) -> int:
        with self._connect() as connection:
            row = connection.execute(
                """
                SELECT COALESCE(MAX(version), 0) + 1
                FROM analysis_baselines
                WHERE project_id = ? AND batch_id = ?
                """,
                (project_id, batch_id),
            ).fetchone()
        return int(row[0])

    def _insert_filter(self, result: FilterResult) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                INSERT INTO filter_results (
                    filter_id, project_id, batch_id, version, status,
                    standard_sha256, parameters_json, file_sha256,
                    result_json, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    result.filter_id,
                    result.project_id,
                    result.batch_id,
                    result.version,
                    result.status,
                    result.identity["standard_input"]["content_sha256"],
                    json.dumps(result.identity["parameters"], sort_keys=True),
                    result.identity["file_sha256"],
                    json.dumps(asdict(result), ensure_ascii=False),
                    result.created_at,
                ),
            )

    def _standard_sha256(self, project_id: str, batch_id: str) -> str:
        path = self._batch_root(project_id, batch_id) / "standard" / "flow.csv"
        if not path.is_file():
            raise BaselinePreconditionError("标准数据尚未确认生成")
        return self._sha256(path)

    def _batch_root(self, project_id: str, batch_id: str) -> Path:
        root = (self.files_root / project_id / "batches" / batch_id).resolve()
        if not root.is_relative_to(self.files_root):
            raise LookupError("项目或批次标识超出项目目录")
        return root

    @staticmethod
    def _parameters(request: FilterRequest) -> dict[str, Any]:
        values = asdict(request)
        values.pop("project_id")
        values.pop("batch_id")
        return values

    @staticmethod
    def _read_filter_notes(path: Path) -> dict[str, list[str]]:
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook["筛选结果"]
        headers = [str(cell.value or "") for cell in sheet[1]]
        note_column = headers.index("筛选说明") + 1
        notes: dict[str, list[str]] = {}
        for row in range(2, sheet.max_row + 1):
            point_id = str(sheet.cell(row=row, column=1).value or "")
            note = str(sheet.cell(row=row, column=note_column).value or "")
            if point_id and point_id != "当天雨量":
                notes[point_id] = [item for item in note.split("；") if item]
        workbook.close()
        return notes

    def _validate_workbook(
        self, project_id: str, batch_id: str, content: bytes
    ) -> None:
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
            if workbook.sheetnames != ["筛选结果"]:
                raise ValueError("筛选文件结构无效：必须仅包含“筛选结果”工作表")
            sheet = workbook["筛选结果"]
            headers = [str(cell.value or "") for cell in sheet[1]]
            if (
                len(headers) < 3
                or headers[0] != "点位编号"
                or headers[-1] != "筛选说明"
            ):
                raise ValueError("筛选文件结构无效：表头不符合筛选结果契约")
            flow = self.standard_data.load_flow(project_id, batch_id)
            expected_dates = sorted(
                pd.to_datetime(flow["timestamp"]).dt.strftime("%Y-%m-%d").unique()
            )
            if headers[1:-1] != expected_dates:
                raise ValueError("筛选文件结构无效：日期列与标准数据不一致")
            if str(sheet.cell(row=2, column=1).value or "") != "当天雨量":
                raise ValueError("筛选文件结构无效：缺少当天雨量参考行")
            actual_points = sorted(
                str(sheet.cell(row=row, column=1).value or "")
                for row in range(3, sheet.max_row + 1)
            )
            expected_points = sorted(flow["point_id"].astype(str).unique())
            if actual_points != expected_points:
                raise ValueError("筛选文件结构无效：点位与标准数据不一致")
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError("筛选文件结构无效，无法读取 xlsx") from exc
        finally:
            if "workbook" in locals():
                workbook.close()

    @staticmethod
    def _sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def _connect(self) -> sqlite3.Connection:
        return sqlite3.connect(self.database)
