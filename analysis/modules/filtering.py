from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import tempfile
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side


@dataclass
class FilterConfig:
    missing_rate_threshold: float = 0.1
    expected_rows_per_day: int = 1440
    rain_day_filter_threshold: float = 2.0
    zero_like_threshold: float = 0.02
    high_zero_ratio_threshold: float = 0.5
    high_zero_ratio_normal_days_threshold: int = 5
    zero_day_drop_min_nonzero_keep_days: int = 3
    mean_lower_ratio: float = 0.5
    mean_upper_ratio: float = 2.0


@dataclass(frozen=True)
class FilterDecision:
    keep: bool
    reason: str


def daily_rain(rain: pd.DataFrame) -> pd.Series:
    if rain.empty:
        return pd.Series(dtype="float64")
    data = rain.copy()
    data["date"] = data["timestamp"].dt.date
    return data.groupby("date")["rain_mm"].sum()


def _calculate_daily_flow(day_groups: dict[object, pd.DataFrame]) -> pd.Series:
    day_flow: dict[object, int] = {}
    for date_key, day_df in day_groups.items():
        mean_flow = pd.to_numeric(day_df["flow_lps"], errors="coerce").fillna(0.0).mean()
        day_flow[date_key] = int(round(float(mean_flow) * 86.4))
    return pd.Series(day_flow, dtype="float64")


def _is_high_zero_ratio_day(series: pd.Series, zero_like_threshold: float, high_zero_ratio_threshold: float) -> bool:
    s = pd.to_numeric(series, errors="coerce").fillna(0.0).astype(float)
    if s.empty:
        return False
    zero_like = s <= zero_like_threshold
    zero_ratio = float(zero_like.mean())
    has_water = bool((s > zero_like_threshold).any())
    return has_water and zero_ratio > high_zero_ratio_threshold


def _calculate_center_ratio_base(day_flow: pd.Series) -> float:
    positive = day_flow[day_flow > 0]
    base = positive if len(positive) > 0 else day_flow
    if len(base) == 0:
        return 0.0
    return float(base.median())


class DataFilter:
    def __init__(self, config: FilterConfig):
        self.config = config

    def _evaluate_day(
        self,
        day: pd.Timestamp,
        day_df: pd.DataFrame,
        monitor_start: pd.Timestamp,
        monitor_end: pd.Timestamp,
        rain_daily: pd.Series,
        *,
        apply_zero_ratio_rule: bool = True,
        apply_missing_rule: bool = True,
        apply_rain_rule: bool = True,
    ) -> FilterDecision:
        date_key = day.date()
        if date_key == monitor_start.date():
            return FilterDecision(False, "监测周期第一天剔除")
        if date_key == monitor_end.date():
            return FilterDecision(False, "监测周期最后一天剔除")

        if apply_rain_rule:
            rain = float(rain_daily.get(date_key, 0.0))
            if rain >= self.config.rain_day_filter_threshold:
                return FilterDecision(False, f"雨天剔除(雨量={rain:.2f})")

        flow = day_df["flow_lps"].fillna(0.0)
        if apply_zero_ratio_rule and _is_high_zero_ratio_day(flow, self.config.zero_like_threshold, self.config.high_zero_ratio_threshold):
            ratio = float((pd.to_numeric(flow, errors="coerce").fillna(0.0) <= self.config.zero_like_threshold).mean())
            return FilterDecision(False, f"有水日近零值比例过高({ratio:.0%})，剔除")

        if apply_missing_rule:
            missing_rate = (self.config.expected_rows_per_day - len(day_df)) / self.config.expected_rows_per_day
            if missing_rate > self.config.missing_rate_threshold:
                return FilterDecision(False, f"缺失率超过{self.config.missing_rate_threshold:.0%}，剔除")

        return FilterDecision(True, "通过基础规则")

    def _should_apply_zero_ratio_rule(self, high_zero_days_count: int, base_candidate_days_count: int) -> bool:
        if high_zero_days_count < self.config.high_zero_ratio_normal_days_threshold:
            return True
        return high_zero_days_count < base_candidate_days_count

    def _apply_mean_ratio_rule(self, day_flow: pd.Series, candidates: set[pd.Timestamp]) -> set[pd.Timestamp]:
        if len(candidates) < 2:
            return candidates
        series = day_flow[day_flow.index.isin([d.date() for d in candidates])]
        center_v = _calculate_center_ratio_base(series)
        lo = center_v * self.config.mean_lower_ratio
        hi = center_v * self.config.mean_upper_ratio
        return {pd.Timestamp(d) for d, v in series.items() if lo <= float(v) <= hi}

    def _apply_high_zero_ratio_rule_with_fallback(
        self,
        base_candidates: set[pd.Timestamp],
        high_zero_days: set[pd.Timestamp],
    ) -> tuple[set[pd.Timestamp], set[pd.Timestamp], bool]:
        removed_high_zero_days = set(base_candidates & high_zero_days)
        keep = set(base_candidates) - removed_high_zero_days
        restored = False
        if len(keep) < self.config.zero_day_drop_min_nonzero_keep_days:
            keep = set(base_candidates)
            restored = True
        return keep, removed_high_zero_days, restored

    def _screen_point(
        self,
        point_id: str,
        point_df: pd.DataFrame,
        rain_daily: pd.Series,
    ) -> tuple[set[pd.Timestamp], dict[pd.Timestamp, float], list[str]]:
        df = point_df.copy()
        df["date"] = df["timestamp"].dt.date
        monitor_start = df["timestamp"].min()
        monitor_end = df["timestamp"].max()
        day_groups = dict(tuple(df.groupby("date")))
        day_flow = _calculate_daily_flow(day_groups)

        base_candidate_days = {
            pd.Timestamp(date_key)
            for date_key, day_df in day_groups.items()
            if self._evaluate_day(
                day=pd.Timestamp(date_key),
                day_df=day_df,
                monitor_start=monitor_start,
                monitor_end=monitor_end,
                rain_daily=rain_daily,
                apply_zero_ratio_rule=False,
                apply_missing_rule=True,
                apply_rain_rule=True,
            ).keep
        }

        high_zero_days = {
            pd.Timestamp(date_key)
            for date_key, day_df in day_groups.items()
            if _is_high_zero_ratio_day(day_df["flow_lps"], self.config.zero_like_threshold, self.config.high_zero_ratio_threshold)
        }

        apply_zero_ratio_rule = self._should_apply_zero_ratio_rule(
            high_zero_days_count=len(high_zero_days & base_candidate_days),
            base_candidate_days_count=len(base_candidate_days),
        )

        notes: list[str] = []
        if not apply_zero_ratio_rule and (high_zero_days & base_candidate_days):
            notes.append(f"该点位共有{len(high_zero_days & base_candidate_days)}天命中高零值比例，按站点常态处理，不据此剔除")

        point_candidates_strict = set(base_candidate_days)
        removed_high_zero_days: set[pd.Timestamp] = set()
        restored_high_zero_days = False

        if apply_zero_ratio_rule:
            point_candidates_strict, removed_high_zero_days, restored_high_zero_days = self._apply_high_zero_ratio_rule_with_fallback(
                base_candidates=base_candidate_days,
                high_zero_days=high_zero_days,
            )

        if removed_high_zero_days and not restored_high_zero_days:
            for d in sorted(removed_high_zero_days):
                notes.append(f"{d.strftime('%m-%d')}:有水日近零值比例过高剔除")
        elif removed_high_zero_days and restored_high_zero_days:
            notes.append(f"高零值比例剔除后保留天不足{self.config.zero_day_drop_min_nonzero_keep_days}天，恢复被该规则剔除的日期后再进行中位数筛选")

        keep_after_mean = self._apply_mean_ratio_rule(day_flow=day_flow, candidates=point_candidates_strict)
        removed_by_mean = point_candidates_strict - keep_after_mean

        for d in sorted(removed_by_mean):
            notes.append(f"{d.strftime('%m-%d')}:均值比例异常剔除")

        if len(keep_after_mean) == 0:
            candidates_no_zero_ratio = {
                pd.Timestamp(date_key)
                for date_key, day_df in day_groups.items()
                if self._evaluate_day(
                    day=pd.Timestamp(date_key),
                    day_df=day_df,
                    monitor_start=monitor_start,
                    monitor_end=monitor_end,
                    rain_daily=rain_daily,
                    apply_zero_ratio_rule=False,
                    apply_missing_rule=True,
                    apply_rain_rule=True,
                ).keep
            }
            if len(candidates_no_zero_ratio) > 0:
                notes.append("回退: 去除高零值比例规则")
                keep_after_mean = self._apply_mean_ratio_rule(day_flow=day_flow, candidates=candidates_no_zero_ratio)

            if len(keep_after_mean) == 0:
                candidates_boundary = {
                    pd.Timestamp(date_key)
                    for date_key, day_df in day_groups.items()
                    if self._evaluate_day(
                        day=pd.Timestamp(date_key),
                        day_df=day_df,
                        monitor_start=monitor_start,
                        monitor_end=monitor_end,
                        rain_daily=rain_daily,
                        apply_zero_ratio_rule=False,
                        apply_missing_rule=False,
                        apply_rain_rule=True,
                    ).keep
                }
                if len(candidates_boundary) > 0:
                    notes.append("回退: 去除均值比例规则")
                    keep_after_mean = candidates_boundary

        matrix = {pd.Timestamp(k): float(v) for k, v in day_flow.items()}
        return set(keep_after_mean), matrix, notes if notes else ["无明显异常"]

    def run(
        self,
        flow: pd.DataFrame,
        rain_daily: pd.Series,
    ) -> tuple[dict[str, set[pd.Timestamp]], dict[str, dict[pd.Timestamp, float]], dict[str, list[str]]]:
        selected: dict[str, set[pd.Timestamp]] = {}
        matrix: dict[str, dict[pd.Timestamp, float]] = {}
        notes: dict[str, list[str]] = {}

        if flow.empty:
            return selected, matrix, notes

        for point_id, point_df in flow.groupby("point_id", sort=True):
            keep, point_matrix, point_notes = self._screen_point(str(point_id), point_df.sort_values("timestamp"), rain_daily)
            selected[str(point_id)] = keep
            matrix[str(point_id)] = point_matrix
            notes[str(point_id)] = point_notes

        return selected, matrix, notes


def _write_filter_excel_uncommitted(
    output_file: Path,
    matrix: dict[str, dict[pd.Timestamp, float]],
    selected: dict[str, set[pd.Timestamp]],
    rain_daily: pd.Series,
    notes: dict[str, list[str]],
) -> None:
    days_sorted = sorted({d for row in matrix.values() for d in row.keys()})
    date_labels = [d.strftime("%Y-%m-%d") for d in days_sorted]

    frame = pd.DataFrame(index=matrix.keys(), columns=date_labels, dtype=float)
    for point_id, day_map in matrix.items():
        for d, val in day_map.items():
            frame.loc[point_id, d.strftime("%Y-%m-%d")] = val

    output_file.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if not rain_daily.empty:
            rain_values = [float(rain_daily.get(pd.Timestamp(c).date(), 0.0)) for c in date_labels]
        else:
            rain_values = [np.nan] * len(date_labels)
        rain_row = pd.DataFrame([rain_values], columns=date_labels, index=["当天雨量"])
        combined = pd.concat([rain_row, frame], axis=0)
        combined.to_excel(writer, sheet_name="筛选结果", index_label="点位编号")

    wb = load_workbook(output_file)
    ws = wb["筛选结果"]
    _apply_filter_excel_formatting(ws, selected, notes)
    wb.save(output_file)


def write_filter_excel(
    output_file: Path,
    matrix: dict[str, dict[pd.Timestamp, float]],
    selected: dict[str, set[pd.Timestamp]],
    rain_daily: pd.Series,
    notes: dict[str, list[str]],
) -> None:
    """Write a complete workbook and atomically publish it at ``output_file``."""
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        dir=output_file.parent,
        prefix=f".{output_file.stem}-",
        suffix=".xlsx",
        delete=False,
    ) as temporary:
        temporary_file = Path(temporary.name)
    try:
        _write_filter_excel_uncommitted(
            temporary_file, matrix, selected, rain_daily, notes
        )
        temporary_file.replace(output_file)
    finally:
        temporary_file.unlink(missing_ok=True)


def _apply_filter_excel_formatting(ws: Any, selected: dict[str, set[pd.Timestamp]], notes: dict[str, list[str]]) -> None:
    green_fill = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050")
    empty_fill = PatternFill(fill_type=None)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    point_to_row: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        point_id = ws.cell(row=row, column=1).value
        if point_id:
            point_to_row[str(point_id)] = row

    date_to_col: dict[str, int] = {}
    for col in range(2, ws.max_column + 1):
        date_to_col[str(ws.cell(row=1, column=col).value)] = col

    for point_id, keep_days in selected.items():
        row = point_to_row.get(point_id)
        if not row:
            continue
        for day in keep_days:
            col = date_to_col.get(day.strftime("%Y-%m-%d"))
            if col and ws.cell(row=row, column=col).value not in (None, "--"):
                ws.cell(row=row, column=col).fill = green_fill

    note_col = ws.max_column + 1
    ws.cell(row=1, column=note_col).value = "筛选说明"
    rain_row_num = None
    for row in range(2, ws.max_row + 1):
        point_id = str(ws.cell(row=row, column=1).value or "")
        if point_id == "当天雨量":
            ws.cell(row=row, column=note_col).value = "雨量参考行"
            rain_row_num = row
            continue
        if point_id in notes:
            ws.cell(row=row, column=note_col).value = "；".join(notes[point_id])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=note_col):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border
            if cell.value is None and cell.column < note_col:
                if rain_row_num and cell.row == rain_row_num:
                    cell.value = ""
                else:
                    cell.value = "--"
                cell.fill = empty_fill


def run_data_filter(
    flow: pd.DataFrame,
    rain: pd.DataFrame,
    output_xlsx: Path,
    config: FilterConfig | None = None,
) -> dict[str, list[str]]:
    cfg = config or FilterConfig()
    rain_daily = daily_rain(rain)
    selected, matrix, notes = DataFilter(cfg).run(flow, rain_daily)
    write_filter_excel(output_xlsx, matrix, selected, rain_daily, notes)
    return {point_id: sorted(day.strftime("%Y-%m-%d") for day in keep_days) for point_id, keep_days in selected.items()}
