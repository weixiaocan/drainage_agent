"""风险分析核心逻辑

分析旱天运行风险和雨天溢流风险。

旱天风险：
- 运行风险（基于最大充满度）
- 溢流风险（基于液位/井深）
- 淤积风险（基于平均流速）

雨天溢流风险：
- 基于降雨场次的最大液位和溢流风险值
"""

from dataclasses import dataclass
from datetime import timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from pipeline.core.data_utils import detect_site_info_columns, read_csv_with_fallback
from pipeline.core.schema import flow_to_legacy_df, normalize_flow_df, parse_flow_filename


@dataclass
class RiskConfig:
    """风险分析配置参数"""
    # 运行风险阈值
    running_risk_low: float = 0.75      # max_fullness < 此值为"运行良好"
    running_risk_medium: float = 1.0    # max_fullness < 此值为"低风险"
    running_risk_high: float = 2.0      # max_fullness <= 此值为"中风险"
    # 溢流风险阈值
    overflow_risk_low: float = 0.7      # overflow_value < 此值为"低溢流风险"
    overflow_risk_medium: float = 0.9   # overflow_value < 此值为"中溢流风险"
    # 淤积风险阈值
    silting_risk_medium: float = 0.3    # avg_velocity > 此值为"中淤积风险"
    # 合流管最小流速
    combined_min_velocity: float = 0.75
    # 分流管最小流速
    separate_min_velocity: float = 0.6
    # 降雨影响延迟
    rain_effect_delay: float = 12.0


def _running_risk(max_fullness: float, cfg: RiskConfig) -> str:
    """判断运行风险等级"""
    if max_fullness < cfg.running_risk_low:
        return "运行良好"
    if max_fullness < cfg.running_risk_medium:
        return "低风险"
    if max_fullness <= cfg.running_risk_high:
        return "中风险"
    return "高风险"


def _overflow_risk(overflow_value: float, cfg: RiskConfig) -> str:
    """判断溢流风险等级"""
    if overflow_value < cfg.overflow_risk_low:
        return "低溢流风险"
    if overflow_value < cfg.overflow_risk_medium:
        return "中溢流风险"
    if overflow_value <= 1.0:
        return "高溢流风险"
    return "已发生溢流"


def _silting_risk(avg_velocity: float, pipe_type: str, cfg: RiskConfig) -> str:
    """判断淤积风险等级"""
    pipe_type = pipe_type or ""
    min_speed = cfg.combined_min_velocity if "合流" in pipe_type else cfg.separate_min_velocity

    if avg_velocity > min_speed:
        return "低淤积风险"
    if avg_velocity > cfg.silting_risk_medium:
        return "中淤积风险"
    return "高淤积风险"


def _load_flow_data(csv_dir: Path) -> dict[str, pd.DataFrame]:
    """加载流量数据"""
    result: dict[str, pd.DataFrame] = {}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        df = read_csv_with_fallback(csv_path)
        df = flow_to_legacy_df(normalize_flow_df(df, csv_path))

        point_name = parse_flow_filename(csv_path).point_id
        result[point_name] = df.sort_values("数据时间").reset_index(drop=True)
    return result


def _load_site_info(site_info_file: Path) -> dict[str, dict]:
    """从点位信息.xlsx读取管径、井深等信息"""
    df = pd.read_excel(site_info_file)

    # 通过列名关键词识别列
    col_mapping = detect_site_info_columns(df)

    # 验证必需列
    if col_mapping["point_name"] is None:
        raise ValueError(f"点位信息文件缺少点位名称列，可用列: {list(df.columns)}")

    result: dict[str, dict] = {}
    for _, row in df.iterrows():
        # 点位名称 - 直接使用原始值，不做正则提取
        point_name = str(row[col_mapping["point_name"]]).strip() if pd.notna(row[col_mapping["point_name"]]) else ""

        if not point_name:
            continue

        # 可选字段
        diameter = 0.0
        if col_mapping["diameter"]:
            try:
                diameter = float(row[col_mapping["diameter"]]) if pd.notna(row[col_mapping["diameter"]]) else 0.0
            except (ValueError, TypeError):
                pass

        depth = 0.0
        if col_mapping["depth"]:
            try:
                depth = float(row[col_mapping["depth"]]) if pd.notna(row[col_mapping["depth"]]) else 0.0
            except (ValueError, TypeError):
                pass

        pipe_type = ""
        if col_mapping["pipe_type"]:
            pipe_type = str(row[col_mapping["pipe_type"]]) if pd.notna(row[col_mapping["pipe_type"]]) else ""

        result[point_name] = {
            "diameter": diameter,
            "depth": depth,
            "pipe_type": pipe_type,
        }

    return result


def _analyze_dry_weather_risk(
    dry_stats: pd.DataFrame,
    site_info: dict[str, dict],
    cfg: RiskConfig,
) -> pd.DataFrame:
    """分析旱天运行风险

    Args:
        dry_stats: 旱天分析统计数据（从综合分析结果.xlsx读取）
        site_info: 点位信息
        cfg: 配置参数

    Returns:
        旱天风险 DataFrame
    """
    rows: list[dict] = []

    for idx, (_, r) in enumerate(dry_stats.iterrows(), start=1):
        point_name = str(r.get("点位编号", ""))
        if not point_name:
            continue

        # 获取点位信息 - 尝试多种匹配方式
        info = site_info.get(point_name)

        # 如果直接匹配失败，尝试从点位编号中提取点位名部分
        # 格式: 设备编号_点位名（如 35943_13, 35943_#1）
        if info is None and "_" in point_name:
            import re
            # 提取第一个下划线之后的内容作为点位名
            point_id = point_name.split("_", 1)[1]
            info = site_info.get(point_id)

        if info is None:
            info = {"diameter": 0, "depth": 0, "pipe_type": ""}

        diameter = info["diameter"]
        well_depth = info["depth"]
        pipe_type = info["pipe_type"]

        # 从旱天统计获取数据
        max_level = float(r.get("最大液位(m)", 0)) if pd.notna(r.get("最大液位(m)")) else 0.0
        avg_velocity = float(r.get("平均流速(m/s)", 0)) if pd.notna(r.get("平均流速(m/s)")) else 0.0
        avg_flow = float(r.get("日均流量(m³/d)", 0)) if pd.notna(r.get("日均流量(m³/d)")) else 0.0

        # 计算风险指标
        max_fullness = max_level / diameter if diameter > 0 else 0.0
        overflow_value = max_level / well_depth if well_depth > 0 else 0.0

        rows.append({
            "序号": idx,
            "点位编号": point_name,
            "管径(m)": round(diameter, 3),
            "井深(m)": round(well_depth, 2),
            "日均流量(m³/d)": round(avg_flow, 2),
            "旱天流速(m/s)": round(avg_velocity, 4),
            "最大液位(m)": round(max_level, 3),
            "最大充满度": round(max_fullness, 2),
            "溢流风险值": round(overflow_value, 2),
            "淤积风险": _silting_risk(avg_velocity, pipe_type, cfg),
            "运行风险": _running_risk(max_fullness, cfg),
            "溢流风险": _overflow_risk(overflow_value, cfg),
        })

    return pd.DataFrame(rows)


def _analyze_rainy_overflow_risk(
    flow_data: dict[str, pd.DataFrame],
    event_data: dict[int, dict],
    site_info: dict[str, dict],
    selected_events: list[int] | None,
    delay_hours: float,
    cfg: RiskConfig,
) -> pd.DataFrame:
    """分析雨天溢流风险

    Args:
        flow_data: 流量数据
        event_data: 场次降雨数据
        site_info: 点位信息
        selected_events: 选中的场次编号
        delay_hours: 降雨影响延迟时间
        cfg: 配置参数

    Returns:
        雨天溢流风险 DataFrame
    """
    event_ids = sorted(event_data.keys())
    if selected_events:
        event_ids = [e for e in event_ids if e in selected_events]

    rows: list[dict] = []

    for event_id in event_ids:
        event = event_data[event_id]
        start = event["start"]
        end = event["end"] + timedelta(hours=delay_hours)
        rain_level = event.get("rain_level", "")

        for point_name, df in flow_data.items():
            # 获取点位信息 - 尝试多种匹配方式
            info = site_info.get(point_name)

            # 如果直接匹配失败，尝试从点位编号中提取点位名部分
            # 格式: 设备编号_点位名（如 35943_13, 35943_#1）
            if info is None and "_" in point_name:
                # 提取第一个下划线之后的内容作为点位名
                point_id = point_name.split("_", 1)[1]
                info = site_info.get(point_id)

            if info is None:
                info = {"diameter": 0, "depth": 0, "pipe_type": ""}

            well_depth = info["depth"]

            # 筛选降雨场次时间范围
            event_df = df[(df["数据时间"] >= start) & (df["数据时间"] <= end)]

            if event_df.empty:
                continue

            max_level = event_df["l"].max()
            overflow_value = max_level / well_depth if well_depth > 0 else 0.0

            rows.append({
                "降雨场次编号": event_id,
                "降雨等级": rain_level,
                "点位编号": point_name,
                "最大液位(m)": round(max_level, 3),
                "井深(m)": round(well_depth, 2),
                "溢流风险值": round(overflow_value, 3),
                "溢流风险": _overflow_risk(overflow_value, cfg),
            })

    return pd.DataFrame(rows)


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def run_risk_analysis(
    flow_dir: Path,
    combined_xlsx: Path,
    site_info_file: Path,
    event_data: dict[int, dict] | None = None,
    selected_events: list[int] | None = None,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行风险分析

    Args:
        flow_dir: 流量数据目录
        combined_xlsx: 综合分析结果 xlsx 文件
        site_info_file: 点位信息 xlsx 文件
        event_data: 场次降雨数据（雨天风险分析需要）
        selected_events: 选中的场次编号
        config: 可选配置参数

    Returns:
        {
            "dry_risk": pd.DataFrame,      # 旱天风险
            "rainy_risk": pd.DataFrame,    # 雨天溢流风险
        }
    """
    # 合并配置
    cfg = RiskConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 读取点位信息
    print(f"读取点位信息: {site_info_file}")
    site_info = _load_site_info(site_info_file)
    print(f"  - 点位数: {len(site_info)}")

    # 读取旱天分析结果
    print(f"读取旱天分析结果: {combined_xlsx}")
    dry_stats = pd.DataFrame()
    try:
        wb = load_workbook(combined_xlsx, data_only=True)
        if "旱天分析" in wb.sheetnames:
            ws = wb["旱天分析"]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            if data:
                dry_stats = pd.DataFrame(data[1:], columns=data[0])
        wb.close()
    except Exception as e:
        print(f"  警告: 读取旱天分析结果失败: {e}")

    # 旱天风险分析
    print("分析旱天运行风险")
    dry_risk_df = _analyze_dry_weather_risk(dry_stats, site_info, cfg)

    # 保存旱天风险
    _save_to_excel(
        dry_risk_df,
        combined_xlsx,
        "旱天风险",
        ["序号", "点位编号", "管径(m)", "井深(m)", "日均流量(m³/d)", "旱天流速(m/s)",
         "最大液位(m)", "最大充满度", "溢流风险值", "淤积风险", "运行风险", "溢流风险"]
    )
    print(f"保存旱天风险: {combined_xlsx}")

    # 雨天溢流风险分析
    rainy_risk_df = pd.DataFrame()
    if event_data:
        print("分析雨天溢流风险")
        flow_data = _load_flow_data(flow_dir)
        print(f"  - 加载流量数据: {len(flow_data)} 个点位")

        rainy_risk_df = _analyze_rainy_overflow_risk(
            flow_data, event_data, site_info, selected_events, cfg.rain_effect_delay, cfg
        )

        # 保存雨天溢流风险
        _save_to_excel(
            rainy_risk_df,
            combined_xlsx,
            "雨天溢流风险",
            ["降雨场次编号", "降雨等级", "点位编号", "最大液位(m)", "井深(m)", "溢流风险值", "溢流风险"]
        )
        print(f"保存雨天溢流风险: {combined_xlsx}")

    # 统计
    print(f"\n风险分析完成:")
    print(f"  - 旱天风险分析: {len(dry_risk_df)} 个点位")
    if not rainy_risk_df.empty:
        print(f"  - 雨天溢流风险分析: {len(rainy_risk_df)} 条记录")

    return {
        "dry_risk": dry_risk_df,
        "rainy_risk": rainy_risk_df,
    }

