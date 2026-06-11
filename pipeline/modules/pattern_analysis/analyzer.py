"""排污规律分析核心逻辑

从 _archive_old_agents/agents/pattern_agent.py 和 utils/pattern_feature_engine.py 提取核心逻辑。

基于旱天特征曲线判断排污规律：
- 第1类：符合生活用水规律（早晚高峰明显）
- 第2类：有波峰但不符合典型规律
- 第3类：曲线平坦/异常
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.dates import DateFormatter, HourLocator
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.signal import find_peaks

# 设置中文字体
mpl.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
mpl.rcParams["font.serif"] = ["SimHei"]
mpl.rcParams["axes.unicode_minus"] = False


@dataclass
class PatternConfig:
    """排污规律分析配置参数"""
    kz_life_min: float = 1.3        # 生活污水Kz下限
    kz_life_max: float = 2.5        # 生活污水Kz上限
    kz_flat_threshold: float = 1.2  # 低于此值认为平坦曲线
    peak_prominence_ratio: float = 0.15  # 峰相对于均值的最小突出度
    peak_distance_minutes: int = 120     # 两个峰之间最小间隔
    peak_valley_ratio_threshold: float = 1.5  # 峰谷比低于此值归为第3类
    mean_flow_low_threshold: float = 0.5      # 均值低于此值归为第3类
    night_ratio_high: float = 0.5   # 夜间占比高于此值认为异常


def _calculate_features(curve: pd.DataFrame) -> dict:
    """计算特征值

    Args:
        curve: 特征曲线 DataFrame，包含 'f', 'l', 'velo' 列

    Returns:
        特征字典
    """
    features = {}
    flow = curve["f"].values

    # 基础统计
    features["peak_value"] = float(flow.max())
    features["min_value"] = float(flow.min())
    features["mean_value"] = float(flow.mean())
    features["std_value"] = float(flow.std())

    # 峰谷比
    if features["min_value"] > 0.0001:
        features["peak_valley_ratio"] = features["peak_value"] / features["min_value"]
    else:
        features["peak_valley_ratio"] = 99.9

    # Kz（时变化系数）
    if features["mean_value"] > 0.0001:
        features["kz"] = features["peak_value"] / features["mean_value"]
    else:
        features["kz"] = 0.0

    # 峰识别
    mean_val = features["mean_value"]
    prominence = mean_val * 0.15
    distance = 120  # 分钟

    peaks, properties = find_peaks(flow, prominence=prominence, distance=distance)

    features["peak_count"] = len(peaks)
    if len(peaks) > 0:
        features["peak_times"] = _get_peak_times(curve.index, peaks)
        features["peak_significance"] = float(max(properties["prominences"]) / mean_val) if mean_val > 0 else 0
    else:
        features["peak_times"] = []
        features["peak_significance"] = 0.0

    # 时段分析
    features["morning_peak"] = _analyze_period(flow, curve.index, 6 * 60, 11 * 60)   # 06:00-11:00
    features["evening_peak"] = _analyze_period(flow, curve.index, 18 * 60, 24 * 60)  # 18:00-24:00
    features["night"] = _analyze_period(flow, curve.index, 1 * 60, 5 * 60)           # 01:00-05:00
    features["daytime"] = _analyze_period(flow, curve.index, 6 * 60, 24 * 60)        # 06:00-24:00

    # 夜间/日间比值
    if features["daytime"]["mean"] > 0:
        features["night_day_ratio"] = features["night"]["mean"] / features["daytime"]["mean"]
    else:
        features["night_day_ratio"] = 0.0

    return features


def _get_peak_times(index: pd.DatetimeIndex, peaks: np.ndarray) -> list[str]:
    """获取峰值时间"""
    times = []
    for p in peaks:
        t = index[p]
        times.append(f"{t.hour:02d}:{t.minute:02d}")
    return times


def _analyze_period(values: np.ndarray, index: pd.DatetimeIndex, start_min: int, end_min: int) -> dict:
    """分析指定时段"""
    # 计算每个时间点对应的分钟数
    minutes = np.array([t.hour * 60 + t.minute for t in index])

    if end_min == 24 * 60:
        mask = minutes >= start_min
    else:
        mask = (minutes >= start_min) & (minutes < end_min)

    period_values = values[mask]

    if len(period_values) > 0:
        return {
            "mean": float(period_values.mean()),
            "max": float(period_values.max()),
            "min": float(period_values.min()),
        }
    return {"mean": 0.0, "max": 0.0, "min": 0.0}


def _extract_peak_valley_periods(curve: pd.DataFrame) -> tuple[list[str], list[str]]:
    """提取波峰时段和波谷时段，返回小时级别的描述"""
    flow = curve["f"].values
    max_val = flow.max()
    min_val = flow.min()

    # 如果最大最小值差异太小，返回空
    if max_val - min_val < 0.1:
        return [], []

    threshold = (max_val + min_val) / 2

    # 标记高于/低于阈值
    above = flow > threshold
    below = flow <= threshold

    # 提取连续区间（小时级别）
    peak_hours = _extract_hours(curve.index, above, min_hours=1)
    valley_hours = _extract_hours(curve.index, below, min_hours=2)

    # 只保留主要的时段（各2个以内）
    return peak_hours[:2], valley_hours[:2]


def _extract_hours(index: pd.DatetimeIndex, mask: np.ndarray, min_hours: int = 1) -> list[str]:
    """提取小时级别的时段描述"""
    if len(mask) == 0:
        return []

    # 统计每个小时中满足条件的时间点数量
    hour_counts = {}  # {hour: count}
    for i, t in enumerate(index):
        h = t.hour
        if mask[i]:
            hour_counts[h] = hour_counts.get(h, 0) + 1

    # 只保留超过半小时的小时
    valid_hours = sorted([h for h, c in hour_counts.items() if c >= 30])

    if not valid_hours:
        return []

    # 合并连续的小时为区间
    periods = []
    start = valid_hours[0]
    end = valid_hours[0]

    for h in valid_hours[1:]:
        if h == end + 1:
            end = h
        else:
            # 只保留时长足够的区间
            duration = end - start + 1
            if duration >= min_hours:
                if start == end:
                    periods.append(f"{start}点")
                else:
                    periods.append(f"{start}-{end}点")
            start = h
            end = h

    # 添加最后一个区间
    duration = end - start + 1
    if duration >= min_hours:
        if start == end:
            periods.append(f"{start}点")
        else:
            periods.append(f"{start}-{end}点")

    return periods


def _classify_pattern(features: dict) -> tuple[int, str]:
    """规则分类（回退用）

    Returns:
        (category, reason)
        category: 1/2/3
    """
    kz = features.get("kz", 0)
    peak_count = features.get("peak_count", 0)
    peak_times = features.get("peak_times", [])
    mean_value = features.get("mean_value", 0)

    # 第3类：流量极低或曲线平坦
    if mean_value < 0.5:
        return 3, "流量接近零"
    if kz < 1.2:
        return 3, f"Kz={kz:.2f}，曲线平坦"

    # 检查是否有早晚高峰
    has_morning_peak = any("06:" <= pt <= "10:59" for pt in peak_times)
    has_evening_peak = any("17:" <= pt <= "22:59" for pt in peak_times)

    # 第1类：符合生活规律（有早晚高峰）
    if peak_count >= 1 and (has_morning_peak or has_evening_peak):
        if 1.3 <= kz <= 3.0:
            return 1, f"Kz={kz:.2f}，有早晚高峰特征"

    # 第2类：不符合典型规律
    return 2, f"Kz={kz:.2f}，波动特征不典型"


def _analyze_with_llm(
    point_name: str,
    curve: pd.DataFrame,
    features: dict,
    peak_periods: list[str],
    valley_periods: list[str],
    llm_client,
) -> tuple[int, str]:
    """使用LLM分析排污规律

    Args:
        point_name: 点位名称
        curve: 特征曲线数据
        features: 计算得到的特征值
        peak_periods: 波峰时段列表
        valley_periods: 波谷时段列表
        llm_client: LLM客户端

    Returns:
        (category, description)
    """
    import json

    # 计算时段流量特征
    flow = curve["f"].values
    total_flow = flow.sum() if flow.sum() > 0 else 1

    # 计算48个半小时平均值（每半小时30分钟）
    halfhourly_avg = []
    for i in range(48):
        start = i * 30
        end = (i + 1) * 30
        halfhourly_avg.append(flow[start:end].mean())
    # 格式化为 "0:00: x.xx, 0:30: x.xx, ..."
    halfhourly_avg_str = ", ".join([f"{i//2}:{'00' if i%2==0 else '30'}: {v:.2f}" for i, v in enumerate(halfhourly_avg)])

    # 夜间(0-7点)、早上(7-10点)、中午(11-15点)、下午(15-17点)、晚上(18-24点)
    night_flow = flow[0:420].mean()          # 0:00-7:00
    morning_flow = flow[420:600].mean()      # 7:00-10:00
    noon_flow = flow[660:900].mean()         # 11:00-15:00
    afternoon_flow = flow[900:1020].mean()   # 15:00-17:00
    evening_flow = flow[1080:1440].mean()    # 18:00-24:00

    night_ratio = flow[0:420].sum() / total_flow
    morning_ratio = flow[420:600].sum() / total_flow
    noon_ratio = flow[660:900].sum() / total_flow
    afternoon_ratio = flow[900:1020].sum() / total_flow
    evening_ratio = flow[1080:1440].sum() / total_flow

    # 加载prompt模板
    try:
        prompt_template = llm_client.load_prompt("pattern_analysis")
    except FileNotFoundError:
        # 如果模板不存在，使用内置模板
        prompt_template = """你是一个排水管网分析专家，请根据统计数据判断排污规律特征。

点位编号：{point_name}
统计数据：
- 日均流量：{mean_flow:.2f} L/s
- 时变化系数Kz：{kz:.2f}
- 峰谷比：{peak_valley_ratio:.2f}
- 峰值时间：{peak_times}
- 流量较高时段：{peak_periods}
- 流量较低时段：{valley_periods}
- 夜间(0-6点)占比：{night_ratio:.1%}
- 早间(6-12点)占比：{morning_ratio:.1%}
- 下午(12-18点)占比：{afternoon_ratio:.1%}
- 晚间(18-24点)占比：{evening_ratio:.1%}

请判断分类并生成简要描述（分类：1类=符合生活规律，2类=不典型，3类=平坦）。
输出JSON格式，包含category和description两个字段。"""

    # 填充模板
    prompt = prompt_template.format(
        point_name=point_name,
        mean_flow=features.get("mean_value", 0),
        max_flow=features.get("peak_value", 0),
        min_flow=features.get("min_value", 0),
        kz=features.get("kz", 0),
        peak_valley_ratio=features.get("peak_valley_ratio", 0),
        peak_count=features.get("peak_count", 0),
        peak_times="、".join(features.get("peak_times", [])),
        peak_periods="、".join(peak_periods) if peak_periods else "无",
        valley_periods="、".join(valley_periods) if valley_periods else "无",
        halfhourly_avg=halfhourly_avg_str,
        night_avg=night_flow,
        morning_avg=morning_flow,
        noon_avg=noon_flow,
        afternoon_avg=afternoon_flow,
        evening_avg=evening_flow,
        night_ratio=night_ratio,
        morning_ratio=morning_ratio,
        noon_ratio=noon_ratio,
        afternoon_ratio=afternoon_ratio,
        evening_ratio=evening_ratio,
    )

    try:
        # 调用LLM（使用JSON格式）
        response = llm_client.chat_json(prompt, temperature=0.1)

        # 解析JSON响应
        result = json.loads(response)
        category = int(result.get("category", 2))
        description = _description_from_llm_result(result)

        # 确保分类在有效范围内
        if category not in [1, 2, 3]:
            category = 2

        kz = features.get("kz", 0)
        mean_val = features.get("mean_value", 0)
        max_val = features.get("peak_value", 0)
        min_val = features.get("min_value", 0)
        peak_count = features.get("peak_count", 0)
        peak_times = features.get("peak_times", [])

        # 计算波动范围
        fluctuation = (max_val - min_val) / mean_val if mean_val > 0 else 0

        # 检查是否有周期性规律（波峰分散在全天各时段，且波峰显著高于日均）
        # 周期性规律：至少4个波峰，且波峰不在早/午/晚三个生活用水高峰时段内集中
        def has_periodic_pattern(peak_times, peak_count, curve_df, mean_val):
            if peak_count < 4:
                return False
            # 检查是否有深夜波峰（0-5点），且波峰值 > 日均
            # 只有显著的深夜波峰才算周期性规律
            night_peaks = []
            for t in peak_times:
                if "00:" <= t <= "04:59":
                    # 获取该时间点的流量值
                    hour, minute = int(t[:2]), int(t[3:5])
                    idx = hour * 60 + minute
                    if idx < len(curve_df):
                        val = curve_df["f"].iloc[idx]
                        # 波峰必须 > 日均才算显著深夜波峰
                        if val > mean_val:
                            night_peaks.append(t)
            # 如果有显著的深夜波峰，说明是周期性规律
            return len(night_peaks) > 0

        periodic = has_periodic_pattern(peak_times, peak_count, curve, mean_val)

        # 强制规则：Kz<1.2 的点位处理
        if kz < 1.2:
            if periodic:
                category = 2
                description = f"Kz={kz:.2f}<1.2但有周期性规律，{description}"
            else:
                category = 3
                description = f"Kz={kz:.2f}<1.2，波动范围{fluctuation*100:.0f}%<30%且波峰数{peak_count}<4，曲线平坦"

        # 强制规则：Kz>=1.2 时，先检查周期性规律，再检查第1类条件
        elif kz >= 1.2:
            if periodic:
                # 有周期性规律，归为第2类
                if category != 2:
                    category = 2
                    description = f"一天内出现{peak_count}个波峰，呈锯齿状规律涨落，疑似泵站调控"
            else:
                # 检查第1类条件
                # 1. 夜间流量最低
                night = features.get("night", {}).get("mean", 0)
                morning = features.get("morning_peak", {}).get("mean", 0)
                evening = features.get("evening_peak", {}).get("mean", 0)
                daytime = features.get("daytime", {}).get("mean", 0)

                night_is_lowest = night < morning and night < evening and night < daytime

                # 2. 晚上有明显高峰
                evening_max = features.get("evening_peak", {}).get("max", 0)
                evening_has_peak = evening_max > mean_val * 1.15 if mean_val > 0 else False

                # 3. 早上或中午有高峰（峰值>日均即可）
                morning_max = features.get("morning_peak", {}).get("max", 0)
                morning_has_peak = morning_max > mean_val if mean_val > 0 else False

                if night_is_lowest and evening_has_peak and morning_has_peak:
                    # 符合第1类条件
                    if category != 1:
                        category = 1
                        description = f"夜间流量最低，晚上高峰明显(峰值{evening_max:.1f}L/s为日均{mean_val:.1f}的{evening_max/mean_val:.1f}倍)，早上有小高峰，符合典型生活用水排放规律"

        return category, _build_report_description(point_name, features, category, peak_periods, valley_periods, description)

    except Exception as e:
        raise RuntimeError(f"LLM分析失败({point_name}): {e}") from e


def _description_from_llm_result(result: dict) -> str:
    """Extract a report-ready description from the structured LLM response."""
    parts: list[str] = []
    for key in ("description", "curve_description", "short_conclusion", "cause_reasoning"):
        value = str(result.get(key) or "").strip()
        if value:
            parts.append(value)
    findings = result.get("key_findings") or []
    if isinstance(findings, list):
        for item in findings[:2]:
            text = str(item or "").strip()
            if text:
                parts.append(text)
    causes = result.get("possible_causes") or []
    if isinstance(causes, list) and causes:
        cause_text = "、".join(str(item).strip() for item in causes if str(item).strip())
        if cause_text:
            parts.append(f"可能原因包括{cause_text}")
    return "；".join(_dedupe_text_parts(parts))


def _build_report_description(
    point_name: str,
    features: dict,
    category: int,
    peak_periods: list[str],
    valley_periods: list[str],
    base_description: str = "",
) -> str:
    """Build a complete sentence suitable for both Excel review and report reuse."""
    category_names = {
        1: "第1类，符合生活用水规律",
        2: "第2类，有波峰或波谷但不符合典型生活用水规律",
        3: "第3类，曲线平坦或无明显波峰波谷",
    }
    parts = [f"{point_name}点位属于{category_names.get(category, '未分类')}"]
    kz = features.get("kz", 0)
    peak_count = features.get("peak_count", 0)
    peak_valley_ratio = features.get("peak_valley_ratio", 0)
    parts.append(f"时变化系数Kz为{kz:.2f}，识别到{peak_count}个波峰")
    if peak_valley_ratio and peak_valley_ratio < 99:
        parts.append(f"峰谷比为{peak_valley_ratio:.2f}")
    period_summary = _build_period_summary(features)
    if period_summary:
        parts.append(period_summary)
    if peak_periods:
        parts.append(f"按小时阈值识别的相对高流量区间为{'、'.join(peak_periods)}")
    if valley_periods:
        parts.append(f"相对低流量区间为{'、'.join(valley_periods)}")
    if base_description:
        parts.append(base_description.rstrip("。"))
    else:
        parts.append(_fallback_pattern_sentence(category))
    return "。".join(_dedupe_text_parts(parts)) + "。"


def _fallback_pattern_sentence(category: int) -> str:
    if category == 1:
        return "曲线在日间或晚间用水时段出现抬升、夜间回落，整体符合生活污水排放规律"
    if category == 2:
        return "曲线存在明显波动，但高低峰出现时段与典型居民生活用水规律不完全一致，需结合汇水范围判断是否受工业排放、商业活动或泵站调度影响"
    if category == 3:
        return "流量曲线整体较为平坦，未形成清晰的早晚高峰特征，需关注持续入渗或恒定排放影响"
    return "流量曲线特征不明确，建议结合现场汇水范围进一步核查"


def _build_period_summary(features: dict) -> str:
    labels = [
        ("夜间1:00-5:00", features.get("night", {}).get("mean", 0)),
        ("早间6:00-11:00", features.get("morning_peak", {}).get("mean", 0)),
        ("日间6:00-24:00", features.get("daytime", {}).get("mean", 0)),
        ("晚间18:00-24:00", features.get("evening_peak", {}).get("mean", 0)),
    ]
    if not any(value > 0 for _, value in labels):
        return ""
    highest_label, highest_value = max(labels, key=lambda item: item[1])
    lowest_label, lowest_value = min(labels, key=lambda item: item[1])
    night = features.get("night", {}).get("mean", 0)
    morning = features.get("morning_peak", {}).get("mean", 0)
    evening = features.get("evening_peak", {}).get("mean", 0)
    daytime = features.get("daytime", {}).get("mean", 0)

    statements = [
        f"分时段均值显示，{highest_label}平均流量最高（{highest_value:.2f}L/s），{lowest_label}平均流量最低（{lowest_value:.2f}L/s）"
    ]
    if night < morning and night < evening and night < daytime:
        statements.append("夜间低流量特征较明显")
    else:
        statements.append("夜间流量未表现为全天最低，排放规律与典型居民生活污水存在差异")
    evening_peak_clear = daytime > 0 and evening > daytime * 1.1
    morning_peak_clear = daytime > 0 and morning > daytime * 1.1
    if evening_peak_clear:
        statements.append("晚间高峰较明显")
    elif morning_peak_clear:
        statements.append("早间高峰较明显")
    else:
        statements.append("早晚高峰不突出")
    return "，".join(statements)


def _dedupe_text_parts(parts: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for part in parts:
        text = str(part or "").strip().strip("。；;")
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _build_description(features: dict, category: int, peak_periods: list[str], valley_periods: list[str]) -> str:
    """生成排污规律描述（兼容旧调用）"""
    return _build_report_description("该", features, category, peak_periods, valley_periods)


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def run_pattern_analysis(
    dry_curve_data: dict[str, pd.DataFrame],
    combined_xlsx: Path,
    config: dict[str, Any] | None = None,
    llm_client=None,
) -> dict[str, Any]:
    """执行排污规律分析

    Args:
        dry_curve_data: 旱天特征曲线数据（从内存传入）
        combined_xlsx: 综合分析结果 xlsx 文件（输出）
        config: 可选配置参数
        llm_client: LLM客户端（必需，用于生成描述）

    Returns:
        {
            "pattern_df": pd.DataFrame,    # 分析结果
            "descriptions": dict,           # 点位描述
        }
    """
    if llm_client is None:
        raise RuntimeError("排污规律分析必须提供 LLMClient，不能回退到规则判断")

    # 合并配置
    cfg = PatternConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    rows: list[dict] = []
    descriptions: dict[str, str] = {}

    for point_name, curve in dry_curve_data.items():
        # 计算特征
        features = _calculate_features(curve)

        # 提取波峰波谷时段
        peak_periods, valley_periods = _extract_peak_valley_periods(curve)

        category, description = _analyze_with_llm(
            point_name, curve, features, peak_periods, valley_periods, llm_client
        )

        # 分类名称
        category_names = {
            1: "第1类-符合生活用水规律",
            2: "第2类-有波峰但不符合典型规律",
            3: "第3类-曲线平坦/异常",
        }

        # 简化的理由
        reason = f"Kz={features['kz']:.2f}"

        row = {
            "点位编号": point_name,
            "分类": category,
            "分类名称": category_names.get(category, "未分类"),
            "Kz值": round(features["kz"], 2),
            "峰谷比": round(features["peak_valley_ratio"], 2) if features["peak_valley_ratio"] < 99 else "N/A",
            "峰数量": features["peak_count"],
            "波峰时段": "、".join(peak_periods) if peak_periods else "",
            "波谷时段": "、".join(valley_periods) if valley_periods else "",
            "诊断理由": reason,
            "排污规律描述": description,
        }
        rows.append(row)
        descriptions[point_name] = description

    # 创建结果 DataFrame
    pattern_df = pd.DataFrame(rows)

    # 输出到综合分析结果.xlsx
    _save_to_excel(
        pattern_df,
        combined_xlsx,
        "排污规律分析",
        ["点位编号", "分类", "分类名称", "Kz值", "峰谷比", "峰数量", "波峰时段", "波谷时段", "诊断理由", "排污规律描述"]
    )
    print(f"保存排污规律分析: {combined_xlsx}")

    # 统计
    cat_counts = pattern_df["分类"].value_counts().to_dict()
    print(f"\n排污规律分析完成:")
    print(f"  - 第1类(符合生活规律): {cat_counts.get(1, 0)} 个点位")
    print(f"  - 第2类(不符合典型规律): {cat_counts.get(2, 0)} 个点位")
    print(f"  - 第3类(曲线平坦/异常): {cat_counts.get(3, 0)} 个点位")

    return {
        "pattern_df": pattern_df,
        "descriptions": descriptions,
    }


def draw_dry_flow_curve(
    dry_curve_data: dict[str, pd.DataFrame],
    dry_curve_data_workday: dict[str, pd.DataFrame],
    dry_curve_data_weekend: dict[str, pd.DataFrame],
    flow_data: dict[str, pd.DataFrame],
    dry_days: dict[str, list[str]],
    day_num: pd.DataFrame,
    output_dir: Path,
) -> int:
    """绘制流量特征曲线（按照原格式）

    Args:
        dry_curve_data: 总体特征曲线
        dry_curve_data_workday: 工作日特征曲线
        dry_curve_data_weekend: 周末特征曲线
        flow_data: 原始流量数据
        dry_days: 各点位旱天日期列表
        day_num: 工作日/周末天数统计
        output_dir: 输出目录

    Returns:
        生成的图表数量
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    count = 0

    for point_name in dry_curve_data.keys():
        # 获取所有旱天数据
        point_dry_days = dry_days.get(point_name, [])
        num_of_day = len(point_dry_days)
        if num_of_day == 0:
            continue

        # 构建每日流量数据（不强制要求1440行）
        day_index = pd.date_range("00:00:00", "23:59:00", freq="min")
        dry_flow_each_day_df = pd.DataFrame(index=day_index)

        if point_name in flow_data:
            df = flow_data[point_name]
            for date in point_dry_days:
                day_data = df[df["数据时间"].dt.strftime("%Y-%m-%d") == date]["f"].values
                if len(day_data) > 0:
                    # 使用实际数据长度，填充到1440
                    padded_data = np.zeros(1440)
                    padded_data[:len(day_data)] = day_data[:1440]
                    dry_flow_each_day_df[date] = padded_data

        if dry_flow_each_day_df.empty or dry_flow_each_day_df.shape[1] == 0:
            continue

        valid_days = list(dry_flow_each_day_df.columns)

        # 绘图
        fig = plt.figure(figsize=(10, 5), dpi=120)
        ax1 = fig.add_subplot(1, 1, 1)

        # 每日曲线用灰色显示
        num_valid = len(valid_days)
        if num_valid == 1:
            dry_flow_each_day_df[valid_days[0]].plot(ax=ax1, color="#D3D3D3", label="每日流量", legend=True, alpha=0.5)
        elif num_valid > 1:
            for j in range(num_valid - 1):
                dry_flow_each_day_df[valid_days[j]].plot(ax=ax1, color="#D3D3D3", label="", legend=False, alpha=0.5)
            dry_flow_each_day_df[valid_days[-1]].plot(ax=ax1, color="#D3D3D3", label="每日流量", legend=True, alpha=0.5)

        # 特征曲线
        workday_num = day_num.loc[point_name, "工作日天数"] if point_name in day_num.index else 0
        weekend_num = day_num.loc[point_name, "周末天数"] if point_name in day_num.index else 0

        if workday_num != 0 and weekend_num != 0:
            dry_curve_data[point_name]["f"].plot(ax=ax1, color="#1E90FF", label="流量特征曲线_总体", legend=True)
        else:
            dry_curve_data[point_name]["f"].plot(ax=ax1, color="#1E90FF", label="流量特征曲线", legend=True)

        # 图形设计
        ax1.xaxis.set_major_formatter(DateFormatter("%H:%M"))
        ax1.xaxis.set_major_locator(HourLocator(byhour=range(0, 24, 2)))
        ax1.set_xlabel("时间")
        ax1.set_ylabel("流量/(L/s)")

        plt.tight_layout()
        img_path = output_dir / f"{point_name}_流量特征曲线.png"
        plt.savefig(img_path, dpi=300, bbox_inches="tight")
        plt.close()
        count += 1

    print(f"生成流量特征曲线图: {count} 张")
    return count


def draw_dry_level_curve(
    dry_curve_data: dict[str, pd.DataFrame],
    dry_curve_data_workday: dict[str, pd.DataFrame],
    dry_curve_data_weekend: dict[str, pd.DataFrame],
    flow_data: dict[str, pd.DataFrame],
    dry_days: dict[str, list[str]],
    day_num: pd.DataFrame,
    output_dir: Path,
) -> int:
    """绘制液位特征曲线（按照原格式）

    Args:
        dry_curve_data: 总体特征曲线
        dry_curve_data_workday: 工作日特征曲线
        dry_curve_data_weekend: 周末特征曲线
        flow_data: 原始流量数据
        dry_days: 各点位旱天日期列表
        day_num: 工作日/周末天数统计
        output_dir: 输出目录

    Returns:
        生成的图表数量
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    count = 0

    for point_name in dry_curve_data.keys():
        point_dry_days = dry_days.get(point_name, [])
        num_of_day = len(point_dry_days)
        if num_of_day == 0:
            continue

        # 构建每日液位数据（不强制要求1440行）
        day_index = pd.date_range("00:00:00", "23:59:00", freq="min")
        dry_level_each_day_df = pd.DataFrame(index=day_index)

        if point_name in flow_data:
            df = flow_data[point_name]
            for date in point_dry_days:
                day_data = df[df["数据时间"].dt.strftime("%Y-%m-%d") == date]["l"].values
                if len(day_data) > 0:
                    # 使用实际数据长度，填充到1440
                    padded_data = np.zeros(1440)
                    padded_data[:len(day_data)] = day_data[:1440]
                    dry_level_each_day_df[date] = padded_data

        if dry_level_each_day_df.empty or dry_level_each_day_df.shape[1] == 0:
            continue

        valid_days = list(dry_level_each_day_df.columns)

        # 绘图
        fig = plt.figure(figsize=(10, 5), dpi=120)
        ax1 = fig.add_subplot(1, 1, 1)

        # 每日液位曲线用灰色显示
        num_valid = len(valid_days)
        if num_valid == 1:
            dry_level_each_day_df[valid_days[0]].plot(ax=ax1, color="#D3D3D3", label="每日液位", legend=True, alpha=0.5)
        elif num_valid > 1:
            for j in range(num_valid - 1):
                dry_level_each_day_df[valid_days[j]].plot(ax=ax1, color="#D3D3D3", label="", legend=False, alpha=0.5)
            dry_level_each_day_df[valid_days[-1]].plot(ax=ax1, color="#D3D3D3", label="每日液位", legend=True, alpha=0.5)

        # 特征曲线
        workday_num = day_num.loc[point_name, "工作日天数"] if point_name in day_num.index else 0
        weekend_num = day_num.loc[point_name, "周末天数"] if point_name in day_num.index else 0

        if workday_num != 0 and weekend_num != 0:
            dry_curve_data[point_name]["l"].plot(ax=ax1, label="液位特征曲线", color="#1E90FF", legend=True)
        else:
            dry_curve_data[point_name]["l"].plot(ax=ax1, label="液位特征曲线", color="#1E90FF", legend=True)

        # 图形设计
        ax1.xaxis.set_major_formatter(DateFormatter("%H:%M"))
        ax1.xaxis.set_major_locator(HourLocator(byhour=range(0, 24, 2)))
        ax1.set_xlabel("时间")
        ax1.set_ylabel("液位/(m)")

        plt.tight_layout()
        img_path = output_dir / f"{point_name}_液位特征曲线.png"
        plt.savefig(img_path, dpi=300, bbox_inches="tight")
        plt.close()
        count += 1

    print(f"生成液位特征曲线图: {count} 张")
    return count


def generate_curve_charts(
    dry_curve_data: dict[str, pd.DataFrame],
    dry_curve_data_workday: dict[str, pd.DataFrame],
    dry_curve_data_weekend: dict[str, pd.DataFrame],
    flow_data: dict[str, pd.DataFrame],
    dry_days: dict[str, list[str]],
    day_num: pd.DataFrame,
    output_dir: Path,
) -> dict[str, int]:
    """生成所有特征曲线图表

    Args:
        dry_curve_data: 总体特征曲线
        dry_curve_data_workday: 工作日特征曲线
        dry_curve_data_weekend: 周末特征曲线
        flow_data: 原始流量数据
        dry_days: 各点位旱天日期列表
        day_num: 工作日/周末天数统计
        output_dir: 输出目录

    Returns:
        {"flow_charts": 流量图数量, "level_charts": 液位图数量}
    """
    flow_count = draw_dry_flow_curve(
        dry_curve_data, dry_curve_data_workday, dry_curve_data_weekend,
        flow_data, dry_days, day_num, output_dir
    )

    level_count = draw_dry_level_curve(
        dry_curve_data, dry_curve_data_workday, dry_curve_data_weekend,
        flow_data, dry_days, day_num, output_dir
    )

    return {"flow_charts": flow_count, "level_charts": level_count}

