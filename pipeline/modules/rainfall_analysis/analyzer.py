"""降雨分析核心逻辑

从 colleague_tool/code/read_prepro_rain_data.py 和 analyze_rain.py 提取核心逻辑。
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from pipeline.core.schema import normalize_rainfall_df


@dataclass
class RainfallConfig:
    """降雨分析配置参数"""
    min_interval: float = 12.0    # 场次降雨划分时间间隔（小时）
    min_rainfall: float = 1.0     # 最小降雨量阈值（mm）


@dataclass
class RainfallData:
    """降雨数据容器"""
    df: pd.DataFrame              # 降雨数据，index 为时间
    freq: str                     # 数据频率: "minute" 或 "hourly"


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV"""
    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as err:
            last_err = err
    if last_err:
        raise last_err
    raise RuntimeError(f"无法读取 CSV: {path}")


def _load_rain_data(rainfall_file: Path) -> RainfallData:
    """加载降雨数据并预处理

    自动检测数据频率（分钟级/小时级/日级），保持原始粒度不变。

    Returns:
        RainfallData 对象，包含数据和频率信息
    """
    df = normalize_rainfall_df(_read_csv_with_fallback(rainfall_file))
    df = df.rename(columns={"timestamp": "time", "rain_mm": "rain"})

    # 检测数据频率
    df_sorted = df.sort_values("time")
    time_diffs = df_sorted["time"].diff().dropna()
    median_diff = time_diffs.median()

    if median_diff <= pd.Timedelta(minutes=5):
        freq = "minute"
        print(f"检测到分钟级降雨数据（间隔 {median_diff}）")
        # 填充连续时间序列
        time_start = df["time"].min()
        time_end = df["time"].max()
        full_index = pd.date_range(time_start, time_end, freq="min")
        full_df = pd.DataFrame({"time": full_index})
        df = full_df.merge(df, on="time", how="left")
        df["rain"] = df["rain"].fillna(0.0)

    elif median_diff <= pd.Timedelta(hours=3):
        freq = "hourly"
        print(f"检测到小时级降雨数据（间隔 {median_diff}），保持小时粒度")
        # 填充连续小时序列
        time_start = df_sorted["time"].min()
        time_end = df_sorted["time"].max()
        full_index = pd.date_range(time_start, time_end, freq="h")
        full_df = pd.DataFrame({"time": full_index})
        df = full_df.merge(df_sorted, on="time", how="left")
        df["rain"] = df["rain"].fillna(0.0)

    elif median_diff <= pd.Timedelta(days=1.5):
        freq = "daily"
        print(f"检测到日级降雨数据（间隔 {median_diff}），保持日粒度")
        # 对于日级数据，直接使用原始数据
        df = df_sorted.copy()
    else:
        raise ValueError(f"不支持的降雨数据频率: {median_diff}，请提供分钟级、小时级或日级数据")

    return RainfallData(df=df.set_index("time"), freq=freq)


def _get_daily_rain(rain_data: pd.DataFrame, freq: str = "minute") -> pd.DataFrame:
    """计算日降雨量

    Args:
        rain_data: 降雨数据 DataFrame，index 为时间
        freq: 数据频率

    Returns:
        日降雨量统计 DataFrame
    """
    if freq == "daily":
        # 日级数据直接使用
        daily = rain_data.reset_index()
        daily.columns = ["日期", "日降雨量(mm)"]
    else:
        # 分钟/小时级数据需要 resample
        daily = rain_data.resample("D").sum()
        daily = daily.reset_index()
        daily.columns = ["日期", "日降雨量(mm)"]
    return daily


def _time_split(df: pd.DataFrame, n_hours: float) -> list[tuple[datetime, datetime]]:
    """场次降雨划分

    Args:
        df: 降雨数据 DataFrame，index 为时间，包含 'rain' 列
        n_hours: 划分时间间隔（小时）

    Returns:
        [(start_time, end_time), ...] 场次降雨起止时间列表
    """
    # 去掉 0 值
    df_nonzero = df[df["rain"] > 0].copy()
    if df_nonzero.empty:
        return []

    timestamps = df_nonzero.index
    rain_rng: list[tuple[datetime, datetime]] = []
    time_nodes = [timestamps[0]]

    for i in range(1, len(timestamps)):
        diff = (timestamps[i] - timestamps[i - 1]).total_seconds()
        if diff >= n_hours * 3600:
            rain_rng.append((time_nodes[-1], timestamps[i - 1]))
            time_nodes.append(timestamps[i])

    rain_rng.append((time_nodes[-1], timestamps[-1]))
    return rain_rng


def _get_rain_info(
    rain_rng: list[tuple[datetime, datetime]],
    rain_data: pd.DataFrame,
    min_rain: float,
    freq: str = "minute",
) -> pd.DataFrame:
    """提取场次降雨特征值

    Args:
        rain_rng: 场次降雨起止时间列表
        rain_data: 降雨数据 DataFrame
        min_rain: 最小降雨量阈值（mm）
        freq: 数据频率 ("minute" 或 "hourly")

    Returns:
        场次降雨统计 DataFrame
    """
    records: list[dict] = []

    for start, end in rain_rng:
        event_data = rain_data.loc[start:end]
        total_rain = event_data["rain"].sum()

        if total_rain > min_rain:
            event_id = len(records) + 1
            duration = (end - start).total_seconds() / 3600  # 小时
            record = {
                "场次编号": event_id,
                "开始时间": start,
                "结束时间": end,
                "总降雨量(mm)": round(total_rain, 2),
                "降雨历时(h)": round(duration, 2),
            }

            if freq == "minute":
                # 分钟级数据：计算分钟级滚动指标
                record["峰值雨强(mm/min)"] = round(event_data["rain"].max(), 2)
                record["最大5分钟降雨量(mm)"] = round(event_data["rain"].rolling(5).sum().max(), 2)
                record["最大10分钟降雨量(mm)"] = round(event_data["rain"].rolling(10).sum().max(), 2)
                record["最大1小时降雨量(mm)"] = round(event_data["rain"].rolling(60).sum().max(), 2)
                record["最大24小时降雨量(mm)"] = round(event_data["rain"].rolling(1440).sum().max(), 2)
            else:
                # 小时级数据：计算小时级滚动指标
                record["峰值雨强(mm/h)"] = round(event_data["rain"].max(), 2)
                record["最大3小时降雨量(mm)"] = round(event_data["rain"].rolling(3).sum().max(), 2)
                record["最大6小时降雨量(mm)"] = round(event_data["rain"].rolling(6).sum().max(), 2)
                record["最大12小时降雨量(mm)"] = round(event_data["rain"].rolling(12).sum().max(), 2)
                record["最大24小时降雨量(mm)"] = round(event_data["rain"].rolling(24).sum().max(), 2)

            record["平均强度(mm/h)"] = round(total_rain / duration, 2) if duration > 0 else 0
            records.append(record)

    return pd.DataFrame(records)


def _classify_rain_level(total_rain: float) -> str:
    """根据总降雨量判断降雨等级"""
    if total_rain < 10:
        return "小雨"
    elif total_rain < 25:
        return "中雨"
    elif total_rain < 50:
        return "大雨"
    elif total_rain < 100:
        return "暴雨"
    elif total_rain < 250:
        return "大暴雨"
    else:
        return "特大暴雨"


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # 打开或创建工作簿
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # 删除已存在的 sheet
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # 创建新 sheet
    ws = wb.create_sheet(sheet_name)

    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d")
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 替换表头
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    # 格式化
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def _add_daily_analysis_sheet(excel_path: Path, daily_rain: pd.DataFrame, rain_overview: pd.DataFrame) -> None:
    """添加降雨日分析 sheet

    包含：
    1. 降雨概况数据
    2. 日降雨量时间序列条形图
    3. 降雨日占比饼图
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font, RegularTextRun
    from openpyxl.styles import Alignment, Border, Side, Font as CellFont

    # 确保目录存在
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # 打开或创建工作簿
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # 删除所有旧的降雨相关 sheet
    old_sheets = [
        "降雨日分析", "降雨场次分析",
        "rainfall_analysis", "降雨分析",
        "日降雨量统计", "场次降雨统计", "降雨概况",
    ]
    for sheet_name in old_sheets:
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    # ========== 写入降雨概况 ==========
    ws.cell(row=1, column=1, value="降雨概况")
    ws.merge_cells('A1:B1')
    ws.cell(row=1, column=1).font = CellFont(bold=True, size=12)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for i, row in enumerate(rain_overview.itertuples(), 2):
        ws.cell(row=i, column=1, value=row.指标)
        ws.cell(row=i, column=2, value=row.数值)

    # 格式化降雨概况表格
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row_idx in range(2, len(rain_overview) + 2):
        for col_idx in range(1, 3):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # ========== 写入日降雨量数据（用于图表） ==========
    start_col = 4  # 从第4列开始写日降雨量数据
    ws.cell(row=1, column=start_col, value="日期")
    ws.cell(row=1, column=start_col + 1, value="日降雨量(mm)")

    for i, row in enumerate(daily_rain.itertuples(), 2):
        date_val = row.日期
        if hasattr(date_val, 'strftime'):
            ws.cell(row=i, column=start_col, value=date_val.strftime("%Y-%m-%d"))
        else:
            ws.cell(row=i, column=start_col, value=str(date_val)[:10])
        ws.cell(row=i, column=start_col + 1, value=row._2)

    data_rows = len(daily_rain) + 1

    # ========== 写入饼图数据 ==========
    pie_col = start_col + 4
    rainy_days = (daily_rain["日降雨量(mm)"] > 0).sum()
    non_rainy_days = len(daily_rain) - rainy_days

    ws.cell(row=1, column=pie_col, value="类型")
    ws.cell(row=1, column=pie_col + 1, value="天数")
    ws.cell(row=2, column=pie_col, value="降雨日")
    ws.cell(row=2, column=pie_col + 1, value=rainy_days)
    ws.cell(row=3, column=pie_col, value="非降雨日")
    ws.cell(row=3, column=pie_col + 1, value=non_rainy_days)

    # ========== 图1: 日降雨量时间序列条形图 ==========
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "日降雨量时间序列"
    bar_chart.y_axis.title = "降雨量(mm)"
    bar_chart.x_axis.title = "日期"
    bar_chart.width = 20
    bar_chart.height = 10

    # 去掉网格线
    bar_chart.y_axis.majorGridlines = None
    bar_chart.y_axis.minorGridlines = None
    bar_chart.x_axis.majorGridlines = None
    bar_chart.x_axis.minorGridlines = None

    # 数据引用
    data_ref = Reference(ws, min_col=start_col + 1, min_row=1, max_row=data_rows)
    cats_ref = Reference(ws, min_col=start_col, min_row=2, max_row=data_rows)

    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)

    ws.add_chart(bar_chart, "A8")

    # ========== 图2: 降雨日占比饼图 ==========
    pie_chart = PieChart()
    pie_chart.title = None
    pie_chart.width = 10
    pie_chart.height = 10
    pie_chart.legend = None

    pie_data = Reference(ws, min_col=pie_col + 1, min_row=1, max_row=3)
    pie_cats = Reference(ws, min_col=pie_col, min_row=2, max_row=3)

    pie_chart.add_data(pie_data, titles_from_data=True)
    pie_chart.set_categories(pie_cats)

    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.dataLabels.showVal = True
    pie_chart.dataLabels.showCatName = True

    ws.add_chart(pie_chart, "A28")

    # ========== 设置字体 ==========
    def set_chart_font(chart_obj):
        if chart_obj.title:
            chart_obj.title.txPr = RichText(
                p=[Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(
                        latin=Font(typeface="Times New Roman"),
                        ea=Font(typeface="宋体")
                    )),
                    r=[RegularTextRun(t=chart_obj.title.txPr.p[0].r[0].t if chart_obj.title.txPr and chart_obj.title.txPr.p and chart_obj.title.txPr.p[0].r else "")]
                )]
            )

        if hasattr(chart_obj, 'y_axis') and chart_obj.y_axis.title:
            y_title_text = chart_obj.y_axis.title.txPr.p[0].r[0].t if chart_obj.y_axis.title.txPr and chart_obj.y_axis.title.txPr.p and chart_obj.y_axis.title.txPr.p[0].r else ""
            chart_obj.y_axis.title.txPr = RichText(
                p=[Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(
                        latin=Font(typeface="Times New Roman"),
                        ea=Font(typeface="宋体")
                    )),
                    r=[RegularTextRun(t=y_title_text)]
                )]
            )

        if hasattr(chart_obj, 'x_axis') and chart_obj.x_axis.title:
            x_title_text = chart_obj.x_axis.title.txPr.p[0].r[0].t if chart_obj.x_axis.title.txPr and chart_obj.x_axis.title.txPr.p and chart_obj.x_axis.title.txPr.p[0].r else ""
            chart_obj.x_axis.title.txPr = RichText(
                p=[Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(
                        latin=Font(typeface="Times New Roman"),
                        ea=Font(typeface="宋体")
                    )),
                    r=[RegularTextRun(t=x_title_text)]
                )]
            )

        if chart_obj.legend:
            chart_obj.legend.txPr = RichText(
                p=[Paragraph(
                    pPr=ParagraphProperties(defRPr=CharacterProperties(
                        latin=Font(typeface="Times New Roman"),
                        ea=Font(typeface="宋体")
                    )),
                    r=[]
                )]
            )

    set_chart_font(bar_chart)

    if pie_chart.dataLabels:
        pie_chart.dataLabels.txPr = RichText(
            p=[Paragraph(
                pPr=ParagraphProperties(defRPr=CharacterProperties(
                    latin=Font(typeface="Times New Roman"),
                    ea=Font(typeface="宋体")
                )),
                r=[]
            )]
        )

    wb.save(excel_path)
    print(f"保存降雨日分析: {excel_path}")


def _add_event_analysis_sheet(excel_path: Path, event_rain: pd.DataFrame, freq: str) -> None:
    """添加降雨场次分析 sheet

    包含：
    1. 场次降雨统计表格
    """
    if event_rain.empty:
        print("无场次降雨数据，跳过")
        return

    wb = load_workbook(excel_path)

    # 删除已存在的 sheet
    sheet_name = "降雨场次分析"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    # 表头
    if freq == "minute":
        headers = ["场次编号", "开始时间", "结束时间", "总降雨量(mm)", "降雨历时(h)",
                   "峰值雨强(mm/min)", "最大5分钟降雨量(mm)", "最大10分钟降雨量(mm)",
                   "最大1小时降雨量(mm)", "最大24小时降雨量(mm)", "平均强度(mm/h)", "降雨等级"]
    else:
        headers = ["场次编号", "开始时间", "结束时间", "总降雨量(mm)", "降雨历时(h)",
                   "峰值雨强(mm/h)", "最大3小时降雨量(mm)", "最大6小时降雨量(mm)",
                   "最大12小时降雨量(mm)", "最大24小时降雨量(mm)", "平均强度(mm/h)", "降雨等级"]

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据
    for row_idx in range(len(event_rain)):
        for col_idx in range(len(event_rain.columns)):
            value = event_rain.iloc[row_idx, col_idx]
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M")
            # 处理 NaN 值
            if pd.isna(value):
                value = ""
            ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

    # 格式化
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)
    print(f"保存降雨场次分析: {excel_path}")


def _get_rain_overview(daily_rain: pd.DataFrame) -> pd.DataFrame:
    """生成降雨概况统计

    Args:
        daily_rain: 日降雨量统计 DataFrame

    Returns:
        降雨概况 DataFrame
    """
    total_days = len(daily_rain)
    rainy_days = (daily_rain["日降雨量(mm)"] > 0).sum()
    non_rainy_days = total_days - rainy_days
    total_rain = daily_rain["日降雨量(mm)"].sum()

    return pd.DataFrame({
        "指标": ["监测总天数", "降雨日数", "非降雨日数", "总降雨量(mm)"],
        "数值": [total_days, rainy_days, non_rainy_days, round(total_rain, 2)],
    })


def _save_rainfall_png_charts(daily_rain: pd.DataFrame, output_dir: Path) -> dict[str, Path]:
    """Save Word-ready rainfall charts as PNG files."""
    output_dir.mkdir(parents=True, exist_ok=True)
    chart_paths = {
        "daily_bar": output_dir / "日降雨量时间序列图.png",
        "rainy_ratio": output_dir / "降雨日占比饼图.png",
    }
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception as exc:
        print(f"生成降雨 PNG 图失败，matplotlib 不可用: {exc}")
        return chart_paths

    plot_df = daily_rain.copy()
    plot_df["日期"] = pd.to_datetime(plot_df["日期"], errors="coerce")
    plot_df["日降雨量(mm)"] = pd.to_numeric(plot_df["日降雨量(mm)"], errors="coerce").fillna(0)

    plt.rcParams["font.family"] = ["Times New Roman", "SimSun"]
    plt.rcParams["font.sans-serif"] = ["SimSun", "宋体", "Microsoft YaHei", "DejaVu Sans"]
    plt.rcParams["font.serif"] = ["Times New Roman", "SimSun"]
    plt.rcParams["axes.unicode_minus"] = False

    fig, ax = plt.subplots(figsize=(9.2, 4.8), dpi=180)
    labels = [d.strftime("%Y-%m-%d") if not pd.isna(d) else "" for d in plot_df["日期"]]
    x = np.arange(len(labels))
    ax.bar(x, plot_df["日降雨量(mm)"], color="#5B9BD5", edgecolor="#2F5597", linewidth=0.6)
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.set_ylabel("降雨量(mm)")
    ax.set_xlabel("日期")
    ax.set_title("日降雨量时间序列")
    ax.tick_params(axis="x", rotation=45, labelsize=7.5)
    ax.tick_params(axis="y", labelsize=9)
    ax.grid(False)
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color("#000000")
        spine.set_linewidth(0.8)
    fig.tight_layout()
    fig.savefig(chart_paths["daily_bar"], bbox_inches="tight")
    plt.close(fig)

    rainy_days = int((plot_df["日降雨量(mm)"] > 0).sum())
    non_rainy_days = int(len(plot_df) - rainy_days)
    total_days = max(1, rainy_days + non_rainy_days)
    pie_labels = ["", ""]
    fig, ax = plt.subplots(figsize=(4.8, 4.8), dpi=180)
    label_iter = iter(["降雨日", "非降雨日"])
    ax.pie(
        [rainy_days, non_rainy_days],
        labels=pie_labels,
        autopct=lambda pct: _pie_autopct(pct, total_days, next(label_iter)),
        pctdistance=0.58,
        startangle=90,
        colors=["#5B9BD5", "#ED7D31"],
        wedgeprops={"edgecolor": "white", "linewidth": 1.0},
        textprops={"fontsize": 10, "color": "black", "ha": "center"},
    )
    ax.axis("equal")
    fig.tight_layout()
    fig.savefig(chart_paths["rainy_ratio"], bbox_inches="tight")
    plt.close(fig)
    print(f"保存降雨分析图: {output_dir}")
    return chart_paths


def _pie_autopct(pct: float, total: int, label: str) -> str:
    count = int(round(pct * total / 100.0))
    return f"{label}\n{count}天\n{pct:.0f}%"


def run_rainfall_analysis(
    rainfall_file: Path,
    combined_xlsx: Path,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行降雨分析

    Args:
        rainfall_file: 降雨数据 CSV 文件
        combined_xlsx: 综合分析结果 xlsx 文件（输出）
        config: 可选配置参数

    Returns:
        {
            "daily_rain": pd.DataFrame,      # 日降雨量统计
            "event_rain": pd.DataFrame,       # 场次降雨统计（日级数据为空）
            "rain_data": pd.DataFrame,        # 预处理后的降雨数据
            "rain_overview": pd.DataFrame,    # 降雨概况
            "event_data_dict": dict,          # 场次降雨详细数据（供后续模块使用）
            "freq": str,                      # 数据频率 ("minute"/"hourly"/"daily")
        }
    """
    # 合并配置
    cfg = RainfallConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 加载降雨数据
    print(f"读取降雨数据: {rainfall_file}")
    rainfall_data = _load_rain_data(rainfall_file)
    rain_data = rainfall_data.df
    freq = rainfall_data.freq
    print(f"  - 时间范围: {rain_data.index.min()} ~ {rain_data.index.max()}")
    print(f"  - 总降雨量: {rain_data['rain'].sum():.2f} mm")
    print(f"  - 数据频率: {freq}")

    # 日降雨量统计
    print("计算日降雨量统计")
    daily_rain = _get_daily_rain(rain_data, freq)
    rainy_days = (daily_rain["日降雨量(mm)"] > 0).sum()
    print(f"  - 降雨日数: {rainy_days}")

    # 降雨概况
    print("计算降雨概况")
    rain_overview = _get_rain_overview(daily_rain)

    # 场次降雨划分（仅支持分钟级和小时级数据）
    event_rain = pd.DataFrame()
    event_data_dict: dict[int, dict] = {}

    if freq == "daily":
        print("日级数据不支持场次降雨划分，跳过")
    else:
        print(f"场次降雨划分: 间隔 {cfg.min_interval} 小时, 最小降雨量 {cfg.min_rainfall} mm")
        rain_rng = _time_split(rain_data, cfg.min_interval)
        event_rain = _get_rain_info(rain_rng, rain_data, cfg.min_rainfall, freq)
        print(f"  - 场次降雨数: {len(event_rain)}")

        # 添加降雨等级
        if not event_rain.empty:
            event_rain["降雨等级"] = event_rain["总降雨量(mm)"].apply(_classify_rain_level)

        # 构建场次降雨数据字典
        for _, row in event_rain.iterrows():
            event_id = int(row["场次编号"])
            start = row["开始时间"]
            end = row["结束时间"]
            event_data_dict[event_id] = {
                "start": start,
                "end": end,
                "total_rain": row["总降雨量(mm)"],
                "duration": row["降雨历时(h)"],
                "rain_level": row["降雨等级"],
            }

    # 输出到 Excel
    print("生成降雨分析结果")

    # Sheet 1: 降雨日分析（概况 + 图表）
    _add_daily_analysis_sheet(combined_xlsx, daily_rain, rain_overview)
    rainfall_chart_paths = _save_rainfall_png_charts(daily_rain, combined_xlsx.parent / "降雨分析图")

    # Sheet 2: 降雨场次分析
    if not event_rain.empty:
        _add_event_analysis_sheet(combined_xlsx, event_rain, freq)

    return {
        "daily_rain": daily_rain,
        "event_rain": event_rain,
        "rain_data": rain_data,
        "rain_overview": rain_overview,
        "rainfall_chart_paths": rainfall_chart_paths,
        "event_data_dict": event_data_dict,
        "freq": freq,
    }

