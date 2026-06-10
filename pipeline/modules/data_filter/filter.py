"""数据筛选核心逻辑

从 _archive_old_agents/agents/filter_agent.py 迁移，剥离旧架构耦合。
"""

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from pipeline.core.data_utils import read_csv_with_fallback
from pipeline.core.schema import flow_to_legacy_df, normalize_flow_df, normalize_rainfall_df, parse_flow_filename


@dataclass
class FilterConfig:
    """筛选配置参数"""
    missing_rate_threshold: float = 0.1
    expected_rows_per_day: int = 1440
    rain_day_filter_threshold: float = 2.0
    zero_like_threshold: float = 0.02
    high_zero_ratio_threshold: float = 0.5
    high_zero_ratio_normal_days_threshold: int = 5
    zero_day_drop_min_nonzero_keep_days: int = 3
    iqr_factor: float = 1.5
    mean_lower_ratio: float = 0.5
    mean_upper_ratio: float = 2.0


@dataclass
class FilterDecision:
    keep: bool
    reason: str


def _load_flow_data(csv_dir: Path) -> dict[str, pd.DataFrame]:
    """加载流量数据目录下所有 CSV"""
    result: dict[str, pd.DataFrame] = {}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        df = read_csv_with_fallback(csv_path)
        df = flow_to_legacy_df(normalize_flow_df(df, csv_path))
        df = df.rename(columns={"f": "__flow__", "l": "__level__", "velo": "__velocity__"})

        point_name = parse_flow_filename(csv_path).point_id
        result[point_name] = df.sort_values("数据时间")
    return result


def _load_rainfall_daily(rainfall_file: Path) -> pd.Series:
    """加载降雨数据，按日汇总

    如果文件不存在或无法解析，返回空 Series（仅进行旱天分析）
    """
    if not rainfall_file.exists():
        return pd.Series(dtype="float64")

    try:
        rain = read_csv_with_fallback(rainfall_file)
        rain = normalize_rainfall_df(rain)
        rain["date"] = rain["timestamp"].dt.date
        return rain.groupby("date")["rain_mm"].sum()
    except Exception:
        return pd.Series(dtype="float64")


def _calculate_daily_flow(day_groups: dict[date, pd.DataFrame], flow_col: str) -> pd.Series:
    """计算日流量 (m³/d)"""
    day_flow: dict[date, int] = {}
    for date_key, day_df in day_groups.items():
        mean_flow = pd.to_numeric(day_df[flow_col], errors="coerce").fillna(0.0).mean()
        day_flow[date_key] = int(round(float(mean_flow) * 86.4))
    return pd.Series(day_flow, dtype="float64")


def _is_high_zero_ratio_day(series: pd.Series, zero_like_threshold: float, high_zero_ratio_threshold: float) -> bool:
    """检测高零值比例日"""
    s = pd.to_numeric(series, errors="coerce").fillna(0.0).astype(float)
    if s.empty:
        return False
    zero_like = s <= zero_like_threshold
    zero_ratio = float(zero_like.mean())
    has_water = bool((s > zero_like_threshold).any())
    return has_water and zero_ratio > high_zero_ratio_threshold


def _calculate_center_ratio_base(day_flow: pd.Series) -> float:
    """计算中位数基准"""
    positive = day_flow[day_flow > 0]
    base = positive if len(positive) > 0 else day_flow
    if len(base) == 0:
        return 0.0
    return float(base.median())


class DataFilter:
    """数据筛选器"""

    def __init__(self, config: FilterConfig):
        self.config = config

    def _evaluate_day(
        self,
        flow_df: pd.DataFrame,
        flow_col: str,
        day: pd.Timestamp,
        day_df: pd.DataFrame,
        monitor_start: pd.Timestamp,
        monitor_end: pd.Timestamp,
        rain_daily: pd.Series,
        apply_zero_ratio_rule: bool = True,
        apply_missing_rule: bool = True,
        apply_rain_rule: bool = True,
    ) -> FilterDecision:
        date_key = day.date()
        if date_key == monitor_start.date():
            return FilterDecision(False, "监测周期第一天剔除")
        if date_key == monitor_end.date():
            return FilterDecision(False, "监测周期最后一天剔除")

        if apply_rain_rule:
            rain = float(rain_daily.get(date_key, 0.0))
            if rain >= self.config.rain_day_filter_threshold:
                return FilterDecision(False, f"雨天剔除(雨量={rain:.2f})")

        flow = day_df[flow_col].fillna(0.0)
        if apply_zero_ratio_rule and _is_high_zero_ratio_day(flow, self.config.zero_like_threshold, self.config.high_zero_ratio_threshold):
            ratio = float((pd.to_numeric(flow, errors="coerce").fillna(0.0) <= self.config.zero_like_threshold).mean())
            return FilterDecision(False, f"有水日近零值比例过高({ratio:.0%})，剔除")

        if apply_missing_rule:
            missing_rate = (self.config.expected_rows_per_day - len(day_df)) / self.config.expected_rows_per_day
            if missing_rate > self.config.missing_rate_threshold:
                return FilterDecision(False, f"缺失率超过{self.config.missing_rate_threshold:.0%}，剔除")

        return FilterDecision(True, "通过基础规则")

    def _should_apply_zero_ratio_rule(self, high_zero_days_count: int, base_candidate_days_count: int) -> bool:
        if high_zero_days_count < self.config.high_zero_ratio_normal_days_threshold:
            return True
        return high_zero_days_count < base_candidate_days_count

    def _apply_mean_ratio_rule(self, day_flow: pd.Series, candidates: set[pd.Timestamp]) -> set[pd.Timestamp]:
        if len(candidates) < 2:
            return candidates
        series = day_flow[day_flow.index.isin([d.date() for d in candidates])]
        center_v = _calculate_center_ratio_base(series)
        lo = center_v * self.config.mean_lower_ratio
        hi = center_v * self.config.mean_upper_ratio
        return {pd.Timestamp(d) for d, v in series.items() if lo <= float(v) <= hi}

    def _apply_high_zero_ratio_rule_with_fallback(
        self,
        base_candidates: set[pd.Timestamp],
        high_zero_days: set[pd.Timestamp],
    ) -> tuple[set[pd.Timestamp], set[pd.Timestamp], bool]:
        removed_high_zero_days = set(base_candidates & high_zero_days)
        keep = set(base_candidates) - removed_high_zero_days
        restored = False
        if len(keep) < self.config.zero_day_drop_min_nonzero_keep_days:
            keep = set(base_candidates)
            restored = True
        return keep, removed_high_zero_days, restored

    def _screen_point(
        self,
        point_name: str,
        flow_df: pd.DataFrame,
        flow_col: str,
        rain_daily: pd.Series,
    ) -> tuple[set[pd.Timestamp], dict[pd.Timestamp, float], list[str]]:
        """筛选单个点位的有效旱天"""
        df = flow_df.copy()
        df["date"] = df["数据时间"].dt.date
        monitor_start = df["数据时间"].min()
        monitor_end = df["数据时间"].max()
        day_groups = dict(tuple(df.groupby("date")))
        day_flow = _calculate_daily_flow(day_groups=day_groups, flow_col=flow_col)

        base_candidate_days = {
            pd.Timestamp(date_key)
            for date_key, day_df in day_groups.items()
            if self._evaluate_day(
                flow_df=flow_df,
                flow_col=flow_col,
                day=pd.Timestamp(date_key),
                day_df=day_df,
                monitor_start=monitor_start,
                monitor_end=monitor_end,
                rain_daily=rain_daily,
                apply_zero_ratio_rule=False,
                apply_missing_rule=True,
                apply_rain_rule=True,
            ).keep
        }

        high_zero_days = {
            pd.Timestamp(date_key)
            for date_key, day_df in day_groups.items()
            if _is_high_zero_ratio_day(day_df[flow_col], self.config.zero_like_threshold, self.config.high_zero_ratio_threshold)
        }

        apply_zero_ratio_rule = self._should_apply_zero_ratio_rule(
            high_zero_days_count=len(high_zero_days & base_candidate_days),
            base_candidate_days_count=len(base_candidate_days),
        )

        notes: list[str] = []
        if not apply_zero_ratio_rule and (high_zero_days & base_candidate_days):
            notes.append(f"该点位共有{len(high_zero_days & base_candidate_days)}天命中高零值比例，按站点常态处理，不据此剔除")

        point_candidates_strict = set(base_candidate_days)
        removed_high_zero_days: set[pd.Timestamp] = set()
        restored_high_zero_days = False

        if apply_zero_ratio_rule:
            point_candidates_strict, removed_high_zero_days, restored_high_zero_days = self._apply_high_zero_ratio_rule_with_fallback(
                base_candidates=base_candidate_days,
                high_zero_days=high_zero_days,
            )

        if removed_high_zero_days and not restored_high_zero_days:
            for d in sorted(removed_high_zero_days):
                notes.append(f"{d.strftime('%m-%d')}:有水日近零值比例过高剔除")
        elif removed_high_zero_days and restored_high_zero_days:
            notes.append(f"高零值比例剔除后保留天不足{self.config.zero_day_drop_min_nonzero_keep_days}天，恢复被该规则剔除的日期后再进行中位数筛选")

        keep_after_mean = self._apply_mean_ratio_rule(day_flow=day_flow, candidates=point_candidates_strict)
        removed_by_mean = point_candidates_strict - keep_after_mean

        for d in sorted(removed_by_mean):
            notes.append(f"{d.strftime('%m-%d')}:均值比例异常剔除")

        # 回退机制
        if len(keep_after_mean) == 0:
            # 尝试放宽规则
            candidates_no_zero_ratio = {
                pd.Timestamp(date_key)
                for date_key, day_df in day_groups.items()
                if self._evaluate_day(
                    flow_df=flow_df,
                    flow_col=flow_col,
                    day=pd.Timestamp(date_key),
                    day_df=day_df,
                    monitor_start=monitor_start,
                    monitor_end=monitor_end,
                    rain_daily=rain_daily,
                    apply_zero_ratio_rule=False,
                    apply_missing_rule=True,
                    apply_rain_rule=True,
                ).keep
            }
            if len(candidates_no_zero_ratio) > 0:
                notes.append("回退: 去除高零值比例规则")
                keep_after_mean = self._apply_mean_ratio_rule(day_flow=day_flow, candidates=candidates_no_zero_ratio)

            if len(keep_after_mean) == 0:
                candidates_boundary = {
                    pd.Timestamp(date_key)
                    for date_key, day_df in day_groups.items()
                    if self._evaluate_day(
                        flow_df=flow_df,
                        flow_col=flow_col,
                        day=pd.Timestamp(date_key),
                        day_df=day_df,
                        monitor_start=monitor_start,
                        monitor_end=monitor_end,
                        rain_daily=rain_daily,
                        apply_zero_ratio_rule=False,
                        apply_missing_rule=False,
                        apply_rain_rule=True,
                    ).keep
                }
                if len(candidates_boundary) > 0:
                    notes.append("回退: 去除均值比例规则")
                    keep_after_mean = candidates_boundary

        final_keep = set(keep_after_mean)

        matrix = {pd.Timestamp(k): float(v) for k, v in day_flow.items()}
        return final_keep, matrix, notes

    def run(
        self,
        flow_data: dict[str, pd.DataFrame],
        rain_daily: pd.Series,
    ) -> tuple[dict[str, set[pd.Timestamp]], dict[str, dict[pd.Timestamp, float]], dict[str, list[str]]]:
        """执行筛选

        Returns:
            selected: {点位编号: 有效日期集合}
            matrix: {点位编号: {日期: 日流量}}
            notes: {点位编号: 筛选说明}
        """
        selected: dict[str, set[pd.Timestamp]] = {}
        matrix: dict[str, dict[pd.Timestamp, float]] = {}
        notes: dict[str, list[str]] = {}

        for point_name, flow_df in flow_data.items():
            keep, point_matrix, point_notes = self._screen_point(
                point_name=point_name,
                flow_df=flow_df,
                flow_col="__flow__",
                rain_daily=rain_daily,
            )
            selected[point_name] = keep
            matrix[point_name] = point_matrix
            notes[point_name] = point_notes if point_notes else ["无明显异常"]

        return selected, matrix, notes


def _write_filter_excel(
    output_file: Path,
    matrix: dict[str, dict[pd.Timestamp, float]],
    selected: dict[str, set[pd.Timestamp]],
    rain_daily: pd.Series,
    notes: dict[str, list[str]],
) -> None:
    """写入筛选结果 Excel"""
    days_sorted = sorted({d for row in matrix.values() for d in row.keys()})
    date_labels = [d.strftime("%Y-%m-%d") for d in days_sorted]

    frame = pd.DataFrame(index=matrix.keys(), columns=date_labels, dtype=float)
    for point_name, day_map in matrix.items():
        for d, val in day_map.items():
            frame.loc[point_name, d.strftime("%Y-%m-%d")] = val

    output_file.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # 始终添加雨量行，无数据时留空
        if not rain_daily.empty:
            rain_values = [float(rain_daily.get(pd.Timestamp(c).date(), 0.0)) for c in date_labels]
        else:
            rain_values = [np.nan] * len(date_labels)
        rain_row = pd.DataFrame(
            [rain_values],
            columns=date_labels,
            index=["当天雨量"],
        )
        combined = pd.concat([rain_row, frame], axis=0)
        combined.to_excel(writer, sheet_name="筛选结果", index_label="点位编号")

    wb = load_workbook(output_file)
    ws = wb["筛选结果"]
    _apply_formatting(ws, selected, notes)
    wb.save(output_file)


def _apply_formatting(
    ws: Any,
    selected: dict[str, set[pd.Timestamp]],
    notes: dict[str, list[str]],
) -> None:
    """应用 Excel 格式化：绿色填充有效旱天"""
    green_fill = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050")
    empty_fill = PatternFill(fill_type=None)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    point_name_to_row: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        point_name = ws.cell(row=row, column=1).value
        if point_name:
            point_name_to_row[str(point_name)] = row

    date_to_col: dict[str, int] = {}
    for col in range(2, ws.max_column + 1):
        date_str = str(ws.cell(row=1, column=col).value)
        date_to_col[date_str] = col

    for point_name, keep_days in selected.items():
        r = point_name_to_row.get(point_name)
        if not r:
            continue
        for d in keep_days:
            c = date_to_col.get(d.strftime("%Y-%m-%d"))
            if c and ws.cell(row=r, column=c).value not in (None, "--"):
                ws.cell(row=r, column=c).fill = green_fill

    # 添加筛选说明列
    note_col = ws.max_column + 1
    ws.cell(row=1, column=note_col).value = "筛选说明"
    rain_row_num = None  # 记录雨量行的行号
    for row in range(2, ws.max_row + 1):
        point_name = str(ws.cell(row=row, column=1).value or "")
        if point_name == "当天雨量":
            ws.cell(row=row, column=note_col).value = "雨量参考行"
            rain_row_num = row
            continue
        if point_name in notes:
            ws.cell(row=row, column=note_col).value = "；".join(notes[point_name])

    # 格式化所有单元格
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=note_col):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border
            if cell.value is None and cell.column < note_col:
                # 雨量行保持为空，其他行显示 "--"
                if rain_row_num and cell.row == rain_row_num:
                    cell.value = ""
                else:
                    cell.value = "--"
                cell.fill = empty_fill


def run_data_filter(
    csv_dir: Path,
    rainfall_file: Path,
    output_xlsx: Path,
    config: dict[str, Any] | None = None,
) -> dict[str, list[str]]:
    """数据筛选主入口

    Args:
        csv_dir: 流量数据 CSV 目录
        rainfall_file: 降雨数据 CSV 文件
        output_xlsx: 输出筛选结果 xlsx 路径
        config: 可选配置参数覆盖

    Returns:
        {点位编号: [有效日期 'yyyy-mm-dd', ...]}
    """
    # 合并配置
    cfg = FilterConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 加载数据
    flow_data = _load_flow_data(csv_dir)
    rain_daily = _load_rainfall_daily(rainfall_file)

    # 执行筛选
    filter_engine = DataFilter(cfg)
    selected, matrix, notes = filter_engine.run(flow_data, rain_daily)

    # 写入 Excel
    _write_filter_excel(output_xlsx, matrix, selected, rain_daily, notes)

    # 返回 dict 格式结果
    result: dict[str, list[str]] = {}
    for point_name, keep_days in selected.items():
        result[point_name] = sorted([d.strftime("%Y-%m-%d") for d in keep_days])

    return result

