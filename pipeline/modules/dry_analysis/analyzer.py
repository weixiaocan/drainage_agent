"""旱天分析核心逻辑

从 colleague_tool/code/analyze_dry_flow.py 提取核心逻辑。
"""

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from pipeline.core.data_utils import read_csv_with_fallback
from pipeline.core.schema import flow_to_legacy_df, normalize_flow_df, parse_flow_filename


@dataclass
class DryAnalysisConfig:
    """旱天分析配置参数"""
    smooth_window: int = 20  # 平滑窗口长度（分钟）
    expected_rows_per_day: int = 1440  # 每日理论数据条数


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV"""
    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as err:
            last_err = err
    if last_err:
        raise last_err
    raise RuntimeError(f"无法读取 CSV: {path}")


def _detect_columns(df: pd.DataFrame) -> tuple[str, str, str, str | None]:
    """检测 CSV 列名"""
    cols = [str(c).strip() for c in df.columns]
    time_col = "数据时间" if "数据时间" in cols else next(c for c in cols if "时间" in c)
    flow_col = "流量(L/s)(均值)" if "流量(L/s)(均值)" in cols else next(c for c in cols if "流量" in c)
    level_col = "液位(m)(均值)" if "液位(m)(均值)" in cols else next(c for c in cols if "液位" in c)
    velocity_col = None
    for c in cols:
        if "流速" in c:
            velocity_col = c
            break
    return time_col, flow_col, level_col, velocity_col


def _parse_point_name(path: Path) -> str:
    """从文件名解析点位编号，如 35891_#1.csv -> #1"""
    stem = path.stem
    if "_" in stem:
        return stem.split("_", 1)[1]
    return stem


def _load_flow_data(csv_dir: Path) -> dict[str, pd.DataFrame]:
    """加载流量数据目录下所有 CSV"""
    result: dict[str, pd.DataFrame] = {}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        df = read_csv_with_fallback(csv_path)
        df = flow_to_legacy_df(normalize_flow_df(df, csv_path))

        point_name = parse_flow_filename(csv_path).point_id
        result[point_name] = df.sort_values("数据时间").reset_index(drop=True)
    return result


def _read_filter_result(filter_file: Path) -> dict[str, list[str]]:
    """从筛选结果 xlsx 读取有效旱天列表（绿色填充单元格）"""
    import re
    wb = load_workbook(filter_file)
    ws = wb["筛选结果"]

    # 读取日期表头
    dates: list[tuple[int, str]] = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and val != "筛选说明":
            dates.append((col, str(val)[:10]))  # 取前10个字符 yyyy-mm-dd

    # 读取每个点位的有效旱天（绿色填充）
    result: dict[str, list[str]] = {}
    for row in range(3, ws.max_row + 1):  # 从第3行开始（跳过表头和雨量行）
        point_name_raw = ws.cell(row=row, column=1).value
        if not point_name_raw:
            continue
        point_name_raw = str(point_name_raw)

        # 提取点位编号（如 "35891_#1" -> "#1", "35943_13" -> "13"）
        if "_" in point_name_raw:
            point_name = point_name_raw.split("_", 1)[1]
        else:
            point_name = point_name_raw

        valid_days: list[str] = []
        for col, date_str in dates:
            cell = ws.cell(row=row, column=col)
            # 检查绿色填充
            fill = cell.fill
            if fill and fill.start_color:
                color = str(fill.start_color.index).upper()
                if color.endswith("92D050"):
                    valid_days.append(date_str)

        result[point_name] = valid_days

    return result


def _load_site_info(site_info_file: Path) -> dict[str, dict[str, float]]:
    """从点位信息 xlsx 读取管径、井深等信息"""
    df = pd.read_excel(site_info_file)

    # 检测列名
    cols = [str(c).strip() for c in df.columns]

    # 查找关键列
    point_col = None
    diameter_col = None
    depth_col = None

    for c in cols:
        if "监测点位" in c or "点位" in c:
            point_col = c
        elif "管径" in c:
            diameter_col = c
        elif "井深" in c:
            depth_col = c

    result: dict[str, dict[str, float]] = {}
    for _, row in df.iterrows():
        # 点位名称
        point_name_full = str(row[point_col]) if point_col and pd.notna(row.get(point_col)) else ""
        if not point_name_full:
            continue

        # 提取点位编号（如 "#1"）
        import re
        match = re.search(r'#\d+', point_name_full)
        if match:
            point_name = match.group()
        else:
            point_name = point_name_full

        # 管径和井深
        diameter = float(row[diameter_col]) if diameter_col and pd.notna(row.get(diameter_col)) else 0.0
        depth = float(row[depth_col]) if depth_col and pd.notna(row.get(depth_col)) else 0.0

        result[point_name] = {
            "diameter": diameter,  # 管径 (m)
            "depth": depth,        # 井深 (m)
        }

    return result


def _preprocess_flow_data(flow_data: dict[str, pd.DataFrame], expected_rows: int = 1440) -> dict[str, pd.DataFrame]:
    """预处理流量数据：缺失数据线性插值"""
    result: dict[str, pd.DataFrame] = {}

    for point_name, df in flow_data.items():
        df = df.copy()
        time_start = df["数据时间"].min()
        time_end = df["数据时间"].max()

        # 生成完整的时间序列
        full_index = pd.date_range(time_start, time_end, freq="min")
        full_df = pd.DataFrame({"数据时间": full_index})

        # 合并并插值
        df = full_df.merge(df, on="数据时间", how="left")
        for col in ["f", "l", "velo"]:
            if col in df.columns:
                df[col] = df[col].interpolate(method="linear").fillna(0.0)

        result[point_name] = df

    return result


def _get_dry_flow(flow_data: dict[str, pd.DataFrame], dry_days: dict[str, list[str]]) -> dict[str, pd.DataFrame]:
    """汇总挑选的旱天数据"""
    dry_flow: dict[str, pd.DataFrame] = {}

    for point_name, df in flow_data.items():
        days = dry_days.get(point_name, [])
        if not days:
            continue

        dfs = []
        for day in days:
            day_start = f"{day} 00:00:00"
            day_end = f"{day} 23:59:00"
            day_df = df[(df["数据时间"] >= day_start) & (df["数据时间"] <= day_end)].copy()
            if len(day_df) > 0:
                dfs.append(day_df)

        if dfs:
            dry_flow[point_name] = pd.concat(dfs, ignore_index=True)

    return dry_flow


def _get_dry_curve_data(
    flow_data: dict[str, pd.DataFrame],
    dry_days: dict[str, list[str]],
) -> tuple[dict[str, pd.DataFrame], dict[str, pd.DataFrame], dict[str, pd.DataFrame], pd.DataFrame]:
    """计算旱天特征曲线

    Returns:
        dry_curve_data: 总体特征曲线
        dry_curve_data_workday: 工作日特征曲线
        dry_curve_data_weekend: 周末特征曲线
        day_num: 工作日/周末天数统计
    """
    day_index = pd.date_range("00:00:00", "23:59:00", freq="min")
    columns = ["f", "l", "velo"]

    dry_curve_data: dict[str, pd.DataFrame] = {}
    dry_curve_data_workday: dict[str, pd.DataFrame] = {}
    dry_curve_data_weekend: dict[str, pd.DataFrame] = {}
    day_num_list: list[tuple[str, int, int]] = []

    for point_name, df in flow_data.items():
        days = dry_days.get(point_name, [])
        if not days:
            continue

        # 初始化累加数组
        day_flow_temp = np.zeros((1440, 3))
        day_flow_workday_temp = np.zeros((1440, 3))
        day_flow_weekend_temp = np.zeros((1440, 3))
        workday_num = 0
        weekend_num = 0

        for day in days:
            day_start = f"{day} 00:00:00"
            day_end = f"{day} 23:59:00"
            day_df = df[(df["数据时间"] >= day_start) & (df["数据时间"] <= day_end)]

            if len(day_df) == 0:
                continue

            # 获取当天数据（1440 个点）
            values = day_df[["f", "l", "velo"]].values if "velo" in day_df.columns else day_df[["f", "l"]].values
            if len(values) == 1440:
                day_flow_temp += values[:, :3] if values.shape[1] >= 3 else np.column_stack([values, np.zeros(1440)])

                # 判断工作日/周末
                weekday = parse(day).weekday() + 1
                if weekday in [1, 2, 3, 4, 5]:
                    workday_num += 1
                    day_flow_workday_temp += values[:, :3] if values.shape[1] >= 3 else np.column_stack([values, np.zeros(1440)])
                else:
                    weekend_num += 1
                    day_flow_weekend_temp += values[:, :3] if values.shape[1] >= 3 else np.column_stack([values, np.zeros(1440)])

        total_days = len(days)
        if total_days > 0:
            dry_curve_data[point_name] = pd.DataFrame(
                day_flow_temp / total_days,
                index=day_index,
                columns=columns[:3] if day_flow_temp.shape[1] >= 3 else columns[:2]
            )

        if workday_num > 0:
            dry_curve_data_workday[point_name] = pd.DataFrame(
                day_flow_workday_temp / workday_num,
                index=day_index,
                columns=columns
            )

        if weekend_num > 0:
            dry_curve_data_weekend[point_name] = pd.DataFrame(
                day_flow_weekend_temp / weekend_num,
                index=day_index,
                columns=columns
            )

        day_num_list.append((point_name, workday_num, weekend_num))

    day_num = pd.DataFrame(day_num_list, columns=["点位编号", "工作日天数", "周末天数"])
    day_num = day_num.set_index("点位编号")

    return dry_curve_data, dry_curve_data_workday, dry_curve_data_weekend, day_num


def _get_dry_curve_smooth_data(dry_curve_data: dict[str, pd.DataFrame], window: int) -> dict[str, pd.DataFrame]:
    """平滑处理特征曲线"""
    result: dict[str, pd.DataFrame] = {}
    for point_name, df in dry_curve_data.items():
        result[point_name] = df.rolling(window, min_periods=1, center=True).mean()
    return result


def _get_dry_flow_sta(
    dry_flow: dict[str, pd.DataFrame],
    dry_curve_data: dict[str, pd.DataFrame],
    site_info: dict[str, dict[str, float]],
) -> pd.DataFrame:
    """计算旱天统计值"""
    rows: list[dict[str, Any]] = []

    for point_name in dry_curve_data.keys():
        curve_df = dry_curve_data[point_name]
        flow_df = dry_flow.get(point_name)

        if flow_df is None:
            continue

        # 获取管径和井深
        info = site_info.get(point_name, {"diameter": 0, "depth": 0})
        diameter = info["diameter"]
        depth = info["depth"]

        row = {
            "点位编号": point_name,
            "日均流量(m³/d)": round(curve_df["f"].mean() * 86.4, 2),
            "日最大流量(L/s)": round(curve_df["f"].max(), 2),
            "日最小流量(L/s)": round(curve_df["f"].min(), 2),
            "最大液位(m)": round(flow_df["l"].max(), 2) if "l" in flow_df.columns else 0,
            "最大充满度": round(flow_df["l"].max() / diameter * 1000, 2) if diameter > 0 else 0,
            "外溢风险": round(flow_df["l"].max() / depth, 2) if depth > 0 else 0,
            "平均流速(m/s)": round(flow_df["velo"].mean(), 6) if "velo" in flow_df.columns else 0,
            "平均液位(m)": round(flow_df["l"].mean(), 2) if "l" in flow_df.columns else 0,
        }
        rows.append(row)

    return pd.DataFrame(rows)


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    # 打开或创建工作簿
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        # 删除默认 sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # 删除已有的特征曲线 sheet（不再需要输出到 Excel）
    sheets_to_remove = [name for name in wb.sheetnames if name.startswith("特征曲线_")]
    for name in sheets_to_remove:
        wb.remove(wb[name])

    # 删除已存在的 sheet
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # 创建新 sheet
    ws = wb.create_sheet(sheet_name)

    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 替换表头
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    # 格式化
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def _save_dry_curve_data_to_excel(
    dry_curve_data: dict[str, pd.DataFrame],
    excel_path: Path,
) -> None:
    """保存旱天特征曲线数据到 Excel（每个点位一个 sheet）

    这是为了让后续模块（如 pattern_analysis）可以读取特征曲线数据。
    """
    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # 删除已有的特征曲线 sheet
    sheets_to_remove = [name for name in wb.sheetnames if name.startswith("特征曲线_")]
    for name in sheets_to_remove:
        wb.remove(wb[name])

    # 为每个点位创建一个 sheet
    for point_name, curve_df in dry_curve_data.items():
        sheet_name = f"特征曲线_{point_name}"
        # 确保 sheet 名称不超过 31 个字符
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(sheet_name)

        # 写入表头
        headers = ["时间", "流量(L/s)", "液位(m)", "流速(m/s)"]
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)

        # 写入数据
        for r_idx, (time, row) in enumerate(curve_df.iterrows(), 2):
            ws.cell(row=r_idx, column=1, value=time.strftime("%H:%M"))
            ws.cell(row=r_idx, column=2, value=round(row.get("f", 0), 4))
            ws.cell(row=r_idx, column=3, value=round(row.get("l", 0), 4))
            if "velo" in row:
                ws.cell(row=r_idx, column=4, value=round(row.get("velo", 0), 6))

    wb.save(excel_path)


def run_dry_analysis(
    flow_dir: Path,
    filter_result: Path,
    combined_xlsx: Path,
    site_info: Path | None = None,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行旱天分析

    Args:
        flow_dir: 流量数据 CSV 目录
        filter_result: 筛选结果 xlsx 文件
        combined_xlsx: 综合分析结果 xlsx 文件（输出）
        site_info: 点位信息 xlsx 文件（获取管径、井深）
        config: 可选配置参数

    Returns:
        {
            "dry_curve_data": dict[str, pd.DataFrame],  # 平滑后特征曲线
            "dry_curve_data_workday": dict[str, pd.DataFrame],  # 工作日特征曲线
            "dry_curve_data_weekend": dict[str, pd.DataFrame],  # 周末特征曲线
            "statistics": pd.DataFrame,  # 统计值
            "day_num": pd.DataFrame,  # 工作日/周末天数
        }
    """
    # 合并配置
    cfg = DryAnalysisConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 加载数据
    print(f"读取流量数据: {flow_dir}")
    flow_data = _load_flow_data(flow_dir)
    print(f"  - 加载点位数: {len(flow_data)}")

    # 预处理（插值）
    print("预处理: 缺失数据线性插值")
    flow_data = _preprocess_flow_data(flow_data, cfg.expected_rows_per_day)

    # 读取筛选结果
    print(f"读取筛选结果: {filter_result}")
    dry_days = _read_filter_result(filter_result)
    total_days = sum(len(days) for days in dry_days.values())
    print(f"  - 有效旱天总数: {total_days}")

    # 读取点位信息
    site_info_dict: dict[str, dict[str, float]] = {}
    if site_info and site_info.exists():
        print(f"读取点位信息: {site_info}")
        site_info_dict = _load_site_info(site_info)

    # 汇总旱天数据
    print("汇总旱天数据")
    dry_flow = _get_dry_flow(flow_data, dry_days)

    # 计算特征曲线
    print("计算旱天特征曲线")
    dry_curve_data, dry_curve_data_workday, dry_curve_data_weekend, day_num = _get_dry_curve_data(flow_data, dry_days)

    # 平滑处理
    print(f"平滑处理: 窗口长度 {cfg.smooth_window}")
    dry_curve_data_smooth = _get_dry_curve_smooth_data(dry_curve_data, cfg.smooth_window)

    # 计算统计值
    print("计算旱天统计指标")
    statistics = _get_dry_flow_sta(dry_flow, dry_curve_data_smooth, site_info_dict)

    # 输出到综合分析结果.xlsx
    _save_to_excel(
        statistics,
        combined_xlsx,
        "旱天分析",
        ["点位编号", "日均流量(m³/d)", "日最大流量(L/s)", "日最小流量(L/s)",
         "最大液位(m)", "最大充满度", "外溢风险", "平均流速(m/s)", "平均液位(m)"]
    )
    print(f"保存旱天分析结果: {combined_xlsx}")

    # 不再保存旱天特征曲线数据到 Excel，通过内存传递给后续模块
    # 如需调试，可取消注释以下代码：
    # _save_dry_curve_data_to_excel(dry_curve_data_smooth, combined_xlsx)

    return {
        "dry_curve_data": dry_curve_data_smooth,
        "dry_curve_data_workday": dry_curve_data_workday,
        "dry_curve_data_weekend": dry_curve_data_weekend,
        "statistics": statistics,
        "day_num": day_num,
    }

