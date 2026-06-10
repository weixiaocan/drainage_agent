"""RDII分析核心逻辑

从 colleague_tool/code/analyze_event_flow.py 和 analyze_event_RDII.py 提取核心逻辑。

RDII (Rainfall-Derived Infiltration and Inflow) = 雨天流量 - 旱天特征曲线
"""

import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import gridspec
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from pipeline.core.data_utils import read_csv_with_fallback
from pipeline.core.schema import flow_to_legacy_df, normalize_flow_df, parse_flow_filename

# 配置中文字体
mpl.rcParams['font.sans-serif'] = ['SimHei']
mpl.rcParams['font.serif'] = ['SimHei']
mpl.rcParams['axes.unicode_minus'] = False


@dataclass
class RDIIConfig:
    """RDII分析配置参数"""
    rain_effect_delay: float = 12.0  # 降雨效应延迟时间（小时）


def _read_csv_with_fallback(path: Path) -> pd.DataFrame:
    """尝试多种编码读取 CSV"""
    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except Exception as err:
            last_err = err
    if last_err:
        raise last_err
    raise RuntimeError(f"无法读取 CSV: {path}")


def _detect_columns(df: pd.DataFrame) -> tuple[str, str, str, str | None]:
    """检测 CSV 列名"""
    cols = [str(c).strip() for c in df.columns]
    time_col = "数据时间" if "数据时间" in cols else next(c for c in cols if "时间" in c)
    flow_col = "流量(L/s)(均值)" if "流量(L/s)(均值)" in cols else next(c for c in cols if "流量" in c)
    level_col = "液位(m)(均值)" if "液位(m)(均值)" in cols else next(c for c in cols if "液位" in c)
    velocity_col = None
    for c in cols:
        if "流速" in c:
            velocity_col = c
            break
    return time_col, flow_col, level_col, velocity_col


def _parse_point_name(path: Path) -> str:
    """从文件名解析点位编号"""
    stem = path.stem
    if "_" in stem:
        return stem.split("_", 1)[1]
    return stem


def _load_flow_data(csv_dir: Path) -> dict[str, pd.DataFrame]:
    """加载流量数据"""
    result: dict[str, pd.DataFrame] = {}
    for csv_path in sorted(csv_dir.glob("*.csv")):
        df = read_csv_with_fallback(csv_path)
        df = flow_to_legacy_df(normalize_flow_df(df, csv_path))

        point_name = parse_flow_filename(csv_path).point_id
        result[point_name] = df.sort_values("数据时间").reset_index(drop=True)
    return result


def _get_event_flow_stats(
    flow_data: dict[str, pd.DataFrame],
    event_data: dict,
    delay_hours: float,
    selected_events: list[int] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """统计降雨事件下各点位的最大液位和平均流量

    Args:
        flow_data: 流量数据字典
        event_data: 场次降雨数据（来自 rainfall_analysis）
        delay_hours: 降雨效应延迟时间
        selected_events: 选中的场次编号列表（如果为 None，使用全部场次）

    Returns:
        (最大液位DataFrame, 平均流量DataFrame)
    """
    point_names = list(flow_data.keys())
    event_ids = sorted(event_data.keys())

    # 过滤选中的场次
    if selected_events:
        event_ids = [e for e in event_ids if e in selected_events]

    max_level_data: dict[str, list] = {"点位编号": point_names}
    avg_flow_data: dict[str, list] = {"点位编号": point_names}

    for event_id in event_ids:
        event = event_data[event_id]
        start = event["start"]
        end = event["end"] + timedelta(hours=delay_hours)

        max_levels = []
        avg_flows = []

        for point_name in point_names:
            df = flow_data[point_name]
            event_df = df[(df["数据时间"] >= start) & (df["数据时间"] <= end)]

            if len(event_df) > 0:
                max_level = event_df["l"].max()
                avg_flow = event_df["f"].mean() * 86.4  # L/s -> m³/d
            else:
                max_level = np.nan
                avg_flow = np.nan

            max_levels.append(round(max_level, 2) if not np.isnan(max_level) else np.nan)
            avg_flows.append(round(avg_flow, 2) if not np.isnan(avg_flow) else np.nan)

        max_level_data[f"场次{event_id}"] = max_levels
        avg_flow_data[f"场次{event_id}"] = avg_flows

    return pd.DataFrame(max_level_data), pd.DataFrame(avg_flow_data)


def _get_rdii_stats(
    flow_data: dict[str, pd.DataFrame],
    dry_curve_data: dict[str, pd.DataFrame],
    event_data: dict,
    delay_hours: float,
    selected_events: list[int] | None = None,
) -> tuple[pd.DataFrame, dict]:
    """计算RDII统计

    Args:
        flow_data: 流量数据字典
        dry_curve_data: 旱天特征曲线
        event_data: 场次降雨数据
        delay_hours: 降雨效应延迟时间
        selected_events: 选中的场次编号列表（如果为 None，使用全部场次）

    Returns:
        (RDII总量DataFrame, RDII曲线数据字典)
    """
    point_names = list(dry_curve_data.keys())
    event_ids = sorted(event_data.keys())

    # 过滤选中的场次
    if selected_events:
        event_ids = [e for e in event_ids if e in selected_events]

    rdii_data: dict[str, list] = {"点位编号": point_names}
    rdii_curve_all: dict[int, dict[str, pd.DataFrame]] = {}

    for event_id in event_ids:
        event = event_data[event_id]
        start = event["start"]
        end = event["end"] + timedelta(hours=delay_hours)

        # 生成时间序列
        delta_minutes = int((end - start).total_seconds() / 60) + 1

        rdii_values = []
        rdii_curve_event: dict[str, pd.DataFrame] = {}

        for point_name in point_names:
            df = flow_data.get(point_name)
            dry_curve = dry_curve_data.get(point_name)

            if df is None or dry_curve is None:
                rdii_values.append(np.nan)
                continue

            # 获取雨天流量数据
            event_df = df[(df["数据时间"] >= start) & (df["数据时间"] <= end)]

            # 如果没有数据，跳过
            if len(event_df) == 0:
                rdii_values.append(np.nan)
                continue

            # 使用实际可用的数据长度
            actual_delta = len(event_df)

            # 获取对应的旱天特征曲线
            # 从开始时间计算偏移量
            start_minute = start.hour * 60 + start.minute
            dry_flow = dry_curve["f"].values
            dry_flow_tiled = np.tile(dry_flow, int(np.ceil(actual_delta / 1440)) + 2)
            dry_flow_segment = dry_flow_tiled[start_minute:start_minute + actual_delta]

            # 计算RDII
            rain_flow = event_df["f"].values
            rdii = rain_flow - dry_flow_segment[:len(rain_flow)]

            # RDII总量 (m³)
            rdii_total = rdii[rdii > 0].sum() * 60 / 1000
            rdii_values.append(round(rdii_total, 2))

            # 保存RDII曲线数据（使用实际数据的时间索引）
            rdii_curve_event[point_name] = pd.DataFrame({
                "时间": event_df["数据时间"].values,
                "雨天流量": rain_flow,
                "旱天流量": dry_flow_segment[:len(rain_flow)],
                "RDII": rdii,
            }).set_index("时间")

        rdii_data[f"场次{event_id}"] = rdii_values
        rdii_curve_all[event_id] = rdii_curve_event

    return pd.DataFrame(rdii_data), rdii_curve_all


def _save_to_excel(data: pd.DataFrame, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
    """保存数据到 Excel 指定 sheet"""
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # 删除旧的冗余sheet（已合并到 event_stats 模块的"降雨事件统计"中）
    legacy_sheets = ["降雨事件最大液位", "降雨事件平均流量"]
    for legacy in legacy_sheets:
        if legacy in wb.sheetnames:
            wb.remove(wb[legacy])

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = header

    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = full_border

    wb.save(excel_path)


def run_rdii_analysis(
    flow_dir: Path,
    dry_curve_data: dict[str, pd.DataFrame],
    event_data: dict[int, dict],
    combined_xlsx: Path,
    selected_events: list[int] | None = None,
    config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """执行RDII分析

    Args:
        flow_dir: 流量数据目录
        dry_curve_data: 旱天特征曲线数据（从内存传入）
        event_data: 场次降雨数据（从内存传入）
        combined_xlsx: 综合分析结果 xlsx 文件（输出）
        selected_events: 选中的场次编号列表（如果为 None，使用全部场次）
        config: 可选配置参数

    Returns:
        {
            "max_level": pd.DataFrame,      # 最大液位统计
            "avg_flow": pd.DataFrame,       # 平均流量统计
            "rdii_total": pd.DataFrame,     # RDII总量统计
            "rdii_curve_data": dict,        # RDII曲线数据
        }
    """
    # 合并配置
    cfg = RDIIConfig()
    if config:
        for key, value in config.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)

    # 加载流量数据
    print(f"读取流量数据: {flow_dir}")
    flow_data = _load_flow_data(flow_dir)
    print(f"  - 点位数: {len(flow_data)}")

    print(f"使用旱天特征曲线: {len(dry_curve_data)} 个点位")
    print(f"使用场次降雨数据: {len(event_data)} 个场次")

    if selected_events:
        print(f"  - 选中场次: {selected_events}")

    # 统计降雨事件下的流量和液位
    print("统计降雨事件下的流量和液位")
    max_level_df, avg_flow_df = _get_event_flow_stats(
        flow_data, event_data, cfg.rain_effect_delay, selected_events
    )

    # 计算RDII
    print(f"计算RDII (延迟时间: {cfg.rain_effect_delay}小时)")
    rdii_total_df, rdii_curve_data = _get_rdii_stats(
        flow_data, dry_curve_data, event_data, cfg.rain_effect_delay, selected_events
    )

    # 保存统计结果到 Excel
    event_ids_used = selected_events if selected_events else sorted(event_data.keys())

    # 注意："降雨事件最大液位"和"降雨事件平均流量"已在 event_stats 模块输出，
    # 此处只输出 RDII 特有的统计结果
    _save_to_excel(
        rdii_total_df,
        combined_xlsx,
        "RDII总量统计",
        ["点位编号"] + [f"场次{e}" for e in event_ids_used]
    )
    print(f"保存RDII总量统计: {combined_xlsx}")

    return {
        "max_level": max_level_df,
        "avg_flow": avg_flow_df,
        "rdii_total": rdii_total_df,
        "rdii_curve_data": rdii_curve_data,
    }


def draw_rdii_curves(
    rdii_curve_data: dict[int, dict[str, pd.DataFrame]],
    rain_data: pd.DataFrame,
    event_data: dict[int, dict],
    output_dir: Path,
    delay_hours: float,
    selected_events: list[int] | None = None,
) -> None:
    """绘制 RDII 过程线图

    按照原 analyze_event_RDII.py 的格式绘制：
    - 上图：降雨过程线（柱状图）
    - 下图：RDII 过程线（雨天流量、旱天流量、RDII 三条线）

    Args:
        rdii_curve_data: RDII 曲线数据，结构为 {event_id: {point_name: DataFrame}}
        rain_data: 降雨数据，index 为时间，包含 'rain' 列
        event_data: 场次降雨数据
        output_dir: 输出目录（config.charts_dir）
        delay_hours: 降雨效应延迟时间（小时）
        selected_events: 选中的场次编号列表（如果为 None，使用全部场次）
    """
    event_ids = sorted(rdii_curve_data.keys())

    # 过滤选中的场次
    if selected_events:
        event_ids = [e for e in event_ids if e in selected_events]

    for event_id in event_ids:
        event = event_data.get(event_id)
        if not event:
            continue

        time_start = event["start"]
        time_end = event["end"] + timedelta(hours=delay_hours)

        # 创建场次目录
        time_name = f"{time_start.month}_{time_start.day}"
        folder_name = f"event{event_id}_{time_name}"
        event_dir = output_dir / "rdii_curve" / folder_name
        event_dir.mkdir(parents=True, exist_ok=True)

        # 获取该场次的降雨数据
        event_rain = rain_data.loc[time_start:time_end].copy()

        # 获取该场次的 RDII 数据
        event_rdii_data = rdii_curve_data[event_id]

        for point_name, rdii_df in event_rdii_data.items():
            _draw_single_rdii_curve(
                rdii_df=rdii_df,
                event_rain=event_rain,
                time_start=time_start,
                time_end=time_end,
                event_id=event_id,
                point_name=point_name,
                output_dir=event_dir,
            )


def _draw_single_rdii_curve(
    rdii_df: pd.DataFrame,
    event_rain: pd.DataFrame,
    time_start: datetime,
    time_end: datetime,
    event_id: int,
    point_name: str,
    output_dir: Path,
) -> None:
    """绘制单个点位的 RDII 过程线

    Args:
        rdii_df: RDII 数据，index 为时间，包含 '雨天流量'、'旱天流量'、'RDII' 列
        event_rain: 降雨数据，index 为时间，包含 'rain' 列
        time_start: 开始时间
        time_end: 结束时间
        event_id: 场次编号
        point_name: 点位名称
        output_dir: 输出目录
    """
    # 准备绑制数据（删除 Overflow 列如果存在）
    data_to_plot = rdii_df.copy()
    if "Overflow" in data_to_plot.columns:
        data_to_plot = data_to_plot.drop(columns=["Overflow"])

    # 创建图表
    fig = plt.figure(figsize=(10, 5))
    gs = gridspec.GridSpec(2, 1, height_ratios=[1, 3])

    ax1 = plt.subplot(gs[1])  # 下图：RDII 过程线
    ax2 = plt.subplot(gs[0])  # 上图：降雨过程线

    # 隐藏上图的 x 轴
    ax2.get_xaxis().set_visible(False)

    # 调整子图间距
    fig.subplots_adjust(hspace=0)

    # 绘制下图：RDII 过程线
    data_to_plot.plot(ax=ax1, legend=True)

    ax1.set_xlabel('时间', fontsize='large')
    ax1.set_ylabel('流量/(L/s)', fontsize='large')

    # 绘制上图：降雨过程线（柱状图）
    # 降雨数据是分钟级的，需要聚合为适当间隔绘制柱状图
    rain_to_plot = event_rain["rain"].copy()

    # 如果数据点太多，进行聚合
    if len(rain_to_plot) > 500:
        # 按10分钟聚合
        rain_to_plot = rain_to_plot.resample("10min").sum()

    rain_to_plot.plot(ax=ax2, kind='bar', width=0.8)

    # 设置上图 y 轴标签
    ax2.set_ylabel('降雨/mm', fontsize='large')

    # 设置 x 轴刻度（减少刻度数量）
    if len(rain_to_plot) > 100:
        # 只显示部分刻度
        n_ticks = min(10, len(rain_to_plot))
        step = len(rain_to_plot) // n_ticks
        ax2.set_xticks(range(0, len(rain_to_plot), step))
        ax2.set_xticklabels(
            [rain_to_plot.index[i].strftime('%m-%d %H:%M') for i in range(0, len(rain_to_plot), step)],
            rotation=45, ha='right'
        )

    # 保存图片
    plt_name = output_dir / f"{point_name}_event{event_id}.png"
    plt.savefig(plt_name, dpi=300, bbox_inches='tight')

    # 清理
    plt.cla()
    plt.clf()
    plt.close(fig)

