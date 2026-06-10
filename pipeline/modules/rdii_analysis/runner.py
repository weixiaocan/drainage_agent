"""RDII分析模块入口

统一接口: run(config: Config, logger, dry_curve_data=None, event_data=None, rain_data=None) -> dict

输出:
    - config.combined_xlsx_path 的 "RDII总量统计" Sheet
    - config.charts_dir/rdii_curve/ 下的 RDII 过程线图
    - 返回值: {max_level, avg_flow, rdii_total, rdii_curve_data}
"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from pipeline.core.config import Config
from openpyxl import load_workbook

from .analyzer import draw_rdii_curves, run_rdii_analysis


def _load_dry_curve_data_from_excel(combined_xlsx: Path, logger: logging.Logger) -> dict[str, pd.DataFrame]:
    """从 Excel 读取旱天特征曲线数据"""
    dry_curve_data: dict[str, pd.DataFrame] = {}

    try:
        wb = load_workbook(combined_xlsx, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("特征曲线_"):
                ws = wb[sheet_name]
                point_name = sheet_name.replace("特征曲线_", "")

                data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:
                        data.append(row)

                if data:
                    df = pd.DataFrame(data, columns=["时间", "流量(L/s)", "液位(m)", "流速(m/s)"])
                    df = df.dropna(subset=["时间"])
                    df["时间"] = pd.date_range("00:00:00", "23:59:00", freq="min")[:len(df)]
                    df = df.set_index("时间")
                    df = df.rename(columns={"流量(L/s)": "f", "液位(m)": "l", "流速(m/s)": "velo"})
                    dry_curve_data[point_name] = df

        wb.close()

    except Exception as e:
        logger.warning(f"读取旱天特征曲线数据失败: {e}")

    return dry_curve_data


def _load_event_data_from_excel(combined_xlsx: Path, logger: logging.Logger) -> dict[int, dict]:
    """从 Excel 读取场次降雨数据"""
    event_data: dict[int, dict] = {}

    try:
        wb = load_workbook(combined_xlsx, data_only=True)

        if "场次降雨统计" in wb.sheetnames:
            ws = wb["场次降雨统计"]

            # 读取表头
            headers = [cell.value for cell in ws[1]]

            # 读取数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue

                event_id = int(row[0])
                event_data[event_id] = {
                    "start": row[1],
                    "end": row[2],
                    "total_rain": float(row[3]) if row[3] else 0,
                    "duration": float(row[4]) if row[4] else 0,
                    "rain_level": row[11] if len(row) > 11 else "",
                }

        wb.close()

    except Exception as e:
        logger.warning(f"读取场次降雨数据失败: {e}")

    return event_data


def run(
    config: Config,
    logger: logging.Logger,
    dry_curve_data: dict[str, pd.DataFrame] | None = None,
    event_data: dict[int, dict] | None = None,
    rain_data: pd.DataFrame | None = None,
) -> dict[str, Any]:
    """
    RDII分析入口。

    输入:
        - 流量数据目录（从 config.flow_data_dir）
        - 旱天特征曲线数据（从内存传入，或从 Excel 读取）
        - 场次降雨数据（从内存传入，或从 Excel 读取）
        - 降雨数据（从内存传入，用于绑制 RDII 过程线图）
        - 选中的场次编号（从 config.selected_rainfall_events）

    输出:
        - config.combined_xlsx_path 的多个 Sheet
        - config.charts_dir/rdii_curve/ 下的 RDII 过程线图

    返回:
        {
            "max_level": pd.DataFrame,
            "avg_flow": pd.DataFrame,
            "rdii_total": pd.DataFrame,
            "rdii_curve_data": dict,
        }
    """
    flow_dir = config.flow_data_dir
    combined_xlsx = config.combined_xlsx_path
    charts_dir = config.charts_dir
    selected_events = config.selected_rainfall_events if config.selected_rainfall_events else None

    logger.info(f"开始RDII分析")
    logger.info(f"  流量数据目录: {flow_dir}")
    logger.info(f"  综合分析结果: {combined_xlsx}")

    # 如果没有传入数据，从 Excel 读取
    if dry_curve_data is None:
        logger.info("  从 Excel 读取旱天特征曲线数据...")
        dry_curve_data = _load_dry_curve_data_from_excel(combined_xlsx, logger)

    if event_data is None:
        logger.info("  从 Excel 读取场次降雨数据...")
        event_data = _load_event_data_from_excel(combined_xlsx, logger)

    if not dry_curve_data:
        logger.warning("未找到旱天特征曲线数据，跳过RDII分析")
        return {
            "max_level": pd.DataFrame(),
            "avg_flow": pd.DataFrame(),
            "rdii_total": pd.DataFrame(),
            "rdii_curve_data": {},
        }

    if not event_data:
        logger.warning("未找到场次降雨数据，跳过RDII分析")
        return {
            "max_level": pd.DataFrame(),
            "avg_flow": pd.DataFrame(),
            "rdii_total": pd.DataFrame(),
            "rdii_curve_data": {},
        }

    logger.info(f"  加载旱天特征曲线: {len(dry_curve_data)} 个点位")
    logger.info(f"  加载场次降雨数据: {len(event_data)} 个场次")

    if selected_events:
        logger.info(f"  选中场次: {selected_events}")

    # 构建配置参数
    analysis_config = {
        "rain_effect_delay": config.rainfall_delay_hours,
    }

    # 执行分析
    result = run_rdii_analysis(
        flow_dir=flow_dir,
        dry_curve_data=dry_curve_data,
        event_data=event_data,
        combined_xlsx=combined_xlsx,
        selected_events=selected_events,
        config=analysis_config,
    )

    logger.info(f"RDII分析完成")
    logger.info(f"  处理场次数: {len(result['rdii_curve_data'])}")
    logger.info(f"  处理点位数: {len(result['rdii_total'])}")

    # 绑制 RDII 过程线图
    if rain_data is not None and result['rdii_curve_data']:
        logger.info("绑制 RDII 过程线图...")
        draw_rdii_curves(
            rdii_curve_data=result['rdii_curve_data'],
            rain_data=rain_data,
            event_data=event_data,
            output_dir=charts_dir,
            delay_hours=config.rainfall_delay_hours,
            selected_events=selected_events,
        )
        logger.info(f"  图表保存至: {charts_dir / 'rdii_curve'}")

    return result

