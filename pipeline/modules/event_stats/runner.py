"""雨天事件统计模块入口

统一接口: run(config: Config, logger, event_data=None) -> dict

输出:
    - config.combined_xlsx_path 的 "雨天事件统计" Sheet
    - 返回值: {event_stats}
"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from pipeline.core.config import Config
from openpyxl import load_workbook

from .analyzer import run_event_stats


def _load_event_data_from_excel(combined_xlsx: Path, logger: logging.Logger) -> dict[int, dict]:
    """从 Excel 读取场次降雨数据"""
    event_data: dict[int, dict] = {}

    try:
        wb = load_workbook(combined_xlsx, data_only=True)

        if "场次降雨统计" in wb.sheetnames:
            ws = wb["场次降雨统计"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue

                event_id = int(row[0])
                event_data[event_id] = {
                    "start": row[1],
                    "end": row[2],
                    "total_rain": float(row[3]) if row[3] else 0,
                    "duration": float(row[4]) if row[4] else 0,
                    "rain_level": row[11] if len(row) > 11 else "",
                }

        wb.close()

    except Exception as e:
        logger.warning(f"读取场次降雨数据失败: {e}")

    return event_data


def run(
    config: Config,
    logger: logging.Logger,
    event_data: dict[int, dict] | None = None,
) -> dict[str, Any]:
    """
    雨天事件统计入口。

    输入:
        - 流量数据目录（从 config.flow_data_dir）
        - 场次降雨数据（从内存传入，或从 Excel 读取）
        - 选中的场次编号（从 config.selected_rainfall_events）

    输出:
        - config.combined_xlsx_path 的 "雨天事件统计" Sheet

    返回:
        {
            "event_stats": pd.DataFrame,
        }
    """
    flow_dir = config.flow_data_dir
    combined_xlsx = config.combined_xlsx_path
    selected_events = config.selected_rainfall_events if config.selected_rainfall_events else None

    logger.info(f"开始雨天事件统计")
    logger.info(f"  流量数据目录: {flow_dir}")
    logger.info(f"  综合分析结果: {combined_xlsx}")

    # 如果没有传入 event_data，从 Excel 读取
    if event_data is None:
        logger.info("  从 Excel 读取场次降雨数据...")
        event_data = _load_event_data_from_excel(combined_xlsx, logger)

    if not event_data:
        logger.warning("未找到场次降雨数据，跳过雨天事件统计")
        return {"event_stats": pd.DataFrame()}

    logger.info(f"  加载场次降雨数据: {len(event_data)} 个场次")

    if selected_events:
        logger.info(f"  选中场次: {selected_events}")

    # 构建配置参数
    analysis_config = {
        "rain_effect_delay": config.rainfall_delay_hours,
    }

    # 执行统计
    result = run_event_stats(
        flow_dir=flow_dir,
        event_data=event_data,
        combined_xlsx=combined_xlsx,
        selected_events=selected_events,
        config=analysis_config,
    )

    logger.info(f"雨天事件统计完成")
    logger.info(f"  处理点位数: {len(result['event_stats'])}")

    return result

